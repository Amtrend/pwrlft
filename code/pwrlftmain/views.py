from django.shortcuts import render, redirect
from .models import *
from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string
from django.core.mail import send_mail
from pwrlft.settings import EMAIL_HOST_USER
from django.contrib.auth.decorators import login_required
from django.db.models import F, Q
from django.core import serializers
# from django.contrib.gis.serializers.geojson import Serializer
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Border, Font, Alignment, Side
from operator import itemgetter


def main_page(request):
    return render(request, 'pwrlftmain/index.html')


def standards_page(request):
    return render(request, 'pwrlftmain/standards.html')


def privacy_page(request):
    return render(request, 'pwrlftmain/privacy.html')


def competitions_page(request):
    competitions = Competitions.objects.all()
    response_data = {
        'competitions': competitions,
    }
    return render(request, 'pwrlftmain/competitions.html', response_data)


def competition_entry_page(request, competition_slug):
    cur_competition = Competitions.objects.get(competition_slug=competition_slug)
    response_data = {
        'cur_competition': cur_competition,
    }
    if "competition_entry-form_btn" in request.POST:
        # print(request.POST)
        gender_st = request.POST.get('competition_entry-form_gender')
        surname_comp = request.POST.get('competition_entry-form_sur-input')
        name_comp = request.POST.get('competition_entry-form_name-input')
        patronymic_comp = request.POST.get('competition_entry-form_patr-input')
        birthday = request.POST.get('competition_entry-form_birthday-input')
        weight_cat_st = request.POST.getlist('competition_entry-form_wcat')
        sports_cat = request.POST.get('competition_entry-form_sp-cat')
        sports_type_st = request.POST.getlist('competition_entry-form_sp-t_chckbx')
        best_res_pwlft_ek_st = request.POST.get('competition_entry-form_pwlekip-input')
        best_res_pwlft_cl_st = request.POST.get('competition_entry-form_pwlclass-input')
        best_res_bpress_ek_st = request.POST.get('competition_entry-form_bpekip-input')
        best_res_bpress_cl_st = request.POST.get('competition_entry-form_bp-input')
        mail = request.POST.get('competition_entry-form_mail-input')
        phone = request.POST.get('competition_entry-form_phone-input')
        surname_trainer = request.POST.get('competition_entry-form_sur-trn-input')
        name_trainer = request.POST.get('competition_entry-form_name-trn-input')
        patronymic_trainer = request.POST.get('competition_entry-form_patr-trn-input')
        club = request.POST.get('competition_entry-form_club-input')
        best_res_pwlft_ek = None
        best_res_pwlft_cl = None
        best_res_bpress_ek = None
        best_res_bpress_cl = None
        if best_res_pwlft_ek_st:
            best_res_pwlft_ek = float(best_res_pwlft_ek_st)
        if best_res_pwlft_cl_st:
            best_res_pwlft_cl = float(best_res_pwlft_cl_st)
        if best_res_bpress_ek_st:
            best_res_bpress_ek = float(best_res_bpress_ek_st)
        if best_res_bpress_cl_st:
            best_res_bpress_cl = float(best_res_bpress_cl_st)
        if gender_st == "competition_entry-form_gender-male":
            gender = "Мужской"
            weight_cat = weight_cat_st[0]
        else:
            gender = "Женский"
            weight_cat = weight_cat_st[1]
        try:
            new_competitor = Competitors(gender=gender, name_comp=name_comp, surname_comp=surname_comp, patronymic_comp=patronymic_comp, name_trainer=name_trainer, surname_trainer=surname_trainer, patronymic_trainer=patronymic_trainer, club=club, birthday=birthday, sports_cat=sports_cat, best_res_pwlft_ek=best_res_pwlft_ek, best_res_pwlft_cl=best_res_pwlft_cl, best_res_bpress_ek=best_res_bpress_ek, best_res_bpress_cl=best_res_bpress_cl, weight_cat=weight_cat, phone=phone, mail=mail, comp_title_id=cur_competition.id)
            new_competitor.save()
            for st in sports_type_st:
                sport_type_id = SportType.objects.get(title=st).id
                new_competitor.sports_type.add(sport_type_id)
        except Exception as e:
            return JsonResponse({"error": e}, status=500)
        if mail:
            competitors_list_link = f"https://powerlifting-kirov.ru/competitions/{competition_slug}/competitors"
            if gender == "Мужской" and weight_cat in ['93', '105', '120', '120+']:
                text_message = f'Благодарим Вас за регистрацию на соревнованиях {cur_competition.title}.\nСоревнования по вашей весовой категории ({weight_cat} кг) будут проходить {cur_competition.end_date}\nВзвешивание - {cur_competition.end_date} с 9:00, выступления участников с 11:00.\nЖелаем Вам удачного выступления и новых достижений!'
                html_message = render_to_string('email/competition_entry_mail.html', {'comp_title': cur_competition.title, 'weight_cat': weight_cat, 'end_date': cur_competition.end_date, 'competitors_list_link': competitors_list_link})
            else:
                text_message = f'Благодарим Вас за регистрацию на соревнованиях {cur_competition.title}.\nСоревнования по вашей весовой категории ({weight_cat} кг) будут проходить {cur_competition.start_date}\nВзвешивание - {cur_competition.start_date} с 9:00, выступления участников с 11:00.\nЖелаем Вам удачного выступления и новых достижений!'
                html_message = render_to_string('email/competition_entry_mail.html',
                                                {'comp_title': cur_competition.title, 'weight_cat': weight_cat,
                                                 'end_date': cur_competition.start_date, 'competitors_list_link': competitors_list_link})
            try:
                send_mail(f'Регистирация на соревнованиях {cur_competition.title}', text_message, EMAIL_HOST_USER, [mail], fail_silently=True, html_message=html_message)
            except Exception as e:
                pass
        return JsonResponse({"np": f"{name_comp} {patronymic_comp}"}, status=200)
    return render(request, 'pwrlftmain/competition_entry.html', response_data)


def competitors_page(request, competition_slug):
    cur_competition = Competitions.objects.get(competition_slug=competition_slug)
    competitors = Competitors.objects.filter(comp_title_id=cur_competition.id)
    # print(cur_competition)
    # print(competitors)
    # print(competitors.filter(sports_type__title="Жим лёжа (без экип.)", gender="Женский"))
    competitiors_pwlclass_m = ""
    competitiors_pwlclass_f = ""
    competitiors_pwlekip_m = ""
    competitiors_pwlekip_f = ""
    competitiors_bpekip_m = ""
    competitiors_bpekip_f = ""
    competitiors_bp_m = ""
    competitiors_bp_f = ""
    for sp_type in cur_competition.sport_types.all():
        if sp_type.title == "Троеборье классическое":
            competitiors_pwlclass_m = competitors.filter(sports_type__title="Троеборье классическое", gender="Мужской")
            competitiors_pwlclass_f = competitors.filter(sports_type__title="Троеборье классическое", gender="Женский")
        if sp_type.title == "Троеборье (экип.)":
            competitiors_pwlekip_m = competitors.filter(sports_type__title="Троеборье (экип.)", gender="Мужской")
            competitiors_pwlekip_f = competitors.filter(sports_type__title="Троеборье (экип.)", gender="Женский")
        if sp_type.title == "Жим лёжа (экип.)":
            competitiors_bpekip_m = competitors.filter(sports_type__title="Жим лёжа (экип.)", gender="Мужской")
            competitiors_bpekip_f = competitors.filter(sports_type__title="Жим лёжа (экип.)", gender="Женский")
        if sp_type.title == "Жим лёжа (без экип.)":
            competitiors_bp_m = competitors.filter(sports_type__title="Жим лёжа (без экип.)", gender="Мужской")
            competitiors_bp_f = competitors.filter(sports_type__title="Жим лёжа (без экип.)", gender="Женский")
    response_data = {
        'cur_competition': cur_competition,
        # 'competitors': competitors,
        'competitiors_pwlclass_m': competitiors_pwlclass_m,
        'competitiors_pwlclass_f': competitiors_pwlclass_f,
        'competitiors_pwlekip_m': competitiors_pwlekip_m,
        'competitiors_pwlekip_f': competitiors_pwlekip_f,
        'competitiors_bpekip_m': competitiors_bpekip_m,
        'competitiors_bpekip_f': competitiors_bpekip_f,
        'competitiors_bp_m': competitiors_bp_m,
        'competitiors_bp_f': competitiors_bp_f,
    }
    return render(request, 'pwrlftmain/competitors.html', response_data)


def competition_protocol_page(request, competition_slug):
    cur_competition = Competitions.objects.get(competition_slug=competition_slug)
    competitors = CompetitionProtocols.objects.filter(competitor__comp_title_id=cur_competition.id)
    competitiors_pwlclass_m = ""
    competitiors_pwlclass_f = ""
    competitiors_pwlekip_m = ""
    competitiors_pwlekip_f = ""
    competitiors_bpekip_m = ""
    competitiors_bpekip_f = ""
    competitiors_bp_m = ""
    competitiors_bp_f = ""
    competitiors_pwlclass_m_wcat = ""
    competitiors_pwlclass_f_wcat = ""
    competitiors_pwlekip_m_wcat = ""
    competitiors_pwlekip_f_wcat = ""
    competitiors_bpekip_m_wcat = ""
    competitiors_bpekip_f_wcat = ""
    competitiors_bp_m_wcat = ""
    competitiors_bp_f_wcat = ""
    for sp_type in cur_competition.sport_types.all():
        if sp_type.title == "Троеборье классическое":
            competitiors_pwlclass_m = competitors.filter(competitor__sports_type__title="Троеборье классическое", competitor__gender="Мужской").order_by("-best_sum_res", "competitor_weight")
            competitiors_pwlclass_m_wcat_st = competitiors_pwlclass_m.values_list("competitor__weight_cat", flat=True)
            competitiors_pwlclass_m_wcat = []
            for comppclmwc in competitiors_pwlclass_m_wcat_st:
                if comppclmwc not in competitiors_pwlclass_m_wcat:
                    competitiors_pwlclass_m_wcat.append(comppclmwc)
            competitiors_pwlclass_f = competitors.filter(competitor__sports_type__title="Троеборье классическое", competitor__gender="Женский").order_by("-best_sum_res", "competitor_weight")
            competitiors_pwlclass_f_wcat_st = competitiors_pwlclass_f.values_list("competitor__weight_cat", flat=True)
            competitiors_pwlclass_f_wcat = []
            for comppclfwc in competitiors_pwlclass_f_wcat_st:
                if comppclfwc not in competitiors_pwlclass_f_wcat:
                    competitiors_pwlclass_f_wcat.append(comppclfwc)
        if sp_type.title == "Троеборье (экип.)":
            competitiors_pwlekip_m = competitors.filter(competitor__sports_type__title="Троеборье (экип.)", competitor__gender="Мужской").order_by("-best_sum_res_ek", "competitor_weight")
            competitiors_pwlekip_m_wcat_st = competitiors_pwlekip_m.values_list("competitor__weight_cat", flat=True)
            competitiors_pwlekip_m_wcat = []
            for comppekmwc in competitiors_pwlekip_m_wcat_st:
                if comppekmwc not in competitiors_pwlekip_m_wcat:
                    competitiors_pwlekip_m_wcat.append(comppekmwc)
            competitiors_pwlekip_f = competitors.filter(competitor__sports_type__title="Троеборье (экип.)", competitor__gender="Женский").order_by("-best_sum_res_ek", "competitor_weight")
            competitiors_pwlekip_f_wcat_st = competitiors_pwlekip_f.values_list("competitor__weight_cat", flat=True)
            competitiors_pwlekip_f_wcat = []
            for comppekfwc in competitiors_pwlekip_f_wcat_st:
                if comppekfwc not in competitiors_pwlekip_f_wcat:
                    competitiors_pwlekip_f_wcat.append(comppekfwc)
        if sp_type.title == "Жим лёжа (экип.)":
            competitiors_bpekip_m = competitors.filter(competitor__sports_type__title="Жим лёжа (экип.)", competitor__gender="Мужской").order_by("-best_bpress_res_ek", "competitor_weight")
            competitiors_bpekip_m_wcat_st = competitiors_bpekip_m.values_list("competitor__weight_cat", flat=True)
            competitiors_bpekip_m_wcat = []
            for combpekmwc in competitiors_bpekip_m_wcat_st:
                if combpekmwc not in competitiors_bpekip_m_wcat:
                    competitiors_bpekip_m_wcat.append(combpekmwc)
            competitiors_bpekip_f = competitors.filter(competitor__sports_type__title="Жим лёжа (экип.)", competitor__gender="Женский").order_by("-best_bpress_res_ek", "competitor_weight")
            competitiors_bpekip_f_wcat_st = competitiors_bpekip_f.values_list("competitor__weight_cat", flat=True)
            competitiors_bpekip_f_wcat = []
            for combpekfwc in competitiors_bpekip_f_wcat_st:
                if combpekfwc not in competitiors_bpekip_f_wcat:
                    competitiors_bpekip_f_wcat.append(combpekfwc)
        if sp_type.title == "Жим лёжа (без экип.)":
            competitiors_bp_m = competitors.filter(competitor__sports_type__title="Жим лёжа (без экип.)", competitor__gender="Мужской").order_by("-best_bpress_res", "competitor_weight")
            competitiors_bp_m_wcat_st = competitiors_bp_m.values_list("competitor__weight_cat", flat=True)
            competitiors_bp_m_wcat = []
            for combpclmwc in competitiors_bp_m_wcat_st:
                if combpclmwc not in competitiors_bp_m_wcat:
                    competitiors_bp_m_wcat.append(combpclmwc)
            competitiors_bp_f = competitors.filter(competitor__sports_type__title="Жим лёжа (без экип.)", competitor__gender="Женский").order_by("-best_bpress_res", "competitor_weight")
            competitiors_bp_f_wcat_st = competitiors_bp_f.values_list("competitor__weight_cat", flat=True)
            competitiors_bp_f_wcat = []
            for combpclfwc in competitiors_bp_f_wcat_st:
                if combpclfwc not in competitiors_bp_f_wcat:
                    competitiors_bp_f_wcat.append(combpclfwc)
    response_data = {
        'cur_competition': cur_competition,
        'competitiors_pwlclass_m': competitiors_pwlclass_m,
        'competitiors_pwlclass_f': competitiors_pwlclass_f,
        'competitiors_pwlekip_m': competitiors_pwlekip_m,
        'competitiors_pwlekip_f': competitiors_pwlekip_f,
        'competitiors_bpekip_m': competitiors_bpekip_m,
        'competitiors_bpekip_f': competitiors_bpekip_f,
        'competitiors_bp_m': competitiors_bp_m,
        'competitiors_bp_f': competitiors_bp_f,
        'competitiors_pwlclass_m_wcat': competitiors_pwlclass_m_wcat,
        'competitiors_pwlclass_f_wcat': competitiors_pwlclass_f_wcat,
        'competitiors_pwlekip_m_wcat': competitiors_pwlekip_m_wcat,
        'competitiors_pwlekip_f_wcat': competitiors_pwlekip_f_wcat,
        'competitiors_bpekip_m_wcat': competitiors_bpekip_m_wcat,
        'competitiors_bpekip_f_wcat': competitiors_bpekip_f_wcat,
        'competitiors_bp_m_wcat': competitiors_bp_m_wcat,
        'competitiors_bp_f_wcat': competitiors_bp_f_wcat,
    }
    return render(request, 'pwrlftmain/competition_protocol.html', response_data)


def news_page(request):
    return render(request, 'pwrlftmain/news.html')


@login_required
def secretary_page(request, competition_slug):
    cur_competition = Competitions.objects.get(competition_slug=competition_slug)
    competitors_st = CompetitionProtocols.objects.filter(competitor__comp_title__competition_slug=competition_slug)
    competitors_stypes_st = competitors_st.values_list('competitor__sports_type__title', flat=True)
    competitors_wcats_st = competitors_st.values_list('competitor__weight_cat', flat=True)
    competitors_stypes = []
    competitor_wcats = []
    for csts in competitors_stypes_st:
        if csts not in competitors_stypes:
            competitors_stypes.append(csts)
    for csws in competitors_wcats_st:
        if csws not in competitor_wcats:
            competitor_wcats.append(csws)
    # before sort by table title
    # if "Троеборье классическое" in competitors_stypes:
    #     competitors = competitors_st.order_by(F("first_attempt_squat_res").asc(nulls_last=True))
    # elif "Троеборье (экип.)" in competitors_stypes:
    #     competitors = competitors_st.order_by(F("first_attempt_squat_res_ek").asc(nulls_last=True))
    # elif "Жим лёжа (без экип.)" in competitors_stypes:
    #     competitors = competitors_st.order_by(F("first_attempt_bpress_res").asc(nulls_last=True))
    # else:
    #     competitors = competitors_st.order_by(F("first_attempt_bpress_res_ek").asc(nulls_last=True))
    competitors = competitors_st
    competitors_streams_st = competitors.values_list('competitor_stream', flat=True).order_by()
    competitors_streams = []
    for comstr in competitors_streams_st:
        if comstr not in competitors_streams:
            competitors_streams.append(comstr)
    selected_stream = request.GET.get('form-secretary_stream-select')
    order_by = request.GET.get('order_by')
    if not order_by:
        order_by = "competitor_weight"
    competitors_str = competitors
    if selected_stream and selected_stream != "все":
        competitors_str = CompetitionProtocols.objects.filter(competitor__comp_title__competition_slug=competition_slug, competitor_stream=selected_stream)
        competitors_stypes_st = competitors_str.values_list('competitor__sports_type__title', flat=True)
        competitors_wcats_st = competitors_str.values_list('competitor__weight_cat', flat=True)
        competitors_stypes = []
        competitor_wcats = []
        for csts in competitors_stypes_st:
            if csts not in competitors_stypes:
                competitors_stypes.append(csts)
        for csws in competitors_wcats_st:
            if csws not in competitor_wcats:
                competitor_wcats.append(csws)
    if order_by:
        if order_by == "first_squat":
            competitors_f = competitors_str.filter(competitor__sports_type__title="Троеборье классическое").annotate(
                first_squat=F('first_attempt_squat_res')).values('competitor__sports_type__title', 'pk', 'first_squat', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res',
                                                                 'first_attempt_squat_off', 'second_attempt_squat_res', 'second_attempt_squat_off', 'third_attempt_squat_res', 'third_attempt_squat_off',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off', 'first_attempt_dlift_res', 'first_attempt_dlift_off', 'second_attempt_dlift_res',
                                                                 'second_attempt_dlift_off', 'third_attempt_dlift_res', 'third_attempt_dlift_off')
            competitors_s = competitors_str.filter(competitor__sports_type__title="Троеборье (экип.)").annotate(first_squat=F('first_attempt_squat_res_ek')).values('competitor__sports_type__title', 'pk', 'first_squat', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res_ek',
                                                                 'first_attempt_squat_off_ek', 'second_attempt_squat_res_ek', 'second_attempt_squat_off_ek', 'third_attempt_squat_res_ek', 'third_attempt_squat_off_ek',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek', 'first_attempt_dlift_res_ek', 'first_attempt_dlift_off_ek', 'second_attempt_dlift_res_ek',
                                                                 'second_attempt_dlift_off_ek', 'third_attempt_dlift_res_ek', 'third_attempt_dlift_off_ek')
            competitors_f_s = list(competitors_f) + list(competitors_s)
            # competitors = sorted(competitors_f_s, key=itemgetter('first_squat'))
            competitors = sorted(competitors_f_s, key=lambda x: (x['first_squat'] is not None, x['first_squat']))
        elif order_by == "-first_squat":
            competitors_f = competitors_str.filter(
                competitor__sports_type__title="Троеборье классическое").annotate(
                first_squat=F('first_attempt_squat_res')).values('competitor__sports_type__title', 'pk', 'first_squat', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res',
                                                                 'first_attempt_squat_off', 'second_attempt_squat_res', 'second_attempt_squat_off', 'third_attempt_squat_res', 'third_attempt_squat_off',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off', 'first_attempt_dlift_res', 'first_attempt_dlift_off', 'second_attempt_dlift_res',
                                                                 'second_attempt_dlift_off', 'third_attempt_dlift_res', 'third_attempt_dlift_off')
            competitors_s = competitors_str.filter(
                competitor__sports_type__title="Троеборье (экип.)").annotate(
                first_squat=F('first_attempt_squat_res_ek')).values('competitor__sports_type__title', 'pk', 'first_squat', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res_ek',
                                                                 'first_attempt_squat_off_ek', 'second_attempt_squat_res_ek', 'second_attempt_squat_off_ek', 'third_attempt_squat_res_ek', 'third_attempt_squat_off_ek',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek', 'first_attempt_dlift_res_ek', 'first_attempt_dlift_off_ek', 'second_attempt_dlift_res_ek',
                                                                 'second_attempt_dlift_off_ek', 'third_attempt_dlift_res_ek', 'third_attempt_dlift_off_ek')
            competitors_f_s = list(competitors_f) + list(competitors_s)
            # competitors = sorted(competitors_f_s, key=itemgetter('first_squat'), reverse=True)
            competitors = sorted(competitors_f_s, key=lambda x: (x['first_squat'] is not None, x['first_squat']), reverse=True)
        if order_by == "second_squat":
            competitors_f = competitors_str.filter(competitor__sports_type__title="Троеборье классическое").annotate(
                second_squat=F('second_attempt_squat_res')).values('competitor__sports_type__title', 'pk', 'second_squat', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res',
                                                                 'first_attempt_squat_off', 'second_attempt_squat_res', 'second_attempt_squat_off', 'third_attempt_squat_res', 'third_attempt_squat_off',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off', 'first_attempt_dlift_res', 'first_attempt_dlift_off', 'second_attempt_dlift_res',
                                                                 'second_attempt_dlift_off', 'third_attempt_dlift_res', 'third_attempt_dlift_off')
            competitors_s = competitors_str.filter(competitor__sports_type__title="Троеборье (экип.)").annotate(second_squat=F('second_attempt_squat_res_ek')).values('competitor__sports_type__title', 'pk', 'second_squat', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res_ek',
                                                                 'first_attempt_squat_off_ek', 'second_attempt_squat_res_ek', 'second_attempt_squat_off_ek', 'third_attempt_squat_res_ek', 'third_attempt_squat_off_ek',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek', 'first_attempt_dlift_res_ek', 'first_attempt_dlift_off_ek', 'second_attempt_dlift_res_ek',
                                                                 'second_attempt_dlift_off_ek', 'third_attempt_dlift_res_ek', 'third_attempt_dlift_off_ek')
            competitors_f_s = list(competitors_f) + list(competitors_s)
            competitors = sorted(competitors_f_s, key=lambda x: (x['second_squat'] is not None, x['second_squat']))
        elif order_by == "-second_squat":
            competitors_f = competitors_str.filter(
                competitor__sports_type__title="Троеборье классическое").annotate(
                second_squat=F('second_attempt_squat_res')).values('competitor__sports_type__title', 'pk', 'second_squat', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res',
                                                                 'first_attempt_squat_off', 'second_attempt_squat_res', 'second_attempt_squat_off', 'third_attempt_squat_res', 'third_attempt_squat_off',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off', 'first_attempt_dlift_res', 'first_attempt_dlift_off', 'second_attempt_dlift_res',
                                                                 'second_attempt_dlift_off', 'third_attempt_dlift_res', 'third_attempt_dlift_off')
            competitors_s = competitors_str.filter(
                competitor__sports_type__title="Троеборье (экип.)").annotate(
                second_squat=F('second_attempt_squat_res_ek')).values('competitor__sports_type__title', 'pk', 'second_squat', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res_ek',
                                                                 'first_attempt_squat_off_ek', 'second_attempt_squat_res_ek', 'second_attempt_squat_off_ek', 'third_attempt_squat_res_ek', 'third_attempt_squat_off_ek',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek', 'first_attempt_dlift_res_ek', 'first_attempt_dlift_off_ek', 'second_attempt_dlift_res_ek',
                                                                 'second_attempt_dlift_off_ek', 'third_attempt_dlift_res_ek', 'third_attempt_dlift_off_ek')
            competitors_f_s = list(competitors_f) + list(competitors_s)
            competitors = sorted(competitors_f_s, key=lambda x: (x['second_squat'] is not None, x['second_squat']), reverse=True)
        if order_by == "third_squat":
            competitors_f = competitors_str.filter(competitor__sports_type__title="Троеборье классическое").annotate(
                third_squat=F('third_attempt_squat_res')).values('competitor__sports_type__title', 'pk', 'third_squat', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res',
                                                                 'first_attempt_squat_off', 'second_attempt_squat_res', 'second_attempt_squat_off', 'third_attempt_squat_res', 'third_attempt_squat_off',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off', 'first_attempt_dlift_res', 'first_attempt_dlift_off', 'second_attempt_dlift_res',
                                                                 'second_attempt_dlift_off', 'third_attempt_dlift_res', 'third_attempt_dlift_off')
            competitors_s = competitors_str.filter(competitor__sports_type__title="Троеборье (экип.)").annotate(third_squat=F('third_attempt_squat_res_ek')).values('competitor__sports_type__title', 'pk', 'third_squat', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res_ek',
                                                                 'first_attempt_squat_off_ek', 'second_attempt_squat_res_ek', 'second_attempt_squat_off_ek', 'third_attempt_squat_res_ek', 'third_attempt_squat_off_ek',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek', 'first_attempt_dlift_res_ek', 'first_attempt_dlift_off_ek', 'second_attempt_dlift_res_ek',
                                                                 'second_attempt_dlift_off_ek', 'third_attempt_dlift_res_ek', 'third_attempt_dlift_off_ek')
            competitors_f_s = list(competitors_f) + list(competitors_s)
            competitors = sorted(competitors_f_s, key=lambda x: (x['third_squat'] is not None, x['third_squat']))
        elif order_by == "-third_squat":
            competitors_f = competitors_str.filter(
                competitor__sports_type__title="Троеборье классическое").annotate(
                third_squat=F('third_attempt_squat_res')).values('competitor__sports_type__title', 'pk', 'third_squat', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res',
                                                                 'first_attempt_squat_off', 'second_attempt_squat_res', 'second_attempt_squat_off', 'third_attempt_squat_res', 'third_attempt_squat_off',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off', 'first_attempt_dlift_res', 'first_attempt_dlift_off', 'second_attempt_dlift_res',
                                                                 'second_attempt_dlift_off', 'third_attempt_dlift_res', 'third_attempt_dlift_off')
            competitors_s = competitors_str.filter(
                competitor__sports_type__title="Троеборье (экип.)").annotate(
                third_squat=F('third_attempt_squat_res_ek')).values('competitor__sports_type__title', 'pk', 'third_squat', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res_ek',
                                                                 'first_attempt_squat_off_ek', 'second_attempt_squat_res_ek', 'second_attempt_squat_off_ek', 'third_attempt_squat_res_ek', 'third_attempt_squat_off_ek',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek', 'first_attempt_dlift_res_ek', 'first_attempt_dlift_off_ek', 'second_attempt_dlift_res_ek',
                                                                 'second_attempt_dlift_off_ek', 'third_attempt_dlift_res_ek', 'third_attempt_dlift_off_ek')
            competitors_f_s = list(competitors_f) + list(competitors_s)
            competitors = sorted(competitors_f_s, key=lambda x: (x['third_squat'] is not None, x['third_squat']), reverse=True)
        if order_by == "first_bpress":
            competitors_f = competitors_str.filter(competitor__sports_type__title="Троеборье классическое").annotate(
                first_bpress=F('first_attempt_bpress_res')).values('competitor__sports_type__title', 'pk', 'first_bpress', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res',
                                                                 'first_attempt_squat_off', 'second_attempt_squat_res', 'second_attempt_squat_off', 'third_attempt_squat_res', 'third_attempt_squat_off',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off', 'first_attempt_dlift_res', 'first_attempt_dlift_off', 'second_attempt_dlift_res',
                                                                 'second_attempt_dlift_off', 'third_attempt_dlift_res', 'third_attempt_dlift_off')
            competitors_s = competitors_str.filter(competitor__sports_type__title="Троеборье (экип.)").annotate(first_bpress=F('first_attempt_bpress_res_ek')).values('competitor__sports_type__title', 'pk', 'first_bpress', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res_ek',
                                                                 'first_attempt_squat_off_ek', 'second_attempt_squat_res_ek', 'second_attempt_squat_off_ek', 'third_attempt_squat_res_ek', 'third_attempt_squat_off_ek',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek', 'first_attempt_dlift_res_ek', 'first_attempt_dlift_off_ek', 'second_attempt_dlift_res_ek',
                                                                 'second_attempt_dlift_off_ek', 'third_attempt_dlift_res_ek', 'third_attempt_dlift_off_ek')
            competitors_t = competitors_str.filter(competitor__sports_type__title="Жим лёжа (без экип.)").exclude(competitor__sports_type__title__in=["Троеборье классическое", "Троеборье (экип.)", "Жим лёжа (экип.)"]).annotate(
                first_bpress=F('first_attempt_bpress_res')).values('competitor__sports_type__title', 'pk', 'first_bpress', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off')
            competitors_f_p = competitors_str.filter(competitor__sports_type__title="Жим лёжа (экип.)").exclude(competitor__sports_type__title__in=["Троеборье классическое", "Троеборье (экип.)", "Жим лёжа (без экип.)"]).annotate(first_bpress=F('first_attempt_bpress_res_ek')).values('competitor__sports_type__title', 'pk', 'first_bpress', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek')
            competitors_f_s = list(competitors_f) + list(competitors_s) + list(competitors_t) + list(competitors_f_p)
            competitors = sorted(competitors_f_s, key=lambda x: (x['first_bpress'] is not None, x['first_bpress']))
        elif order_by == "-first_bpress":
            competitors_f = competitors_str.filter(
                competitor__sports_type__title="Троеборье классическое").annotate(
                first_bpress=F('first_attempt_bpress_res')).values('competitor__sports_type__title', 'pk', 'first_bpress', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res',
                                                                 'first_attempt_squat_off', 'second_attempt_squat_res', 'second_attempt_squat_off', 'third_attempt_squat_res', 'third_attempt_squat_off',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off', 'first_attempt_dlift_res', 'first_attempt_dlift_off', 'second_attempt_dlift_res',
                                                                 'second_attempt_dlift_off', 'third_attempt_dlift_res', 'third_attempt_dlift_off')
            competitors_s = competitors_str.filter(
                competitor__sports_type__title="Троеборье (экип.)").annotate(
                first_bpress=F('first_attempt_bpress_res_ek')).values('competitor__sports_type__title', 'pk', 'first_bpress', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res_ek',
                                                                 'first_attempt_squat_off_ek', 'second_attempt_squat_res_ek', 'second_attempt_squat_off_ek', 'third_attempt_squat_res_ek', 'third_attempt_squat_off_ek',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek', 'first_attempt_dlift_res_ek', 'first_attempt_dlift_off_ek', 'second_attempt_dlift_res_ek',
                                                                 'second_attempt_dlift_off_ek', 'third_attempt_dlift_res_ek', 'third_attempt_dlift_off_ek')
            competitors_t = competitors_str.filter(competitor__sports_type__title="Жим лёжа (без экип.)").exclude(competitor__sports_type__title__in=["Троеборье классическое", "Троеборье (экип.)", "Жим лёжа (экип.)"]).annotate(
                first_bpress=F('first_attempt_bpress_res')).values('competitor__sports_type__title', 'pk', 'first_bpress', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off')
            competitors_f_p = competitors_str.filter(competitor__sports_type__title="Жим лёжа (экип.)").exclude(competitor__sports_type__title__in=["Троеборье классическое", "Троеборье (экип.)", "Жим лёжа (без экип.)"]).annotate(
                first_bpress=F('first_attempt_bpress_res_ek')).values('competitor__sports_type__title', 'pk', 'first_bpress', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek')
            competitors_f_s = list(competitors_f) + list(competitors_s) + list(competitors_t) + list(competitors_f_p)
            competitors = sorted(competitors_f_s, key=lambda x: (x['first_bpress'] is not None, x['first_bpress']), reverse=True)
        if order_by == "second_bpress":
            competitors_f = competitors_str.filter(competitor__sports_type__title="Троеборье классическое").annotate(
                second_bpress=F('second_attempt_bpress_res')).values('competitor__sports_type__title', 'pk', 'second_bpress', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res',
                                                                 'first_attempt_squat_off', 'second_attempt_squat_res', 'second_attempt_squat_off', 'third_attempt_squat_res', 'third_attempt_squat_off',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off', 'first_attempt_dlift_res', 'first_attempt_dlift_off', 'second_attempt_dlift_res',
                                                                 'second_attempt_dlift_off', 'third_attempt_dlift_res', 'third_attempt_dlift_off')
            competitors_s = competitors_str.filter(competitor__sports_type__title="Троеборье (экип.)").annotate(second_bpress=F('second_attempt_bpress_res_ek')).values('competitor__sports_type__title', 'pk', 'second_bpress', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res_ek',
                                                                 'first_attempt_squat_off_ek', 'second_attempt_squat_res_ek', 'second_attempt_squat_off_ek', 'third_attempt_squat_res_ek', 'third_attempt_squat_off_ek',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek', 'first_attempt_dlift_res_ek', 'first_attempt_dlift_off_ek', 'second_attempt_dlift_res_ek',
                                                                 'second_attempt_dlift_off_ek', 'third_attempt_dlift_res_ek', 'third_attempt_dlift_off_ek')
            competitors_t = competitors_str.filter(competitor__sports_type__title="Жим лёжа (без экип.)").exclude(competitor__sports_type__title__in=["Троеборье классическое", "Троеборье (экип.)", "Жим лёжа (экип.)"]).annotate(
                second_bpress=F('second_attempt_bpress_res')).values('competitor__sports_type__title', 'pk', 'second_bpress', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off')
            competitors_f_p = competitors_str.filter(competitor__sports_type__title="Жим лёжа (экип.)").exclude(competitor__sports_type__title__in=["Троеборье классическое", "Троеборье (экип.)", "Жим лёжа (без экип.)"]).annotate(second_bpress=F('second_attempt_bpress_res_ek')).values('competitor__sports_type__title', 'pk', 'second_bpress', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek')
            competitors_f_s = list(competitors_f) + list(competitors_s) + list(competitors_t) + list(competitors_f_p)
            competitors = sorted(competitors_f_s, key=lambda x: (x['second_bpress'] is not None, x['second_bpress']))
        elif order_by == "-second_bpress":
            competitors_f = competitors_str.filter(
                competitor__sports_type__title="Троеборье классическое").annotate(
                second_bpress=F('second_attempt_bpress_res')).values('competitor__sports_type__title', 'pk', 'second_bpress', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res',
                                                                 'first_attempt_squat_off', 'second_attempt_squat_res', 'second_attempt_squat_off', 'third_attempt_squat_res', 'third_attempt_squat_off',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off', 'first_attempt_dlift_res', 'first_attempt_dlift_off', 'second_attempt_dlift_res',
                                                                 'second_attempt_dlift_off', 'third_attempt_dlift_res', 'third_attempt_dlift_off')
            competitors_s = competitors_str.filter(
                competitor__sports_type__title="Троеборье (экип.)").annotate(
                second_bpress=F('second_attempt_bpress_res_ek')).values('competitor__sports_type__title', 'pk', 'second_bpress', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res_ek',
                                                                 'first_attempt_squat_off_ek', 'second_attempt_squat_res_ek', 'second_attempt_squat_off_ek', 'third_attempt_squat_res_ek', 'third_attempt_squat_off_ek',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek', 'first_attempt_dlift_res_ek', 'first_attempt_dlift_off_ek', 'second_attempt_dlift_res_ek',
                                                                 'second_attempt_dlift_off_ek', 'third_attempt_dlift_res_ek', 'third_attempt_dlift_off_ek')
            competitors_t = competitors_str.filter(competitor__sports_type__title="Жим лёжа (без экип.)").exclude(competitor__sports_type__title__in=["Троеборье классическое", "Троеборье (экип.)", "Жим лёжа (экип.)"]).annotate(
                second_bpress=F('second_attempt_bpress_res')).values('competitor__sports_type__title', 'pk', 'second_bpress', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off')
            competitors_f_p = competitors_str.filter(competitor__sports_type__title="Жим лёжа (экип.)").exclude(competitor__sports_type__title__in=["Троеборье классическое", "Троеборье (экип.)", "Жим лёжа (без экип.)"]).annotate(
                second_bpress=F('second_attempt_bpress_res_ek')).values('competitor__sports_type__title', 'pk', 'second_bpress', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek')
            competitors_f_s = list(competitors_f) + list(competitors_s) + list(competitors_t) + list(competitors_f_p)
            competitors = sorted(competitors_f_s, key=lambda x: (x['second_bpress'] is not None, x['second_bpress']), reverse=True)
        if order_by == "third_bpress":
            competitors_f = competitors_str.filter(competitor__sports_type__title="Троеборье классическое").annotate(
                third_bpress=F('third_attempt_bpress_res')).values('competitor__sports_type__title', 'pk', 'third_bpress', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res',
                                                                 'first_attempt_squat_off', 'second_attempt_squat_res', 'second_attempt_squat_off', 'third_attempt_squat_res', 'third_attempt_squat_off',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off', 'first_attempt_dlift_res', 'first_attempt_dlift_off', 'second_attempt_dlift_res',
                                                                 'second_attempt_dlift_off', 'third_attempt_dlift_res', 'third_attempt_dlift_off')
            competitors_s = competitors_str.filter(competitor__sports_type__title="Троеборье (экип.)").annotate(third_bpress=F('third_attempt_bpress_res_ek')).values('competitor__sports_type__title', 'pk', 'third_bpress', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res_ek',
                                                                 'first_attempt_squat_off_ek', 'second_attempt_squat_res_ek', 'second_attempt_squat_off_ek', 'third_attempt_squat_res_ek', 'third_attempt_squat_off_ek',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek', 'first_attempt_dlift_res_ek', 'first_attempt_dlift_off_ek', 'second_attempt_dlift_res_ek',
                                                                 'second_attempt_dlift_off_ek', 'third_attempt_dlift_res_ek', 'third_attempt_dlift_off_ek')
            competitors_t = competitors_str.filter(competitor__sports_type__title="Жим лёжа (без экип.)").exclude(competitor__sports_type__title__in=["Троеборье классическое", "Троеборье (экип.)", "Жим лёжа (экип.)"]).annotate(
                third_bpress=F('third_attempt_bpress_res')).values('competitor__sports_type__title', 'pk', 'third_bpress', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off')
            competitors_f_p = competitors_str.filter(competitor__sports_type__title="Жим лёжа (экип.)").exclude(competitor__sports_type__title__in=["Троеборье классическое", "Троеборье (экип.)", "Жим лёжа (без экип.)"]).annotate(third_bpress=F('third_attempt_bpress_res_ek')).values('competitor__sports_type__title', 'pk', 'third_bpress', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek')
            competitors_f_s = list(competitors_f) + list(competitors_s) + list(competitors_t) + list(competitors_f_p)
            competitors = sorted(competitors_f_s, key=lambda x: (x['third_bpress'] is not None, x['third_bpress']))
        elif order_by == "-third_bpress":
            competitors_f = competitors_str.filter(
                competitor__sports_type__title="Троеборье классическое").annotate(
                third_bpress=F('third_attempt_bpress_res')).values('competitor__sports_type__title', 'pk', 'third_bpress', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res',
                                                                 'first_attempt_squat_off', 'second_attempt_squat_res', 'second_attempt_squat_off', 'third_attempt_squat_res', 'third_attempt_squat_off',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off', 'first_attempt_dlift_res', 'first_attempt_dlift_off', 'second_attempt_dlift_res',
                                                                 'second_attempt_dlift_off', 'third_attempt_dlift_res', 'third_attempt_dlift_off')
            competitors_s = competitors_str.filter(
                competitor__sports_type__title="Троеборье (экип.)").annotate(
                third_bpress=F('third_attempt_bpress_res_ek')).values('competitor__sports_type__title', 'pk', 'third_bpress', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res_ek',
                                                                 'first_attempt_squat_off_ek', 'second_attempt_squat_res_ek', 'second_attempt_squat_off_ek', 'third_attempt_squat_res_ek', 'third_attempt_squat_off_ek',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek', 'first_attempt_dlift_res_ek', 'first_attempt_dlift_off_ek', 'second_attempt_dlift_res_ek',
                                                                 'second_attempt_dlift_off_ek', 'third_attempt_dlift_res_ek', 'third_attempt_dlift_off_ek')
            competitors_t = competitors_str.filter(competitor__sports_type__title="Жим лёжа (без экип.)").exclude(competitor__sports_type__title__in=["Троеборье классическое", "Троеборье (экип.)", "Жим лёжа (экип.)"]).annotate(
                third_bpress=F('third_attempt_bpress_res')).values('competitor__sports_type__title', 'pk', 'third_bpress', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off')
            competitors_f_p = competitors_str.filter(competitor__sports_type__title="Жим лёжа (экип.)").exclude(competitor__sports_type__title__in=["Троеборье классическое", "Троеборье (экип.)", "Жим лёжа (без экип.)"]).annotate(third_bpress=F('third_attempt_bpress_res_ek')).values('competitor__sports_type__title', 'pk', 'third_bpress', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek')
            competitors_f_s = list(competitors_f) + list(competitors_s) + list(competitors_t) + list(competitors_f_p)
            competitors = sorted(competitors_f_s, key=lambda x: (x['third_bpress'] is not None, x['third_bpress']), reverse=True)
        if order_by == "first_dlift":
            competitors_f = competitors_str.filter(competitor__sports_type__title="Троеборье классическое").annotate(
                first_dlift=F('first_attempt_dlift_res')).values('competitor__sports_type__title', 'pk', 'first_dlift', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res',
                                                                 'first_attempt_squat_off', 'second_attempt_squat_res', 'second_attempt_squat_off', 'third_attempt_squat_res', 'third_attempt_squat_off',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off', 'first_attempt_dlift_res', 'first_attempt_dlift_off', 'second_attempt_dlift_res',
                                                                 'second_attempt_dlift_off', 'third_attempt_dlift_res', 'third_attempt_dlift_off')
            competitors_s = competitors_str.filter(competitor__sports_type__title="Троеборье (экип.)").annotate(first_dlift=F('first_attempt_dlift_res_ek')).values('competitor__sports_type__title', 'pk', 'first_dlift', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res_ek',
                                                                 'first_attempt_squat_off_ek', 'second_attempt_squat_res_ek', 'second_attempt_squat_off_ek', 'third_attempt_squat_res_ek', 'third_attempt_squat_off_ek',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek', 'first_attempt_dlift_res_ek', 'first_attempt_dlift_off_ek', 'second_attempt_dlift_res_ek',
                                                                 'second_attempt_dlift_off_ek', 'third_attempt_dlift_res_ek', 'third_attempt_dlift_off_ek')
            competitors_f_s = list(competitors_f) + list(competitors_s)
            competitors = sorted(competitors_f_s, key=lambda x: (x['first_dlift'] is not None, x['first_dlift']))
        elif order_by == "-first_dlift":
            competitors_f = competitors_str.filter(
                competitor__sports_type__title="Троеборье классическое").annotate(
                first_dlift=F('first_attempt_dlift_res')).values('competitor__sports_type__title', 'pk', 'first_dlift', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res',
                                                                 'first_attempt_squat_off', 'second_attempt_squat_res', 'second_attempt_squat_off', 'third_attempt_squat_res', 'third_attempt_squat_off',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off', 'first_attempt_dlift_res', 'first_attempt_dlift_off', 'second_attempt_dlift_res',
                                                                 'second_attempt_dlift_off', 'third_attempt_dlift_res', 'third_attempt_dlift_off')
            competitors_s = competitors_str.filter(
                competitor__sports_type__title="Троеборье (экип.)").annotate(
                first_dlift=F('first_attempt_dlift_res_ek')).values('competitor__sports_type__title', 'pk', 'first_dlift', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res_ek',
                                                                 'first_attempt_squat_off_ek', 'second_attempt_squat_res_ek', 'second_attempt_squat_off_ek', 'third_attempt_squat_res_ek', 'third_attempt_squat_off_ek',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek', 'first_attempt_dlift_res_ek', 'first_attempt_dlift_off_ek', 'second_attempt_dlift_res_ek',
                                                                 'second_attempt_dlift_off_ek', 'third_attempt_dlift_res_ek', 'third_attempt_dlift_off_ek')
            competitors_f_s = list(competitors_f) + list(competitors_s)
            competitors = sorted(competitors_f_s, key=lambda x: (x['first_dlift'] is not None, x['first_dlift']), reverse=True)
        if order_by == "second_dlift":
            competitors_f = competitors_str.filter(competitor__sports_type__title="Троеборье классическое").annotate(
                second_dlift=F('second_attempt_dlift_res')).values('competitor__sports_type__title', 'pk', 'second_dlift', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res',
                                                                 'first_attempt_squat_off', 'second_attempt_squat_res', 'second_attempt_squat_off', 'third_attempt_squat_res', 'third_attempt_squat_off',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off', 'first_attempt_dlift_res', 'first_attempt_dlift_off', 'second_attempt_dlift_res',
                                                                 'second_attempt_dlift_off', 'third_attempt_dlift_res', 'third_attempt_dlift_off')
            competitors_s = competitors_str.filter(competitor__sports_type__title="Троеборье (экип.)").annotate(second_dlift=F('second_attempt_dlift_res_ek')).values('competitor__sports_type__title', 'pk', 'second_dlift', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res_ek',
                                                                 'first_attempt_squat_off_ek', 'second_attempt_squat_res_ek', 'second_attempt_squat_off_ek', 'third_attempt_squat_res_ek', 'third_attempt_squat_off_ek',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek', 'first_attempt_dlift_res_ek', 'first_attempt_dlift_off_ek', 'second_attempt_dlift_res_ek',
                                                                 'second_attempt_dlift_off_ek', 'third_attempt_dlift_res_ek', 'third_attempt_dlift_off_ek')
            competitors_f_s = list(competitors_f) + list(competitors_s)
            competitors = sorted(competitors_f_s, key=lambda x: (x['second_dlift'] is not None, x['second_dlift']))
        elif order_by == "-second_dlift":
            competitors_f = competitors_str.filter(
                competitor__sports_type__title="Троеборье классическое").annotate(
                second_dlift=F('second_attempt_dlift_res')).values('competitor__sports_type__title', 'pk', 'second_dlift', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res',
                                                                 'first_attempt_squat_off', 'second_attempt_squat_res', 'second_attempt_squat_off', 'third_attempt_squat_res', 'third_attempt_squat_off',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off', 'first_attempt_dlift_res', 'first_attempt_dlift_off', 'second_attempt_dlift_res',
                                                                 'second_attempt_dlift_off', 'third_attempt_dlift_res', 'third_attempt_dlift_off')
            competitors_s = competitors_str.filter(
                competitor__sports_type__title="Троеборье (экип.)").annotate(
                second_dlift=F('second_attempt_dlift_res_ek')).values('competitor__sports_type__title', 'pk', 'second_dlift', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res_ek',
                                                                 'first_attempt_squat_off_ek', 'second_attempt_squat_res_ek', 'second_attempt_squat_off_ek', 'third_attempt_squat_res_ek', 'third_attempt_squat_off_ek',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek', 'first_attempt_dlift_res_ek', 'first_attempt_dlift_off_ek', 'second_attempt_dlift_res_ek',
                                                                 'second_attempt_dlift_off_ek', 'third_attempt_dlift_res_ek', 'third_attempt_dlift_off_ek')
            competitors_f_s = list(competitors_f) + list(competitors_s)
            competitors = sorted(competitors_f_s, key=lambda x: (x['second_dlift'] is not None, x['second_dlift']), reverse=True)
        if order_by == "third_dlift":
            competitors_f = competitors_str.filter(competitor__sports_type__title="Троеборье классическое").annotate(
                third_dlift=F('third_attempt_dlift_res')).values('competitor__sports_type__title', 'pk', 'third_dlift', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res',
                                                                 'first_attempt_squat_off', 'second_attempt_squat_res', 'second_attempt_squat_off', 'third_attempt_squat_res', 'third_attempt_squat_off',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off', 'first_attempt_dlift_res', 'first_attempt_dlift_off', 'second_attempt_dlift_res',
                                                                 'second_attempt_dlift_off', 'third_attempt_dlift_res', 'third_attempt_dlift_off')
            competitors_s = competitors_str.filter(competitor__sports_type__title="Троеборье (экип.)").annotate(third_dlift=F('third_attempt_dlift_res_ek')).values('competitor__sports_type__title', 'pk', 'third_dlift', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res_ek',
                                                                 'first_attempt_squat_off_ek', 'second_attempt_squat_res_ek', 'second_attempt_squat_off_ek', 'third_attempt_squat_res_ek', 'third_attempt_squat_off_ek',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek', 'first_attempt_dlift_res_ek', 'first_attempt_dlift_off_ek', 'second_attempt_dlift_res_ek',
                                                                 'second_attempt_dlift_off_ek', 'third_attempt_dlift_res_ek', 'third_attempt_dlift_off_ek')
            competitors_f_s = list(competitors_f) + list(competitors_s)
            competitors = sorted(competitors_f_s, key=lambda x: (x['third_dlift'] is not None, x['third_dlift']))
        elif order_by == "-third_dlift":
            competitors_f = competitors_str.filter(
                competitor__sports_type__title="Троеборье классическое").annotate(
                third_dlift=F('third_attempt_dlift_res')).values('competitor__sports_type__title', 'pk', 'third_dlift', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res',
                                                                 'first_attempt_squat_off', 'second_attempt_squat_res', 'second_attempt_squat_off', 'third_attempt_squat_res', 'third_attempt_squat_off',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off', 'first_attempt_dlift_res', 'first_attempt_dlift_off', 'second_attempt_dlift_res',
                                                                 'second_attempt_dlift_off', 'third_attempt_dlift_res', 'third_attempt_dlift_off')
            competitors_s = competitors_str.filter(
                competitor__sports_type__title="Троеборье (экип.)").annotate(
                third_dlift=F('third_attempt_dlift_res_ek')).values('competitor__sports_type__title', 'pk', 'third_dlift', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res_ek',
                                                                 'first_attempt_squat_off_ek', 'second_attempt_squat_res_ek', 'second_attempt_squat_off_ek', 'third_attempt_squat_res_ek', 'third_attempt_squat_off_ek',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek', 'first_attempt_dlift_res_ek', 'first_attempt_dlift_off_ek', 'second_attempt_dlift_res_ek',
                                                                 'second_attempt_dlift_off_ek', 'third_attempt_dlift_res_ek', 'third_attempt_dlift_off_ek')
            competitors_f_s = list(competitors_f) + list(competitors_s)
            competitors = sorted(competitors_f_s, key=lambda x: (x['third_dlift'] is not None, x['third_dlift']), reverse=True)
        if order_by == "competitor_weight":
            competitors_f = competitors_str.filter(
                competitor__sports_type__title="Троеборье классическое").values('competitor__sports_type__title', 'pk', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res',
                                                                 'first_attempt_squat_off', 'second_attempt_squat_res', 'second_attempt_squat_off', 'third_attempt_squat_res', 'third_attempt_squat_off',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off', 'first_attempt_dlift_res', 'first_attempt_dlift_off', 'second_attempt_dlift_res',
                                                                 'second_attempt_dlift_off', 'third_attempt_dlift_res', 'third_attempt_dlift_off')
            competitors_s = competitors_str.filter(competitor__sports_type__title="Троеборье (экип.)").values('competitor__sports_type__title', 'pk', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res_ek',
                                                                 'first_attempt_squat_off_ek', 'second_attempt_squat_res_ek', 'second_attempt_squat_off_ek', 'third_attempt_squat_res_ek', 'third_attempt_squat_off_ek',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek', 'first_attempt_dlift_res_ek', 'first_attempt_dlift_off_ek', 'second_attempt_dlift_res_ek',
                                                                 'second_attempt_dlift_off_ek', 'third_attempt_dlift_res_ek', 'third_attempt_dlift_off_ek')
            competitors_t = competitors_str.filter(competitor__sports_type__title="Жим лёжа (без экип.)").exclude(competitor__sports_type__title__in=["Троеборье классическое", "Троеборье (экип.)", "Жим лёжа (экип.)"]).values('competitor__sports_type__title', 'pk', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off')
            competitors_f_p = competitors_str.filter(competitor__sports_type__title="Жим лёжа (экип.)").exclude(competitor__sports_type__title__in=["Троеборье классическое", "Троеборье (экип.)", "Жим лёжа (без экип.)"]).values('competitor__sports_type__title', 'pk', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek')
            competitors_f_s = list(competitors_f) + list(competitors_s) + list(competitors_t) + list(competitors_f_p)
            competitors = sorted(competitors_f_s, key=lambda x: (x['competitor_weight'] is not None, x['competitor_weight']))
        elif order_by == "-competitor_weight":
            competitors_f = competitors_str.filter(
                competitor__sports_type__title="Троеборье классическое").values('competitor__sports_type__title', 'pk', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res',
                                                                 'first_attempt_squat_off', 'second_attempt_squat_res', 'second_attempt_squat_off', 'third_attempt_squat_res', 'third_attempt_squat_off',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off', 'first_attempt_dlift_res', 'first_attempt_dlift_off', 'second_attempt_dlift_res',
                                                                 'second_attempt_dlift_off', 'third_attempt_dlift_res', 'third_attempt_dlift_off')
            competitors_s = competitors_str.filter(competitor__sports_type__title="Троеборье (экип.)").values('competitor__sports_type__title', 'pk', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res_ek',
                                                                 'first_attempt_squat_off_ek', 'second_attempt_squat_res_ek', 'second_attempt_squat_off_ek', 'third_attempt_squat_res_ek', 'third_attempt_squat_off_ek',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek', 'first_attempt_dlift_res_ek', 'first_attempt_dlift_off_ek', 'second_attempt_dlift_res_ek',
                                                                 'second_attempt_dlift_off_ek', 'third_attempt_dlift_res_ek', 'third_attempt_dlift_off_ek')
            competitors_t = competitors_str.filter(competitor__sports_type__title="Жим лёжа (без экип.)").exclude(competitor__sports_type__title__in=["Троеборье классическое", "Троеборье (экип.)", "Жим лёжа (экип.)"]).values('competitor__sports_type__title', 'pk', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off')
            competitors_f_p = competitors_str.filter(competitor__sports_type__title="Жим лёжа (экип.)").exclude(competitor__sports_type__title__in=["Троеборье классическое", "Троеборье (экип.)", "Жим лёжа (без экип.)"]).values('competitor__sports_type__title', 'pk', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek')
            competitors_f_s = list(competitors_f) + list(competitors_s) + list(competitors_t) + list(competitors_f_p)
            competitors = sorted(competitors_f_s, key=lambda x: (x['competitor_weight'] is not None, x['competitor_weight']), reverse=True)
        if order_by == "competitor__surname_comp":
            competitors_f = competitors_str.filter(
                competitor__sports_type__title="Троеборье классическое").values('competitor__sports_type__title', 'pk', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res',
                                                                 'first_attempt_squat_off', 'second_attempt_squat_res', 'second_attempt_squat_off', 'third_attempt_squat_res', 'third_attempt_squat_off',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off', 'first_attempt_dlift_res', 'first_attempt_dlift_off', 'second_attempt_dlift_res',
                                                                 'second_attempt_dlift_off', 'third_attempt_dlift_res', 'third_attempt_dlift_off')
            competitors_s = competitors_str.filter(competitor__sports_type__title="Троеборье (экип.)").values('competitor__sports_type__title', 'pk', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res_ek',
                                                                 'first_attempt_squat_off_ek', 'second_attempt_squat_res_ek', 'second_attempt_squat_off_ek', 'third_attempt_squat_res_ek', 'third_attempt_squat_off_ek',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek', 'first_attempt_dlift_res_ek', 'first_attempt_dlift_off_ek', 'second_attempt_dlift_res_ek',
                                                                 'second_attempt_dlift_off_ek', 'third_attempt_dlift_res_ek', 'third_attempt_dlift_off_ek')
            competitors_t = competitors_str.filter(competitor__sports_type__title="Жим лёжа (без экип.)").exclude(competitor__sports_type__title__in=["Троеборье классическое", "Троеборье (экип.)", "Жим лёжа (экип.)"]).values('competitor__sports_type__title', 'pk', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off')
            competitors_f_p = competitors_str.filter(competitor__sports_type__title="Жим лёжа (экип.)").exclude(competitor__sports_type__title__in=["Троеборье классическое", "Троеборье (экип.)", "Жим лёжа (без экип.)"]).values('competitor__sports_type__title', 'pk', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek')
            competitors_f_s = list(competitors_f) + list(competitors_s) + list(competitors_t) + list(competitors_f_p)
            competitors = sorted(competitors_f_s, key=lambda x: (x['competitor__surname_comp'] is not None, x['competitor__surname_comp']))
        elif order_by == "-competitor__surname_comp":
            competitors_f = competitors_str.filter(
                competitor__sports_type__title="Троеборье классическое").values('competitor__sports_type__title', 'pk', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res',
                                                                 'first_attempt_squat_off', 'second_attempt_squat_res', 'second_attempt_squat_off', 'third_attempt_squat_res', 'third_attempt_squat_off',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off', 'first_attempt_dlift_res', 'first_attempt_dlift_off', 'second_attempt_dlift_res',
                                                                 'second_attempt_dlift_off', 'third_attempt_dlift_res', 'third_attempt_dlift_off')
            competitors_s = competitors_str.filter(competitor__sports_type__title="Троеборье (экип.)").values('competitor__sports_type__title', 'pk', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight', 'pillar_height_squat', 'first_attempt_squat_res_ek',
                                                                 'first_attempt_squat_off_ek', 'second_attempt_squat_res_ek', 'second_attempt_squat_off_ek', 'third_attempt_squat_res_ek', 'third_attempt_squat_off_ek',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek', 'first_attempt_dlift_res_ek', 'first_attempt_dlift_off_ek', 'second_attempt_dlift_res_ek',
                                                                 'second_attempt_dlift_off_ek', 'third_attempt_dlift_res_ek', 'third_attempt_dlift_off_ek')
            competitors_t = competitors_str.filter(competitor__sports_type__title="Жим лёжа (без экип.)").exclude(competitor__sports_type__title__in=["Троеборье классическое", "Троеборье (экип.)", "Жим лёжа (экип.)"]).values('competitor__sports_type__title', 'pk', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res', 'first_attempt_bpress_off', 'second_attempt_bpress_res', 'second_attempt_bpress_off',
                                                                 'third_attempt_bpress_res', 'third_attempt_bpress_off')
            competitors_f_p = competitors_str.filter(competitor__sports_type__title="Жим лёжа (экип.)").exclude(competitor__sports_type__title__in=["Троеборье классическое", "Троеборье (экип.)", "Жим лёжа (без экип.)"]).values('competitor__sports_type__title', 'pk', 'competitor__surname_comp', 'competitor__name_comp',
                                                                 'competitor__patronymic_comp', 'competitor_weight',
                                                                 'pillar_height_bpress', 'first_attempt_bpress_res_ek', 'first_attempt_bpress_off_ek', 'second_attempt_bpress_res_ek', 'second_attempt_bpress_off_ek',
                                                                 'third_attempt_bpress_res_ek', 'third_attempt_bpress_off_ek')
            competitors_f_s = list(competitors_f) + list(competitors_s) + list(competitors_t) + list(competitors_f_p)
            competitors = sorted(competitors_f_s, key=lambda x: (x['competitor__surname_comp'] is not None, x['competitor__surname_comp']), reverse=True)
        # before sort by table title
        # if "Троеборье классическое" in competitors_stypes:
        #     competitors = competitors_str.order_by(F("first_attempt_squat_res").asc(nulls_last=True))
        #     if None not in competitors.filter(Q(competitor__sports_type__title="Троеборье классическое") & Q(competitor__sports_type__title="Троеборье (экип.)")).values_list('third_attempt_squat_off', flat=True):
        #         competitors = competitors_str.order_by(F("first_attempt_bpress_res").asc(nulls_last=True))
        #         if None not in competitors.filter(Q(competitor__sports_type__title="Троеборье классическое") & Q(competitor__sports_type__title="Троеборье (экип.)")).values_list('third_attempt_bpress_off', flat=True):
        #             competitors = competitors_str.order_by(F("first_attempt_dlift_res").asc(nulls_last=True))
        # elif "Троеборье (экип.)" in competitors_stypes:
        #     competitors = competitors_str.order_by(F("first_attempt_squat_res_ek").asc(nulls_last=True))
        #     if None not in competitors.filter(competitor__sports_type__title="Троеборье (экип.)").values_list('third_attempt_squat_off_ek', flat=True):
        #         competitors = competitors_str.order_by(F("first_attempt_bpress_res_ek").asc(nulls_last=True))
        #         if None not in competitors.filter(competitor__sports_type__title="Троеборье (экип.)").values_list('third_attempt_bpress_off_ek', flat=True):
        #             competitors = competitors_str.order_by(F("first_attempt_dlift_res_ek").asc(nulls_last=True))
        # elif "Жим лёжа (без экип.)" in competitors_stypes:
        #     competitors = competitors_str.order_by(F("first_attempt_bpress_res").asc(nulls_last=True))
        # else:
        #     competitors = competitors_str.order_by(F("first_attempt_bpress_res_ek").asc(nulls_last=True))
    cur_transl = CompetitionProtocols.objects.exclude(competitor_translation=None)
    if cur_transl:
        end_cur_transl = f"{cur_transl.first().pk}_{cur_transl.first().competitor_translation}"
    else:
        end_cur_transl = ''
    response_data = {
        'cur_competition': cur_competition,
        'competitors': competitors,
        'competitors_streams': competitors_streams,
        'competitors_stypes': competitors_stypes,
        'competitor_wcats': competitor_wcats,
        'cur_transl': end_cur_transl,
    }
    if 'saveval' in request.POST:
        targ_el_block_st = request.POST.get('saveattrname')
        targ_el_block_val = request.POST.get('saveattrval')
        targ_el_block = targ_el_block_st.split('secr-page_noform-')[1]
        targ_el_block_end = targ_el_block.split('_')[0]
        targ_el_pk = targ_el_block.split('_')[1]
        if '-res' in targ_el_block_end:
            if '-ek' not in targ_el_block_end:
                if 'squat' in targ_el_block_end:
                    if 'fat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.first_attempt_squat_res = targ_el_block_val
                        targ_model.save()
                    if 'sat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.second_attempt_squat_res = targ_el_block_val
                        targ_model.save()
                    if 'tat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.third_attempt_squat_res = targ_el_block_val
                        targ_model.save()
                if 'bpress' in targ_el_block_end:
                    if 'fat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.first_attempt_bpress_res = targ_el_block_val
                        targ_model.save()
                    if 'sat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.second_attempt_bpress_res = targ_el_block_val
                        targ_model.save()
                    if 'tat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.third_attempt_bpress_res = targ_el_block_val
                        targ_model.save()
                if 'dlift' in targ_el_block_end:
                    if 'fat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.first_attempt_dlift_res = targ_el_block_val
                        targ_model.save()
                    if 'sat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.second_attempt_dlift_res = targ_el_block_val
                        targ_model.save()
                    if 'tat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.third_attempt_dlift_res = targ_el_block_val
                        targ_model.save()
            else:
                if 'squat' in targ_el_block_end:
                    if 'fat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.first_attempt_squat_res_ek = targ_el_block_val
                        targ_model.save()
                    if 'sat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.second_attempt_squat_res_ek = targ_el_block_val
                        targ_model.save()
                    if 'tat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.third_attempt_squat_res_ek = targ_el_block_val
                        targ_model.save()
                if 'bpress' in targ_el_block_end:
                    if 'fat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.first_attempt_bpress_res_ek = targ_el_block_val
                        targ_model.save()
                    if 'sat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.second_attempt_bpress_res_ek = targ_el_block_val
                        targ_model.save()
                    if 'tat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.third_attempt_bpress_res_ek = targ_el_block_val
                        targ_model.save()
                if 'dlift' in targ_el_block_end:
                    if 'fat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.first_attempt_dlift_res_ek = targ_el_block_val
                        targ_model.save()
                    if 'sat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.second_attempt_dlift_res_ek = targ_el_block_val
                        targ_model.save()
                    if 'tat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.third_attempt_dlift_res_ek = targ_el_block_val
                        targ_model.save()
        if '-off' in targ_el_block_end:
            if '-ek' not in targ_el_block_end:
                if 'squat' in targ_el_block_end:
                    if 'fat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.first_attempt_squat_off = targ_el_block_val
                        targ_model.save()
                    if 'sat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.second_attempt_squat_off = targ_el_block_val
                        targ_model.save()
                    if 'tat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.third_attempt_squat_off = targ_el_block_val
                        targ_model.save()
                if 'bpress' in targ_el_block_end:
                    if 'fat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.first_attempt_bpress_off = targ_el_block_val
                        targ_model.save()
                    if 'sat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.second_attempt_bpress_off = targ_el_block_val
                        targ_model.save()
                    if 'tat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.third_attempt_bpress_off = targ_el_block_val
                        targ_model.save()
                if 'dlift' in targ_el_block_end:
                    if 'fat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.first_attempt_dlift_off = targ_el_block_val
                        targ_model.save()
                    if 'sat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.second_attempt_dlift_off = targ_el_block_val
                        targ_model.save()
                    if 'tat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.third_attempt_dlift_off = targ_el_block_val
                        targ_model.save()
            else:
                if 'squat' in targ_el_block_end:
                    if 'fat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.first_attempt_squat_off_ek = targ_el_block_val
                        targ_model.save()
                    if 'sat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.second_attempt_squat_off_ek = targ_el_block_val
                        targ_model.save()
                    if 'tat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.third_attempt_squat_off_ek = targ_el_block_val
                        targ_model.save()
                if 'bpress' in targ_el_block_end:
                    if 'fat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.first_attempt_bpress_off_ek = targ_el_block_val
                        targ_model.save()
                    if 'sat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.second_attempt_bpress_off_ek = targ_el_block_val
                        targ_model.save()
                    if 'tat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.third_attempt_bpress_off_ek = targ_el_block_val
                        targ_model.save()
                if 'dlift' in targ_el_block_end:
                    if 'fat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.first_attempt_dlift_off_ek = targ_el_block_val
                        targ_model.save()
                    if 'sat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.second_attempt_dlift_off_ek = targ_el_block_val
                        targ_model.save()
                    if 'tat' in targ_el_block_end:
                        targ_model = CompetitionProtocols.objects.get(pk=targ_el_pk)
                        targ_model.third_attempt_dlift_off_ek = targ_el_block_val
                        targ_model.save()
        if 'pil-h-squat' in targ_el_block_end:
            CompetitionProtocols.objects.filter(pk=targ_el_pk).update(pillar_height_squat=targ_el_block_val)
        if 'pil-h-bpress' in targ_el_block_end:
            CompetitionProtocols.objects.filter(pk=targ_el_pk).update(pillar_height_bpress=targ_el_block_val)
        return JsonResponse({"saved_val": targ_el_block_val}, status=200)
    if 'showmeatt' in request.POST:
        # print(request.POST)
        show_or_not = request.POST.get('showmeatt')
        targ_pk = request.POST.get('pk')
        targ_block_type = request.POST.get('typeatt')
        if show_or_not == 'yes':
            cur_transl = CompetitionProtocols.objects.exclude(competitor_translation=None)
            if cur_transl:
                return JsonResponse({"error": "Нельзя запускать две трансляции одновременно. Сначала закончите существующую."}, status=302)
            else:
                CompetitionProtocols.objects.filter(pk=targ_pk).update(competitor_translation=targ_block_type)
                return JsonResponse({"answer": "yes"}, status=200)
        else:
            CompetitionProtocols.objects.filter(pk=targ_pk).update(competitor_translation=None)
            return JsonResponse({"answer": "no"}, status=200)
    # if 'secr-page_form-btn-plclass' in request.POST:
    #     # print(request.POST)
    #     comp_pk = request.POST.get('secr-page_form-btn-plclass')
    #     pil_h_squat = request.POST.get('secr-page_form-pil_h_squat')
    #     pil_h_bpress = request.POST.get('secr-page_form-pil_h_bpress')
    #     if pil_h_squat:
    #         pil_h_squat = pil_h_squat
    #     else:
    #         pil_h_squat = None
    #     if pil_h_bpress:
    #         pil_h_bpress = pil_h_bpress
    #     else:
    #         pil_h_bpress = None
    #     first_attempt_squat_res = request.POST.get('secr-page_form-fat-squat-res')
    #     if first_attempt_squat_res:
    #         first_attempt_squat_res = Decimal(first_attempt_squat_res)
    #     else:
    #         first_attempt_squat_res = None
    #     first_attempt_squat_off = request.POST.get('secr-page_form-fat-squat-off')
    #     second_attempt_squat_res = request.POST.get('secr-page_form-sat-squat-res')
    #     if second_attempt_squat_res:
    #         second_attempt_squat_res = Decimal(second_attempt_squat_res)
    #     else:
    #         second_attempt_squat_res = None
    #     second_attempt_squat_off = request.POST.get('secr-page_form-sat-squat-off')
    #     third_attempt_squat_res = request.POST.get('secr-page_form-tat-squat-res')
    #     if third_attempt_squat_res:
    #         third_attempt_squat_res = Decimal(third_attempt_squat_res)
    #     else:
    #         third_attempt_squat_res = None
    #     third_attempt_squat_off = request.POST.get('secr-page_form-tat-squat-off')
    #     first_attempt_bpress_res = request.POST.get('secr-page_form-fat-bpress-res')
    #     if first_attempt_bpress_res:
    #         first_attempt_bpress_res = Decimal(first_attempt_bpress_res)
    #     else:
    #         first_attempt_bpress_res = None
    #     first_attempt_bpress_off = request.POST.get('secr-page_form-fat-bpress-off')
    #     second_attempt_bpress_res = request.POST.get('secr-page_form-sat-bpress-res')
    #     if second_attempt_bpress_res:
    #         second_attempt_bpress_res = Decimal(second_attempt_bpress_res)
    #     else:
    #         second_attempt_bpress_res = None
    #     second_attempt_bpress_off = request.POST.get('secr-page_form-sat-bpress-off')
    #     third_attempt_bpress_res = request.POST.get('secr-page_form-tat-bpress-res')
    #     if third_attempt_bpress_res:
    #         third_attempt_bpress_res = Decimal(third_attempt_bpress_res)
    #     else:
    #         third_attempt_bpress_res = None
    #     third_attempt_bpress_off = request.POST.get('secr-page_form-tat-bpress-off')
    #     first_attempt_dlift_res = request.POST.get('secr-page_form-fat-dlift-res')
    #     if first_attempt_dlift_res:
    #         first_attempt_dlift_res = Decimal(first_attempt_dlift_res)
    #     else:
    #         first_attempt_dlift_res = None
    #     first_attempt_dlift_off = request.POST.get('secr-page_form-fat-dlift-off')
    #     second_attempt_dlift_res = request.POST.get('secr-page_form-sat-dlift-res')
    #     if second_attempt_dlift_res:
    #         second_attempt_dlift_res = Decimal(second_attempt_dlift_res)
    #     else:
    #         second_attempt_dlift_res = None
    #     second_attempt_dlift_off = request.POST.get('secr-page_form-sat-dlift-off')
    #     third_attempt_dlift_res = request.POST.get('secr-page_form-tat-dlift-res')
    #     if third_attempt_dlift_res:
    #         third_attempt_dlift_res = Decimal(third_attempt_dlift_res)
    #     else:
    #         third_attempt_dlift_res = None
    #     third_attempt_dlift_off = request.POST.get('secr-page_form-tat-dlift-off')
    #     cur_comp = CompetitionProtocols.objects.get(pk=comp_pk)
    #     # best_squat_res = ''
    #     # best_bpress_res = ''
    #     # best_dlift_res = ''
    #     # best_sum_res = ''
    #     comp_transl = request.POST.get('secr-page_chckbx')
    #     cur_transl = CompetitionProtocols.objects.exclude(competitor_translation=None)
    #     # print(comp_transl)
    #     # print(cur_transl)
    #     if comp_transl and cur_transl:
    #         # print(comp_transl)
    #         # print(cur_transl[0].competitor_translation)
    #         if comp_transl != cur_transl[0].competitor_translation:
    #             return JsonResponse({"error": "Нельзя запускать две трансляции одновременно. Сначала закончите существующую."}, status=302)
    #     if cur_transl:
    #         cur_transl_val = cur_transl.values_list('competitor_translation', flat=True)[0]
    #         cur_transl_pk = cur_transl.values_list('id', flat=True)[0]
    #         if comp_pk == str(cur_transl_pk):
    #             if cur_transl_val == "Присед 1 класс" and first_attempt_squat_off:
    #                 # CompetitionProtocols.objects.filter(pk=cur_transl_pk).update(competitor_translation=None)
    #                 cur_comp.pillar_height_squat = pil_h_squat
    #                 cur_comp.pillar_height_bpress = pil_h_bpress
    #                 cur_comp.first_attempt_squat_res = first_attempt_squat_res
    #                 cur_comp.first_attempt_squat_off = first_attempt_squat_off
    #                 cur_comp.second_attempt_squat_res = second_attempt_squat_res
    #                 cur_comp.second_attempt_squat_off = second_attempt_squat_off
    #                 cur_comp.third_attempt_squat_res = third_attempt_squat_res
    #                 cur_comp.third_attempt_squat_off = third_attempt_squat_off
    #                 cur_comp.first_attempt_bpress_res = first_attempt_bpress_res
    #                 cur_comp.first_attempt_bpress_off = first_attempt_bpress_off
    #                 cur_comp.second_attempt_bpress_res = second_attempt_bpress_res
    #                 cur_comp.second_attempt_bpress_off = second_attempt_bpress_off
    #                 cur_comp.third_attempt_bpress_res = third_attempt_bpress_res
    #                 cur_comp.third_attempt_bpress_off = third_attempt_bpress_off
    #                 cur_comp.first_attempt_dlift_res = first_attempt_dlift_res
    #                 cur_comp.first_attempt_dlift_off = first_attempt_dlift_off
    #                 cur_comp.second_attempt_dlift_res = second_attempt_dlift_res
    #                 cur_comp.second_attempt_dlift_off = second_attempt_dlift_off
    #                 cur_comp.third_attempt_dlift_res = third_attempt_dlift_res
    #                 cur_comp.third_attempt_dlift_off = third_attempt_dlift_off
    #                 cur_comp.competitor_translation = comp_transl
    #                 cur_comp.save()
    #                 return JsonResponse({"pil_h_squat": pil_h_squat, "pil_h_bpress": pil_h_bpress, "first_attempt_squat_res": first_attempt_squat_res, "first_attempt_squat_off": first_attempt_squat_off,
    #                                      "second_attempt_squat_res": second_attempt_squat_res, "second_attempt_squat_off": second_attempt_squat_off,
    #                                      "third_attempt_squat_res": third_attempt_squat_res, "third_attempt_squat_off": third_attempt_squat_off,
    #                                      "first_attempt_bpress_res": first_attempt_bpress_res, "first_attempt_bpress_off": first_attempt_bpress_off,
    #                                      "second_attempt_bpress_res": second_attempt_bpress_res, "second_attempt_bpress_off": second_attempt_bpress_off,
    #                                      "third_attempt_bpress_res": third_attempt_bpress_res, "third_attempt_bpress_off": third_attempt_bpress_off,
    #                                      "first_attempt_dlift_res": first_attempt_dlift_res, "first_attempt_dlift_off": first_attempt_dlift_off,
    #                                      "second_attempt_dlift_res": second_attempt_dlift_res, "second_attempt_dlift_off": second_attempt_dlift_off,
    #                                      "third_attempt_dlift_res": third_attempt_dlift_res, "third_attempt_dlift_off": third_attempt_dlift_off, "comp_transl": comp_transl}, status=200)
    #             if cur_transl_val == "Присед 2 класс" and second_attempt_squat_off:
    #                 cur_comp.pillar_height_squat = pil_h_squat
    #                 cur_comp.pillar_height_bpress = pil_h_bpress
    #                 cur_comp.first_attempt_squat_res = first_attempt_squat_res
    #                 cur_comp.first_attempt_squat_off = first_attempt_squat_off
    #                 cur_comp.second_attempt_squat_res = second_attempt_squat_res
    #                 cur_comp.second_attempt_squat_off = second_attempt_squat_off
    #                 cur_comp.third_attempt_squat_res = third_attempt_squat_res
    #                 cur_comp.third_attempt_squat_off = third_attempt_squat_off
    #                 cur_comp.first_attempt_bpress_res = first_attempt_bpress_res
    #                 cur_comp.first_attempt_bpress_off = first_attempt_bpress_off
    #                 cur_comp.second_attempt_bpress_res = second_attempt_bpress_res
    #                 cur_comp.second_attempt_bpress_off = second_attempt_bpress_off
    #                 cur_comp.third_attempt_bpress_res = third_attempt_bpress_res
    #                 cur_comp.third_attempt_bpress_off = third_attempt_bpress_off
    #                 cur_comp.first_attempt_dlift_res = first_attempt_dlift_res
    #                 cur_comp.first_attempt_dlift_off = first_attempt_dlift_off
    #                 cur_comp.second_attempt_dlift_res = second_attempt_dlift_res
    #                 cur_comp.second_attempt_dlift_off = second_attempt_dlift_off
    #                 cur_comp.third_attempt_dlift_res = third_attempt_dlift_res
    #                 cur_comp.third_attempt_dlift_off = third_attempt_dlift_off
    #                 cur_comp.competitor_translation = comp_transl
    #                 cur_comp.save()
    #                 return JsonResponse({"pil_h_squat": pil_h_squat, "pil_h_bpress": pil_h_bpress, "first_attempt_squat_res": first_attempt_squat_res,
    #                                      "first_attempt_squat_off": first_attempt_squat_off,
    #                                      "second_attempt_squat_res": second_attempt_squat_res,
    #                                      "second_attempt_squat_off": second_attempt_squat_off,
    #                                      "third_attempt_squat_res": third_attempt_squat_res,
    #                                      "third_attempt_squat_off": third_attempt_squat_off,
    #                                      "first_attempt_bpress_res": first_attempt_bpress_res,
    #                                      "first_attempt_bpress_off": first_attempt_bpress_off,
    #                                      "second_attempt_bpress_res": second_attempt_bpress_res,
    #                                      "second_attempt_bpress_off": second_attempt_bpress_off,
    #                                      "third_attempt_bpress_res": third_attempt_bpress_res,
    #                                      "third_attempt_bpress_off": third_attempt_bpress_off,
    #                                      "first_attempt_dlift_res": first_attempt_dlift_res,
    #                                      "first_attempt_dlift_off": first_attempt_dlift_off,
    #                                      "second_attempt_dlift_res": second_attempt_dlift_res,
    #                                      "second_attempt_dlift_off": second_attempt_dlift_off,
    #                                      "third_attempt_dlift_res": third_attempt_dlift_res,
    #                                      "third_attempt_dlift_off": third_attempt_dlift_off,
    #                                      "comp_transl": comp_transl}, status=200)
    #             if cur_transl_val == "Присед 3 класс" and third_attempt_squat_off:
    #                 cur_comp.pillar_height_squat = pil_h_squat
    #                 cur_comp.pillar_height_bpress = pil_h_bpress
    #                 cur_comp.first_attempt_squat_res = first_attempt_squat_res
    #                 cur_comp.first_attempt_squat_off = first_attempt_squat_off
    #                 cur_comp.second_attempt_squat_res = second_attempt_squat_res
    #                 cur_comp.second_attempt_squat_off = second_attempt_squat_off
    #                 cur_comp.third_attempt_squat_res = third_attempt_squat_res
    #                 cur_comp.third_attempt_squat_off = third_attempt_squat_off
    #                 cur_comp.first_attempt_bpress_res = first_attempt_bpress_res
    #                 cur_comp.first_attempt_bpress_off = first_attempt_bpress_off
    #                 cur_comp.second_attempt_bpress_res = second_attempt_bpress_res
    #                 cur_comp.second_attempt_bpress_off = second_attempt_bpress_off
    #                 cur_comp.third_attempt_bpress_res = third_attempt_bpress_res
    #                 cur_comp.third_attempt_bpress_off = third_attempt_bpress_off
    #                 cur_comp.first_attempt_dlift_res = first_attempt_dlift_res
    #                 cur_comp.first_attempt_dlift_off = first_attempt_dlift_off
    #                 cur_comp.second_attempt_dlift_res = second_attempt_dlift_res
    #                 cur_comp.second_attempt_dlift_off = second_attempt_dlift_off
    #                 cur_comp.third_attempt_dlift_res = third_attempt_dlift_res
    #                 cur_comp.third_attempt_dlift_off = third_attempt_dlift_off
    #                 cur_comp.competitor_translation = comp_transl
    #                 cur_comp.save()
    #                 return JsonResponse({"pil_h_squat": pil_h_squat, "pil_h_bpress": pil_h_bpress, "first_attempt_squat_res": first_attempt_squat_res,
    #                                      "first_attempt_squat_off": first_attempt_squat_off,
    #                                      "second_attempt_squat_res": second_attempt_squat_res,
    #                                      "second_attempt_squat_off": second_attempt_squat_off,
    #                                      "third_attempt_squat_res": third_attempt_squat_res,
    #                                      "third_attempt_squat_off": third_attempt_squat_off,
    #                                      "first_attempt_bpress_res": first_attempt_bpress_res,
    #                                      "first_attempt_bpress_off": first_attempt_bpress_off,
    #                                      "second_attempt_bpress_res": second_attempt_bpress_res,
    #                                      "second_attempt_bpress_off": second_attempt_bpress_off,
    #                                      "third_attempt_bpress_res": third_attempt_bpress_res,
    #                                      "third_attempt_bpress_off": third_attempt_bpress_off,
    #                                      "first_attempt_dlift_res": first_attempt_dlift_res,
    #                                      "first_attempt_dlift_off": first_attempt_dlift_off,
    #                                      "second_attempt_dlift_res": second_attempt_dlift_res,
    #                                      "second_attempt_dlift_off": second_attempt_dlift_off,
    #                                      "third_attempt_dlift_res": third_attempt_dlift_res,
    #                                      "third_attempt_dlift_off": third_attempt_dlift_off,
    #                                      "comp_transl": comp_transl}, status=200)
    #             if cur_transl_val == "Жим 1 класс" and first_attempt_bpress_off:
    #                 cur_comp.pillar_height_squat = pil_h_squat
    #                 cur_comp.pillar_height_bpress = pil_h_bpress
    #                 cur_comp.first_attempt_squat_res = first_attempt_squat_res
    #                 cur_comp.first_attempt_squat_off = first_attempt_squat_off
    #                 cur_comp.second_attempt_squat_res = second_attempt_squat_res
    #                 cur_comp.second_attempt_squat_off = second_attempt_squat_off
    #                 cur_comp.third_attempt_squat_res = third_attempt_squat_res
    #                 cur_comp.third_attempt_squat_off = third_attempt_squat_off
    #                 cur_comp.first_attempt_bpress_res = first_attempt_bpress_res
    #                 cur_comp.first_attempt_bpress_off = first_attempt_bpress_off
    #                 cur_comp.second_attempt_bpress_res = second_attempt_bpress_res
    #                 cur_comp.second_attempt_bpress_off = second_attempt_bpress_off
    #                 cur_comp.third_attempt_bpress_res = third_attempt_bpress_res
    #                 cur_comp.third_attempt_bpress_off = third_attempt_bpress_off
    #                 cur_comp.first_attempt_dlift_res = first_attempt_dlift_res
    #                 cur_comp.first_attempt_dlift_off = first_attempt_dlift_off
    #                 cur_comp.second_attempt_dlift_res = second_attempt_dlift_res
    #                 cur_comp.second_attempt_dlift_off = second_attempt_dlift_off
    #                 cur_comp.third_attempt_dlift_res = third_attempt_dlift_res
    #                 cur_comp.third_attempt_dlift_off = third_attempt_dlift_off
    #                 cur_comp.competitor_translation = comp_transl
    #                 cur_comp.save()
    #                 return JsonResponse({"pil_h_squat": pil_h_squat, "pil_h_bpress": pil_h_bpress, "first_attempt_squat_res": first_attempt_squat_res,
    #                                      "first_attempt_squat_off": first_attempt_squat_off,
    #                                      "second_attempt_squat_res": second_attempt_squat_res,
    #                                      "second_attempt_squat_off": second_attempt_squat_off,
    #                                      "third_attempt_squat_res": third_attempt_squat_res,
    #                                      "third_attempt_squat_off": third_attempt_squat_off,
    #                                      "first_attempt_bpress_res": first_attempt_bpress_res,
    #                                      "first_attempt_bpress_off": first_attempt_bpress_off,
    #                                      "second_attempt_bpress_res": second_attempt_bpress_res,
    #                                      "second_attempt_bpress_off": second_attempt_bpress_off,
    #                                      "third_attempt_bpress_res": third_attempt_bpress_res,
    #                                      "third_attempt_bpress_off": third_attempt_bpress_off,
    #                                      "first_attempt_dlift_res": first_attempt_dlift_res,
    #                                      "first_attempt_dlift_off": first_attempt_dlift_off,
    #                                      "second_attempt_dlift_res": second_attempt_dlift_res,
    #                                      "second_attempt_dlift_off": second_attempt_dlift_off,
    #                                      "third_attempt_dlift_res": third_attempt_dlift_res,
    #                                      "third_attempt_dlift_off": third_attempt_dlift_off,
    #                                      "comp_transl": comp_transl}, status=200)
    #             if cur_transl_val == "Жим 2 класс" and second_attempt_bpress_off:
    #                 cur_comp.pillar_height_squat = pil_h_squat
    #                 cur_comp.pillar_height_bpress = pil_h_bpress
    #                 cur_comp.first_attempt_squat_res = first_attempt_squat_res
    #                 cur_comp.first_attempt_squat_off = first_attempt_squat_off
    #                 cur_comp.second_attempt_squat_res = second_attempt_squat_res
    #                 cur_comp.second_attempt_squat_off = second_attempt_squat_off
    #                 cur_comp.third_attempt_squat_res = third_attempt_squat_res
    #                 cur_comp.third_attempt_squat_off = third_attempt_squat_off
    #                 cur_comp.first_attempt_bpress_res = first_attempt_bpress_res
    #                 cur_comp.first_attempt_bpress_off = first_attempt_bpress_off
    #                 cur_comp.second_attempt_bpress_res = second_attempt_bpress_res
    #                 cur_comp.second_attempt_bpress_off = second_attempt_bpress_off
    #                 cur_comp.third_attempt_bpress_res = third_attempt_bpress_res
    #                 cur_comp.third_attempt_bpress_off = third_attempt_bpress_off
    #                 cur_comp.first_attempt_dlift_res = first_attempt_dlift_res
    #                 cur_comp.first_attempt_dlift_off = first_attempt_dlift_off
    #                 cur_comp.second_attempt_dlift_res = second_attempt_dlift_res
    #                 cur_comp.second_attempt_dlift_off = second_attempt_dlift_off
    #                 cur_comp.third_attempt_dlift_res = third_attempt_dlift_res
    #                 cur_comp.third_attempt_dlift_off = third_attempt_dlift_off
    #                 cur_comp.competitor_translation = comp_transl
    #                 cur_comp.save()
    #                 return JsonResponse({"pil_h_squat": pil_h_squat, "pil_h_bpress": pil_h_bpress, "first_attempt_squat_res": first_attempt_squat_res,
    #                                      "first_attempt_squat_off": first_attempt_squat_off,
    #                                      "second_attempt_squat_res": second_attempt_squat_res,
    #                                      "second_attempt_squat_off": second_attempt_squat_off,
    #                                      "third_attempt_squat_res": third_attempt_squat_res,
    #                                      "third_attempt_squat_off": third_attempt_squat_off,
    #                                      "first_attempt_bpress_res": first_attempt_bpress_res,
    #                                      "first_attempt_bpress_off": first_attempt_bpress_off,
    #                                      "second_attempt_bpress_res": second_attempt_bpress_res,
    #                                      "second_attempt_bpress_off": second_attempt_bpress_off,
    #                                      "third_attempt_bpress_res": third_attempt_bpress_res,
    #                                      "third_attempt_bpress_off": third_attempt_bpress_off,
    #                                      "first_attempt_dlift_res": first_attempt_dlift_res,
    #                                      "first_attempt_dlift_off": first_attempt_dlift_off,
    #                                      "second_attempt_dlift_res": second_attempt_dlift_res,
    #                                      "second_attempt_dlift_off": second_attempt_dlift_off,
    #                                      "third_attempt_dlift_res": third_attempt_dlift_res,
    #                                      "third_attempt_dlift_off": third_attempt_dlift_off,
    #                                      "comp_transl": comp_transl}, status=200)
    #             if cur_transl_val == "Жим 3 класс" and third_attempt_bpress_off:
    #                 cur_comp.pillar_height_squat = pil_h_squat
    #                 cur_comp.pillar_height_bpress = pil_h_bpress
    #                 cur_comp.first_attempt_squat_res = first_attempt_squat_res
    #                 cur_comp.first_attempt_squat_off = first_attempt_squat_off
    #                 cur_comp.second_attempt_squat_res = second_attempt_squat_res
    #                 cur_comp.second_attempt_squat_off = second_attempt_squat_off
    #                 cur_comp.third_attempt_squat_res = third_attempt_squat_res
    #                 cur_comp.third_attempt_squat_off = third_attempt_squat_off
    #                 cur_comp.first_attempt_bpress_res = first_attempt_bpress_res
    #                 cur_comp.first_attempt_bpress_off = first_attempt_bpress_off
    #                 cur_comp.second_attempt_bpress_res = second_attempt_bpress_res
    #                 cur_comp.second_attempt_bpress_off = second_attempt_bpress_off
    #                 cur_comp.third_attempt_bpress_res = third_attempt_bpress_res
    #                 cur_comp.third_attempt_bpress_off = third_attempt_bpress_off
    #                 cur_comp.first_attempt_dlift_res = first_attempt_dlift_res
    #                 cur_comp.first_attempt_dlift_off = first_attempt_dlift_off
    #                 cur_comp.second_attempt_dlift_res = second_attempt_dlift_res
    #                 cur_comp.second_attempt_dlift_off = second_attempt_dlift_off
    #                 cur_comp.third_attempt_dlift_res = third_attempt_dlift_res
    #                 cur_comp.third_attempt_dlift_off = third_attempt_dlift_off
    #                 cur_comp.competitor_translation = comp_transl
    #                 cur_comp.save()
    #                 return JsonResponse({"pil_h_squat": pil_h_squat, "pil_h_bpress": pil_h_bpress, "first_attempt_squat_res": first_attempt_squat_res,
    #                                      "first_attempt_squat_off": first_attempt_squat_off,
    #                                      "second_attempt_squat_res": second_attempt_squat_res,
    #                                      "second_attempt_squat_off": second_attempt_squat_off,
    #                                      "third_attempt_squat_res": third_attempt_squat_res,
    #                                      "third_attempt_squat_off": third_attempt_squat_off,
    #                                      "first_attempt_bpress_res": first_attempt_bpress_res,
    #                                      "first_attempt_bpress_off": first_attempt_bpress_off,
    #                                      "second_attempt_bpress_res": second_attempt_bpress_res,
    #                                      "second_attempt_bpress_off": second_attempt_bpress_off,
    #                                      "third_attempt_bpress_res": third_attempt_bpress_res,
    #                                      "third_attempt_bpress_off": third_attempt_bpress_off,
    #                                      "first_attempt_dlift_res": first_attempt_dlift_res,
    #                                      "first_attempt_dlift_off": first_attempt_dlift_off,
    #                                      "second_attempt_dlift_res": second_attempt_dlift_res,
    #                                      "second_attempt_dlift_off": second_attempt_dlift_off,
    #                                      "third_attempt_dlift_res": third_attempt_dlift_res,
    #                                      "third_attempt_dlift_off": third_attempt_dlift_off,
    #                                      "comp_transl": comp_transl}, status=200)
    #             if cur_transl_val == "Тяга 1 класс" and first_attempt_dlift_off:
    #                 cur_comp.pillar_height_squat = pil_h_squat
    #                 cur_comp.pillar_height_bpress = pil_h_bpress
    #                 cur_comp.first_attempt_squat_res = first_attempt_squat_res
    #                 cur_comp.first_attempt_squat_off = first_attempt_squat_off
    #                 cur_comp.second_attempt_squat_res = second_attempt_squat_res
    #                 cur_comp.second_attempt_squat_off = second_attempt_squat_off
    #                 cur_comp.third_attempt_squat_res = third_attempt_squat_res
    #                 cur_comp.third_attempt_squat_off = third_attempt_squat_off
    #                 cur_comp.first_attempt_bpress_res = first_attempt_bpress_res
    #                 cur_comp.first_attempt_bpress_off = first_attempt_bpress_off
    #                 cur_comp.second_attempt_bpress_res = second_attempt_bpress_res
    #                 cur_comp.second_attempt_bpress_off = second_attempt_bpress_off
    #                 cur_comp.third_attempt_bpress_res = third_attempt_bpress_res
    #                 cur_comp.third_attempt_bpress_off = third_attempt_bpress_off
    #                 cur_comp.first_attempt_dlift_res = first_attempt_dlift_res
    #                 cur_comp.first_attempt_dlift_off = first_attempt_dlift_off
    #                 cur_comp.second_attempt_dlift_res = second_attempt_dlift_res
    #                 cur_comp.second_attempt_dlift_off = second_attempt_dlift_off
    #                 cur_comp.third_attempt_dlift_res = third_attempt_dlift_res
    #                 cur_comp.third_attempt_dlift_off = third_attempt_dlift_off
    #                 cur_comp.competitor_translation = comp_transl
    #                 cur_comp.save()
    #                 return JsonResponse({"pil_h_squat": pil_h_squat, "pil_h_bpress": pil_h_bpress, "first_attempt_squat_res": first_attempt_squat_res,
    #                                      "first_attempt_squat_off": first_attempt_squat_off,
    #                                      "second_attempt_squat_res": second_attempt_squat_res,
    #                                      "second_attempt_squat_off": second_attempt_squat_off,
    #                                      "third_attempt_squat_res": third_attempt_squat_res,
    #                                      "third_attempt_squat_off": third_attempt_squat_off,
    #                                      "first_attempt_bpress_res": first_attempt_bpress_res,
    #                                      "first_attempt_bpress_off": first_attempt_bpress_off,
    #                                      "second_attempt_bpress_res": second_attempt_bpress_res,
    #                                      "second_attempt_bpress_off": second_attempt_bpress_off,
    #                                      "third_attempt_bpress_res": third_attempt_bpress_res,
    #                                      "third_attempt_bpress_off": third_attempt_bpress_off,
    #                                      "first_attempt_dlift_res": first_attempt_dlift_res,
    #                                      "first_attempt_dlift_off": first_attempt_dlift_off,
    #                                      "second_attempt_dlift_res": second_attempt_dlift_res,
    #                                      "second_attempt_dlift_off": second_attempt_dlift_off,
    #                                      "third_attempt_dlift_res": third_attempt_dlift_res,
    #                                      "third_attempt_dlift_off": third_attempt_dlift_off,
    #                                      "comp_transl": comp_transl}, status=200)
    #             if cur_transl_val == "Тяга 2 класс" and second_attempt_dlift_off:
    #                 cur_comp.pillar_height_squat = pil_h_squat
    #                 cur_comp.pillar_height_bpress = pil_h_bpress
    #                 cur_comp.first_attempt_squat_res = first_attempt_squat_res
    #                 cur_comp.first_attempt_squat_off = first_attempt_squat_off
    #                 cur_comp.second_attempt_squat_res = second_attempt_squat_res
    #                 cur_comp.second_attempt_squat_off = second_attempt_squat_off
    #                 cur_comp.third_attempt_squat_res = third_attempt_squat_res
    #                 cur_comp.third_attempt_squat_off = third_attempt_squat_off
    #                 cur_comp.first_attempt_bpress_res = first_attempt_bpress_res
    #                 cur_comp.first_attempt_bpress_off = first_attempt_bpress_off
    #                 cur_comp.second_attempt_bpress_res = second_attempt_bpress_res
    #                 cur_comp.second_attempt_bpress_off = second_attempt_bpress_off
    #                 cur_comp.third_attempt_bpress_res = third_attempt_bpress_res
    #                 cur_comp.third_attempt_bpress_off = third_attempt_bpress_off
    #                 cur_comp.first_attempt_dlift_res = first_attempt_dlift_res
    #                 cur_comp.first_attempt_dlift_off = first_attempt_dlift_off
    #                 cur_comp.second_attempt_dlift_res = second_attempt_dlift_res
    #                 cur_comp.second_attempt_dlift_off = second_attempt_dlift_off
    #                 cur_comp.third_attempt_dlift_res = third_attempt_dlift_res
    #                 cur_comp.third_attempt_dlift_off = third_attempt_dlift_off
    #                 cur_comp.competitor_translation = comp_transl
    #                 cur_comp.save()
    #                 return JsonResponse({"pil_h_squat": pil_h_squat, "pil_h_bpress": pil_h_bpress, "first_attempt_squat_res": first_attempt_squat_res,
    #                                      "first_attempt_squat_off": first_attempt_squat_off,
    #                                      "second_attempt_squat_res": second_attempt_squat_res,
    #                                      "second_attempt_squat_off": second_attempt_squat_off,
    #                                      "third_attempt_squat_res": third_attempt_squat_res,
    #                                      "third_attempt_squat_off": third_attempt_squat_off,
    #                                      "first_attempt_bpress_res": first_attempt_bpress_res,
    #                                      "first_attempt_bpress_off": first_attempt_bpress_off,
    #                                      "second_attempt_bpress_res": second_attempt_bpress_res,
    #                                      "second_attempt_bpress_off": second_attempt_bpress_off,
    #                                      "third_attempt_bpress_res": third_attempt_bpress_res,
    #                                      "third_attempt_bpress_off": third_attempt_bpress_off,
    #                                      "first_attempt_dlift_res": first_attempt_dlift_res,
    #                                      "first_attempt_dlift_off": first_attempt_dlift_off,
    #                                      "second_attempt_dlift_res": second_attempt_dlift_res,
    #                                      "second_attempt_dlift_off": second_attempt_dlift_off,
    #                                      "third_attempt_dlift_res": third_attempt_dlift_res,
    #                                      "third_attempt_dlift_off": third_attempt_dlift_off,
    #                                      "comp_transl": comp_transl}, status=200)
    #             if cur_transl_val == "Тяга 3 класс" and third_attempt_dlift_off:
    #                 cur_comp.pillar_height_squat = pil_h_squat
    #                 cur_comp.pillar_height_bpress = pil_h_bpress
    #                 cur_comp.first_attempt_squat_res = first_attempt_squat_res
    #                 cur_comp.first_attempt_squat_off = first_attempt_squat_off
    #                 cur_comp.second_attempt_squat_res = second_attempt_squat_res
    #                 cur_comp.second_attempt_squat_off = second_attempt_squat_off
    #                 cur_comp.third_attempt_squat_res = third_attempt_squat_res
    #                 cur_comp.third_attempt_squat_off = third_attempt_squat_off
    #                 cur_comp.first_attempt_bpress_res = first_attempt_bpress_res
    #                 cur_comp.first_attempt_bpress_off = first_attempt_bpress_off
    #                 cur_comp.second_attempt_bpress_res = second_attempt_bpress_res
    #                 cur_comp.second_attempt_bpress_off = second_attempt_bpress_off
    #                 cur_comp.third_attempt_bpress_res = third_attempt_bpress_res
    #                 cur_comp.third_attempt_bpress_off = third_attempt_bpress_off
    #                 cur_comp.first_attempt_dlift_res = first_attempt_dlift_res
    #                 cur_comp.first_attempt_dlift_off = first_attempt_dlift_off
    #                 cur_comp.second_attempt_dlift_res = second_attempt_dlift_res
    #                 cur_comp.second_attempt_dlift_off = second_attempt_dlift_off
    #                 cur_comp.third_attempt_dlift_res = third_attempt_dlift_res
    #                 cur_comp.third_attempt_dlift_off = third_attempt_dlift_off
    #                 cur_comp.competitor_translation = comp_transl
    #                 cur_comp.save()
    #                 return JsonResponse({"pil_h_squat": pil_h_squat, "pil_h_bpress": pil_h_bpress, "first_attempt_squat_res": first_attempt_squat_res,
    #                                      "first_attempt_squat_off": first_attempt_squat_off,
    #                                      "second_attempt_squat_res": second_attempt_squat_res,
    #                                      "second_attempt_squat_off": second_attempt_squat_off,
    #                                      "third_attempt_squat_res": third_attempt_squat_res,
    #                                      "third_attempt_squat_off": third_attempt_squat_off,
    #                                      "first_attempt_bpress_res": first_attempt_bpress_res,
    #                                      "first_attempt_bpress_off": first_attempt_bpress_off,
    #                                      "second_attempt_bpress_res": second_attempt_bpress_res,
    #                                      "second_attempt_bpress_off": second_attempt_bpress_off,
    #                                      "third_attempt_bpress_res": third_attempt_bpress_res,
    #                                      "third_attempt_bpress_off": third_attempt_bpress_off,
    #                                      "first_attempt_dlift_res": first_attempt_dlift_res,
    #                                      "first_attempt_dlift_off": first_attempt_dlift_off,
    #                                      "second_attempt_dlift_res": second_attempt_dlift_res,
    #                                      "second_attempt_dlift_off": second_attempt_dlift_off,
    #                                      "third_attempt_dlift_res": third_attempt_dlift_res,
    #                                      "third_attempt_dlift_off": third_attempt_dlift_off,
    #                                      "comp_transl": comp_transl}, status=200)
    #     if comp_transl:
    #         cur_comp.pillar_height_squat = pil_h_squat
    #         cur_comp.pillar_height_bpress = pil_h_bpress
    #         cur_comp.first_attempt_squat_res = first_attempt_squat_res
    #         cur_comp.first_attempt_squat_off = first_attempt_squat_off
    #         cur_comp.second_attempt_squat_res = second_attempt_squat_res
    #         cur_comp.second_attempt_squat_off = second_attempt_squat_off
    #         cur_comp.third_attempt_squat_res = third_attempt_squat_res
    #         cur_comp.third_attempt_squat_off = third_attempt_squat_off
    #         cur_comp.first_attempt_bpress_res = first_attempt_bpress_res
    #         cur_comp.first_attempt_bpress_off = first_attempt_bpress_off
    #         cur_comp.second_attempt_bpress_res = second_attempt_bpress_res
    #         cur_comp.second_attempt_bpress_off = second_attempt_bpress_off
    #         cur_comp.third_attempt_bpress_res = third_attempt_bpress_res
    #         cur_comp.third_attempt_bpress_off = third_attempt_bpress_off
    #         cur_comp.first_attempt_dlift_res = first_attempt_dlift_res
    #         cur_comp.first_attempt_dlift_off = first_attempt_dlift_off
    #         cur_comp.second_attempt_dlift_res = second_attempt_dlift_res
    #         cur_comp.second_attempt_dlift_off = second_attempt_dlift_off
    #         cur_comp.third_attempt_dlift_res = third_attempt_dlift_res
    #         cur_comp.third_attempt_dlift_off = third_attempt_dlift_off
    #         cur_comp.competitor_translation = comp_transl
    #         cur_comp.save()
    #         return JsonResponse(
    #             {"pil_h_squat": pil_h_squat, "pil_h_bpress": pil_h_bpress, "first_attempt_squat_res": first_attempt_squat_res, "first_attempt_squat_off": first_attempt_squat_off,
    #              "second_attempt_squat_res": second_attempt_squat_res,
    #              "second_attempt_squat_off": second_attempt_squat_off,
    #              "third_attempt_squat_res": third_attempt_squat_res, "third_attempt_squat_off": third_attempt_squat_off,
    #              "first_attempt_bpress_res": first_attempt_bpress_res,
    #              "first_attempt_bpress_off": first_attempt_bpress_off,
    #              "second_attempt_bpress_res": second_attempt_bpress_res,
    #              "second_attempt_bpress_off": second_attempt_bpress_off,
    #              "third_attempt_bpress_res": third_attempt_bpress_res,
    #              "third_attempt_bpress_off": third_attempt_bpress_off,
    #              "first_attempt_dlift_res": first_attempt_dlift_res, "first_attempt_dlift_off": first_attempt_dlift_off,
    #              "second_attempt_dlift_res": second_attempt_dlift_res,
    #              "second_attempt_dlift_off": second_attempt_dlift_off,
    #              "third_attempt_dlift_res": third_attempt_dlift_res, "third_attempt_dlift_off": third_attempt_dlift_off,
    #              "comp_transl": comp_transl}, status=200)
    #     else:
    #         cur_comp.pillar_height_squat = pil_h_squat
    #         cur_comp.pillar_height_bpress = pil_h_bpress
    #         cur_comp.first_attempt_squat_res = first_attempt_squat_res
    #         cur_comp.first_attempt_squat_off = first_attempt_squat_off
    #         cur_comp.second_attempt_squat_res = second_attempt_squat_res
    #         cur_comp.second_attempt_squat_off = second_attempt_squat_off
    #         cur_comp.third_attempt_squat_res = third_attempt_squat_res
    #         cur_comp.third_attempt_squat_off = third_attempt_squat_off
    #         cur_comp.first_attempt_bpress_res = first_attempt_bpress_res
    #         cur_comp.first_attempt_bpress_off = first_attempt_bpress_off
    #         cur_comp.second_attempt_bpress_res = second_attempt_bpress_res
    #         cur_comp.second_attempt_bpress_off = second_attempt_bpress_off
    #         cur_comp.third_attempt_bpress_res = third_attempt_bpress_res
    #         cur_comp.third_attempt_bpress_off = third_attempt_bpress_off
    #         cur_comp.first_attempt_dlift_res = first_attempt_dlift_res
    #         cur_comp.first_attempt_dlift_off = first_attempt_dlift_off
    #         cur_comp.second_attempt_dlift_res = second_attempt_dlift_res
    #         cur_comp.second_attempt_dlift_off = second_attempt_dlift_off
    #         cur_comp.third_attempt_dlift_res = third_attempt_dlift_res
    #         cur_comp.third_attempt_dlift_off = third_attempt_dlift_off
    #         cur_comp.competitor_translation = None
    #         cur_comp.save()
    #         return JsonResponse(
    #             {"pil_h_squat": pil_h_squat, "pil_h_bpress": pil_h_bpress, "first_attempt_squat_res": first_attempt_squat_res, "first_attempt_squat_off": first_attempt_squat_off,
    #              "second_attempt_squat_res": second_attempt_squat_res,
    #              "second_attempt_squat_off": second_attempt_squat_off,
    #              "third_attempt_squat_res": third_attempt_squat_res, "third_attempt_squat_off": third_attempt_squat_off,
    #              "first_attempt_bpress_res": first_attempt_bpress_res,
    #              "first_attempt_bpress_off": first_attempt_bpress_off,
    #              "second_attempt_bpress_res": second_attempt_bpress_res,
    #              "second_attempt_bpress_off": second_attempt_bpress_off,
    #              "third_attempt_bpress_res": third_attempt_bpress_res,
    #              "third_attempt_bpress_off": third_attempt_bpress_off,
    #              "first_attempt_dlift_res": first_attempt_dlift_res, "first_attempt_dlift_off": first_attempt_dlift_off,
    #              "second_attempt_dlift_res": second_attempt_dlift_res,
    #              "second_attempt_dlift_off": second_attempt_dlift_off,
    #              "third_attempt_dlift_res": third_attempt_dlift_res, "third_attempt_dlift_off": third_attempt_dlift_off,
    #              "comp_transl": comp_transl}, status=200)
    #     # , "testjson": serializers.serialize("json", CompetitionProtocols.objects.all().order_by(
    #         # F("first_attempt_squat_res").asc(nulls_last=True)), use_natural_foreign_keys=True)
    # if 'secr-page_form-btn-bpclass' in request.POST:
    #     comp_pk = request.POST.get('secr-page_form-btn-bpclass')
    #     pil_h_bpress = request.POST.get('secr-page_form-pil_h_bpress')
    #     if pil_h_bpress:
    #         pil_h_bpress = pil_h_bpress
    #     else:
    #         pil_h_bpress = None
    #     first_attempt_bpress_res = request.POST.get('secr-page_form-fat-bpress-res')
    #     if first_attempt_bpress_res:
    #         first_attempt_bpress_res = Decimal(first_attempt_bpress_res)
    #     else:
    #         first_attempt_bpress_res = None
    #     first_attempt_bpress_off = request.POST.get('secr-page_form-fat-bpress-off')
    #     second_attempt_bpress_res = request.POST.get('secr-page_form-sat-bpress-res')
    #     if second_attempt_bpress_res:
    #         second_attempt_bpress_res = Decimal(second_attempt_bpress_res)
    #     else:
    #         second_attempt_bpress_res = None
    #     second_attempt_bpress_off = request.POST.get('secr-page_form-sat-bpress-off')
    #     third_attempt_bpress_res = request.POST.get('secr-page_form-tat-bpress-res')
    #     if third_attempt_bpress_res:
    #         third_attempt_bpress_res = Decimal(third_attempt_bpress_res)
    #     else:
    #         third_attempt_bpress_res = None
    #     third_attempt_bpress_off = request.POST.get('secr-page_form-tat-bpress-off')
    #     cur_comp = CompetitionProtocols.objects.get(pk=comp_pk)
    #     comp_transl = request.POST.get('secr-page_chckbx')
    #     cur_transl = CompetitionProtocols.objects.exclude(competitor_translation=None)
    #     if comp_transl and cur_transl:
    #         if comp_transl != cur_transl[0].competitor_translation:
    #             return JsonResponse(
    #                 {"error": "Нельзя запускать две трансляции одновременно. Сначала закончите существующую."},
    #                 status=302)
    #     if cur_transl:
    #         cur_transl_val = cur_transl.values_list('competitor_translation', flat=True)[0]
    #         cur_transl_pk = cur_transl.values_list('id', flat=True)[0]
    #         if comp_pk == str(cur_transl_pk):
    #             if cur_transl_val == "Жим 1 класс" and first_attempt_bpress_off:
    #                 cur_comp.pillar_height_bpress = pil_h_bpress
    #                 cur_comp.first_attempt_bpress_res = first_attempt_bpress_res
    #                 cur_comp.first_attempt_bpress_off = first_attempt_bpress_off
    #                 cur_comp.second_attempt_bpress_res = second_attempt_bpress_res
    #                 cur_comp.second_attempt_bpress_off = second_attempt_bpress_off
    #                 cur_comp.third_attempt_bpress_res = third_attempt_bpress_res
    #                 cur_comp.third_attempt_bpress_off = third_attempt_bpress_off
    #                 cur_comp.competitor_translation = comp_transl
    #                 cur_comp.save()
    #                 return JsonResponse({"pil_h_bpress": pil_h_bpress, "first_attempt_bpress_res": first_attempt_bpress_res,
    #                                      "first_attempt_bpress_off": first_attempt_bpress_off,
    #                                      "second_attempt_bpress_res": second_attempt_bpress_res,
    #                                      "second_attempt_bpress_off": second_attempt_bpress_off,
    #                                      "third_attempt_bpress_res": third_attempt_bpress_res,
    #                                      "third_attempt_bpress_off": third_attempt_bpress_off,
    #                                      "comp_transl": comp_transl}, status=200)
    #             if cur_transl_val == "Жим 2 класс" and second_attempt_bpress_off:
    #                 cur_comp.pillar_height_bpress = pil_h_bpress
    #                 cur_comp.first_attempt_bpress_res = first_attempt_bpress_res
    #                 cur_comp.first_attempt_bpress_off = first_attempt_bpress_off
    #                 cur_comp.second_attempt_bpress_res = second_attempt_bpress_res
    #                 cur_comp.second_attempt_bpress_off = second_attempt_bpress_off
    #                 cur_comp.third_attempt_bpress_res = third_attempt_bpress_res
    #                 cur_comp.third_attempt_bpress_off = third_attempt_bpress_off
    #                 cur_comp.competitor_translation = comp_transl
    #                 cur_comp.save()
    #                 return JsonResponse({"pil_h_bpress": pil_h_bpress, "first_attempt_bpress_res": first_attempt_bpress_res,
    #                                      "first_attempt_bpress_off": first_attempt_bpress_off,
    #                                      "second_attempt_bpress_res": second_attempt_bpress_res,
    #                                      "second_attempt_bpress_off": second_attempt_bpress_off,
    #                                      "third_attempt_bpress_res": third_attempt_bpress_res,
    #                                      "third_attempt_bpress_off": third_attempt_bpress_off,
    #                                      "comp_transl": comp_transl}, status=200)
    #             if cur_transl_val == "Жим 3 класс" and third_attempt_bpress_off:
    #                 cur_comp.pillar_height_bpress = pil_h_bpress
    #                 cur_comp.first_attempt_bpress_res = first_attempt_bpress_res
    #                 cur_comp.first_attempt_bpress_off = first_attempt_bpress_off
    #                 cur_comp.second_attempt_bpress_res = second_attempt_bpress_res
    #                 cur_comp.second_attempt_bpress_off = second_attempt_bpress_off
    #                 cur_comp.third_attempt_bpress_res = third_attempt_bpress_res
    #                 cur_comp.third_attempt_bpress_off = third_attempt_bpress_off
    #                 cur_comp.competitor_translation = comp_transl
    #                 cur_comp.save()
    #                 return JsonResponse({"pil_h_bpress": pil_h_bpress, "first_attempt_bpress_res": first_attempt_bpress_res,
    #                                      "first_attempt_bpress_off": first_attempt_bpress_off,
    #                                      "second_attempt_bpress_res": second_attempt_bpress_res,
    #                                      "second_attempt_bpress_off": second_attempt_bpress_off,
    #                                      "third_attempt_bpress_res": third_attempt_bpress_res,
    #                                      "third_attempt_bpress_off": third_attempt_bpress_off,
    #                                      "comp_transl": comp_transl}, status=200)
    #     if comp_transl:
    #         cur_comp.pillar_height_bpress = pil_h_bpress
    #         cur_comp.first_attempt_bpress_res = first_attempt_bpress_res
    #         cur_comp.first_attempt_bpress_off = first_attempt_bpress_off
    #         cur_comp.second_attempt_bpress_res = second_attempt_bpress_res
    #         cur_comp.second_attempt_bpress_off = second_attempt_bpress_off
    #         cur_comp.third_attempt_bpress_res = third_attempt_bpress_res
    #         cur_comp.third_attempt_bpress_off = third_attempt_bpress_off
    #         cur_comp.competitor_translation = comp_transl
    #         cur_comp.save()
    #         return JsonResponse(
    #             {"pil_h_bpress": pil_h_bpress, "first_attempt_bpress_res": first_attempt_bpress_res,
    #              "first_attempt_bpress_off": first_attempt_bpress_off,
    #              "second_attempt_bpress_res": second_attempt_bpress_res,
    #              "second_attempt_bpress_off": second_attempt_bpress_off,
    #              "third_attempt_bpress_res": third_attempt_bpress_res,
    #              "third_attempt_bpress_off": third_attempt_bpress_off,
    #              "comp_transl": comp_transl}, status=200)
    #     else:
    #         cur_comp.pillar_height_bpress = pil_h_bpress
    #         cur_comp.first_attempt_bpress_res = first_attempt_bpress_res
    #         cur_comp.first_attempt_bpress_off = first_attempt_bpress_off
    #         cur_comp.second_attempt_bpress_res = second_attempt_bpress_res
    #         cur_comp.second_attempt_bpress_off = second_attempt_bpress_off
    #         cur_comp.third_attempt_bpress_res = third_attempt_bpress_res
    #         cur_comp.third_attempt_bpress_off = third_attempt_bpress_off
    #         cur_comp.competitor_translation = None
    #         cur_comp.save()
    #         return JsonResponse(
    #             {"pil_h_bpress": pil_h_bpress, "first_attempt_bpress_res": first_attempt_bpress_res,
    #              "first_attempt_bpress_off": first_attempt_bpress_off,
    #              "second_attempt_bpress_res": second_attempt_bpress_res,
    #              "second_attempt_bpress_off": second_attempt_bpress_off,
    #              "third_attempt_bpress_res": third_attempt_bpress_res,
    #              "third_attempt_bpress_off": third_attempt_bpress_off,
    #              "comp_transl": comp_transl}, status=200)
    # if 'secr-page_form-btn-plek' in request.POST:
    #     comp_pk = request.POST.get('secr-page_form-btn-plek')
    #     pil_h_squat = request.POST.get('secr-page_form-pil_h_squat')
    #     pil_h_bpress = request.POST.get('secr-page_form-pil_h_bpress')
    #     if pil_h_squat:
    #         pil_h_squat = pil_h_squat
    #     else:
    #         pil_h_squat = None
    #     if pil_h_bpress:
    #         pil_h_bpress = pil_h_bpress
    #     else:
    #         pil_h_bpress = None
    #     first_attempt_squat_res = request.POST.get('secr-page_form-fat-squat-res')
    #     if first_attempt_squat_res:
    #         first_attempt_squat_res = Decimal(first_attempt_squat_res)
    #     else:
    #         first_attempt_squat_res = None
    #     first_attempt_squat_off = request.POST.get('secr-page_form-fat-squat-off')
    #     second_attempt_squat_res = request.POST.get('secr-page_form-sat-squat-res')
    #     if second_attempt_squat_res:
    #         second_attempt_squat_res = Decimal(second_attempt_squat_res)
    #     else:
    #         second_attempt_squat_res = None
    #     second_attempt_squat_off = request.POST.get('secr-page_form-sat-squat-off')
    #     third_attempt_squat_res = request.POST.get('secr-page_form-tat-squat-res')
    #     if third_attempt_squat_res:
    #         third_attempt_squat_res = Decimal(third_attempt_squat_res)
    #     else:
    #         third_attempt_squat_res = None
    #     third_attempt_squat_off = request.POST.get('secr-page_form-tat-squat-off')
    #     first_attempt_bpress_res = request.POST.get('secr-page_form-fat-bpress-res')
    #     if first_attempt_bpress_res:
    #         first_attempt_bpress_res = Decimal(first_attempt_bpress_res)
    #     else:
    #         first_attempt_bpress_res = None
    #     first_attempt_bpress_off = request.POST.get('secr-page_form-fat-bpress-off')
    #     second_attempt_bpress_res = request.POST.get('secr-page_form-sat-bpress-res')
    #     if second_attempt_bpress_res:
    #         second_attempt_bpress_res = Decimal(second_attempt_bpress_res)
    #     else:
    #         second_attempt_bpress_res = None
    #     second_attempt_bpress_off = request.POST.get('secr-page_form-sat-bpress-off')
    #     third_attempt_bpress_res = request.POST.get('secr-page_form-tat-bpress-res')
    #     if third_attempt_bpress_res:
    #         third_attempt_bpress_res = Decimal(third_attempt_bpress_res)
    #     else:
    #         third_attempt_bpress_res = None
    #     third_attempt_bpress_off = request.POST.get('secr-page_form-tat-bpress-off')
    #     first_attempt_dlift_res = request.POST.get('secr-page_form-fat-dlift-res')
    #     if first_attempt_dlift_res:
    #         first_attempt_dlift_res = Decimal(first_attempt_dlift_res)
    #     else:
    #         first_attempt_dlift_res = None
    #     first_attempt_dlift_off = request.POST.get('secr-page_form-fat-dlift-off')
    #     second_attempt_dlift_res = request.POST.get('secr-page_form-sat-dlift-res')
    #     if second_attempt_dlift_res:
    #         second_attempt_dlift_res = Decimal(second_attempt_dlift_res)
    #     else:
    #         second_attempt_dlift_res = None
    #     second_attempt_dlift_off = request.POST.get('secr-page_form-sat-dlift-off')
    #     third_attempt_dlift_res = request.POST.get('secr-page_form-tat-dlift-res')
    #     if third_attempt_dlift_res:
    #         third_attempt_dlift_res = Decimal(third_attempt_dlift_res)
    #     else:
    #         third_attempt_dlift_res = None
    #     third_attempt_dlift_off = request.POST.get('secr-page_form-tat-dlift-off')
    #     cur_comp = CompetitionProtocols.objects.get(pk=comp_pk)
    #     comp_transl = request.POST.get('secr-page_chckbx')
    #     cur_transl = CompetitionProtocols.objects.exclude(competitor_translation=None)
    #     if comp_transl and cur_transl:
    #         if comp_transl != cur_transl[0].competitor_translation:
    #             return JsonResponse(
    #                 {"error": "Нельзя запускать две трансляции одновременно. Сначала закончите существующую."},
    #                 status=302)
    #     if cur_transl:
    #         cur_transl_val = cur_transl.values_list('competitor_translation', flat=True)[0]
    #         cur_transl_pk = cur_transl.values_list('id', flat=True)[0]
    #         if comp_pk == str(cur_transl_pk):
    #             if cur_transl_val == "Присед 1 экип" and first_attempt_squat_off:
    #                 cur_comp.pillar_height_squat = pil_h_squat
    #                 cur_comp.pillar_height_bpress = pil_h_bpress
    #                 cur_comp.first_attempt_squat_res_ek = first_attempt_squat_res
    #                 cur_comp.first_attempt_squat_off_ek = first_attempt_squat_off
    #                 cur_comp.second_attempt_squat_res_ek = second_attempt_squat_res
    #                 cur_comp.second_attempt_squat_off_ek = second_attempt_squat_off
    #                 cur_comp.third_attempt_squat_res_ek = third_attempt_squat_res
    #                 cur_comp.third_attempt_squat_off_ek = third_attempt_squat_off
    #                 cur_comp.first_attempt_bpress_res_ek = first_attempt_bpress_res
    #                 cur_comp.first_attempt_bpress_off_ek = first_attempt_bpress_off
    #                 cur_comp.second_attempt_bpress_res_ek = second_attempt_bpress_res
    #                 cur_comp.second_attempt_bpress_off_ek = second_attempt_bpress_off
    #                 cur_comp.third_attempt_bpress_res_ek = third_attempt_bpress_res
    #                 cur_comp.third_attempt_bpress_off_ek = third_attempt_bpress_off
    #                 cur_comp.first_attempt_dlift_res_ek = first_attempt_dlift_res
    #                 cur_comp.first_attempt_dlift_off_ek = first_attempt_dlift_off
    #                 cur_comp.second_attempt_dlift_res_ek = second_attempt_dlift_res
    #                 cur_comp.second_attempt_dlift_off_ek = second_attempt_dlift_off
    #                 cur_comp.third_attempt_dlift_res_ek = third_attempt_dlift_res
    #                 cur_comp.third_attempt_dlift_off_ek = third_attempt_dlift_off
    #                 cur_comp.competitor_translation = comp_transl
    #                 cur_comp.save()
    #                 return JsonResponse({"pil_h_squat": pil_h_squat, "pil_h_bpress": pil_h_bpress, "first_attempt_squat_res": first_attempt_squat_res,
    #                                      "first_attempt_squat_off": first_attempt_squat_off,
    #                                      "second_attempt_squat_res": second_attempt_squat_res,
    #                                      "second_attempt_squat_off": second_attempt_squat_off,
    #                                      "third_attempt_squat_res": third_attempt_squat_res,
    #                                      "third_attempt_squat_off": third_attempt_squat_off,
    #                                      "first_attempt_bpress_res": first_attempt_bpress_res,
    #                                      "first_attempt_bpress_off": first_attempt_bpress_off,
    #                                      "second_attempt_bpress_res": second_attempt_bpress_res,
    #                                      "second_attempt_bpress_off": second_attempt_bpress_off,
    #                                      "third_attempt_bpress_res": third_attempt_bpress_res,
    #                                      "third_attempt_bpress_off": third_attempt_bpress_off,
    #                                      "first_attempt_dlift_res": first_attempt_dlift_res,
    #                                      "first_attempt_dlift_off": first_attempt_dlift_off,
    #                                      "second_attempt_dlift_res": second_attempt_dlift_res,
    #                                      "second_attempt_dlift_off": second_attempt_dlift_off,
    #                                      "third_attempt_dlift_res": third_attempt_dlift_res,
    #                                      "third_attempt_dlift_off": third_attempt_dlift_off,
    #                                      "comp_transl": comp_transl}, status=200)
    #             if cur_transl_val == "Присед 2 экип" and second_attempt_squat_off:
    #                 cur_comp.pillar_height_squat = pil_h_squat
    #                 cur_comp.pillar_height_bpress = pil_h_bpress
    #                 cur_comp.first_attempt_squat_res_ek = first_attempt_squat_res
    #                 cur_comp.first_attempt_squat_off_ek = first_attempt_squat_off
    #                 cur_comp.second_attempt_squat_res_ek = second_attempt_squat_res
    #                 cur_comp.second_attempt_squat_off_ek = second_attempt_squat_off
    #                 cur_comp.third_attempt_squat_res_ek = third_attempt_squat_res
    #                 cur_comp.third_attempt_squat_off_ek = third_attempt_squat_off
    #                 cur_comp.first_attempt_bpress_res_ek = first_attempt_bpress_res
    #                 cur_comp.first_attempt_bpress_off_ek = first_attempt_bpress_off
    #                 cur_comp.second_attempt_bpress_res_ek = second_attempt_bpress_res
    #                 cur_comp.second_attempt_bpress_off_ek = second_attempt_bpress_off
    #                 cur_comp.third_attempt_bpress_res_ek = third_attempt_bpress_res
    #                 cur_comp.third_attempt_bpress_off_ek = third_attempt_bpress_off
    #                 cur_comp.first_attempt_dlift_res_ek = first_attempt_dlift_res
    #                 cur_comp.first_attempt_dlift_off_ek = first_attempt_dlift_off
    #                 cur_comp.second_attempt_dlift_res_ek = second_attempt_dlift_res
    #                 cur_comp.second_attempt_dlift_off_ek = second_attempt_dlift_off
    #                 cur_comp.third_attempt_dlift_res_ek = third_attempt_dlift_res
    #                 cur_comp.third_attempt_dlift_off_ek = third_attempt_dlift_off
    #                 cur_comp.competitor_translation = comp_transl
    #                 cur_comp.save()
    #                 return JsonResponse({"pil_h_squat": pil_h_squat, "pil_h_bpress": pil_h_bpress, "first_attempt_squat_res": first_attempt_squat_res,
    #                                      "first_attempt_squat_off": first_attempt_squat_off,
    #                                      "second_attempt_squat_res": second_attempt_squat_res,
    #                                      "second_attempt_squat_off": second_attempt_squat_off,
    #                                      "third_attempt_squat_res": third_attempt_squat_res,
    #                                      "third_attempt_squat_off": third_attempt_squat_off,
    #                                      "first_attempt_bpress_res": first_attempt_bpress_res,
    #                                      "first_attempt_bpress_off": first_attempt_bpress_off,
    #                                      "second_attempt_bpress_res": second_attempt_bpress_res,
    #                                      "second_attempt_bpress_off": second_attempt_bpress_off,
    #                                      "third_attempt_bpress_res": third_attempt_bpress_res,
    #                                      "third_attempt_bpress_off": third_attempt_bpress_off,
    #                                      "first_attempt_dlift_res": first_attempt_dlift_res,
    #                                      "first_attempt_dlift_off": first_attempt_dlift_off,
    #                                      "second_attempt_dlift_res": second_attempt_dlift_res,
    #                                      "second_attempt_dlift_off": second_attempt_dlift_off,
    #                                      "third_attempt_dlift_res": third_attempt_dlift_res,
    #                                      "third_attempt_dlift_off": third_attempt_dlift_off,
    #                                      "comp_transl": comp_transl}, status=200)
    #             if cur_transl_val == "Присед 3 экип" and third_attempt_squat_off:
    #                 cur_comp.pillar_height_squat = pil_h_squat
    #                 cur_comp.pillar_height_bpress = pil_h_bpress
    #                 cur_comp.first_attempt_squat_res_ek = first_attempt_squat_res
    #                 cur_comp.first_attempt_squat_off_ek = first_attempt_squat_off
    #                 cur_comp.second_attempt_squat_res_ek = second_attempt_squat_res
    #                 cur_comp.second_attempt_squat_off_ek = second_attempt_squat_off
    #                 cur_comp.third_attempt_squat_res_ek = third_attempt_squat_res
    #                 cur_comp.third_attempt_squat_off_ek = third_attempt_squat_off
    #                 cur_comp.first_attempt_bpress_res_ek = first_attempt_bpress_res
    #                 cur_comp.first_attempt_bpress_off_ek = first_attempt_bpress_off
    #                 cur_comp.second_attempt_bpress_res_ek = second_attempt_bpress_res
    #                 cur_comp.second_attempt_bpress_off_ek = second_attempt_bpress_off
    #                 cur_comp.third_attempt_bpress_res_ek = third_attempt_bpress_res
    #                 cur_comp.third_attempt_bpress_off_ek = third_attempt_bpress_off
    #                 cur_comp.first_attempt_dlift_res_ek = first_attempt_dlift_res
    #                 cur_comp.first_attempt_dlift_off_ek = first_attempt_dlift_off
    #                 cur_comp.second_attempt_dlift_res_ek = second_attempt_dlift_res
    #                 cur_comp.second_attempt_dlift_off_ek = second_attempt_dlift_off
    #                 cur_comp.third_attempt_dlift_res_ek = third_attempt_dlift_res
    #                 cur_comp.third_attempt_dlift_off_ek = third_attempt_dlift_off
    #                 cur_comp.competitor_translation = comp_transl
    #                 cur_comp.save()
    #                 return JsonResponse({"pil_h_squat": pil_h_squat, "pil_h_bpress": pil_h_bpress, "first_attempt_squat_res": first_attempt_squat_res,
    #                                      "first_attempt_squat_off": first_attempt_squat_off,
    #                                      "second_attempt_squat_res": second_attempt_squat_res,
    #                                      "second_attempt_squat_off": second_attempt_squat_off,
    #                                      "third_attempt_squat_res": third_attempt_squat_res,
    #                                      "third_attempt_squat_off": third_attempt_squat_off,
    #                                      "first_attempt_bpress_res": first_attempt_bpress_res,
    #                                      "first_attempt_bpress_off": first_attempt_bpress_off,
    #                                      "second_attempt_bpress_res": second_attempt_bpress_res,
    #                                      "second_attempt_bpress_off": second_attempt_bpress_off,
    #                                      "third_attempt_bpress_res": third_attempt_bpress_res,
    #                                      "third_attempt_bpress_off": third_attempt_bpress_off,
    #                                      "first_attempt_dlift_res": first_attempt_dlift_res,
    #                                      "first_attempt_dlift_off": first_attempt_dlift_off,
    #                                      "second_attempt_dlift_res": second_attempt_dlift_res,
    #                                      "second_attempt_dlift_off": second_attempt_dlift_off,
    #                                      "third_attempt_dlift_res": third_attempt_dlift_res,
    #                                      "third_attempt_dlift_off": third_attempt_dlift_off,
    #                                      "comp_transl": comp_transl}, status=200)
    #             if cur_transl_val == "Жим 1 экип" and first_attempt_bpress_off:
    #                 cur_comp.pillar_height_squat = pil_h_squat
    #                 cur_comp.pillar_height_bpress = pil_h_bpress
    #                 cur_comp.first_attempt_squat_res_ek = first_attempt_squat_res
    #                 cur_comp.first_attempt_squat_off_ek = first_attempt_squat_off
    #                 cur_comp.second_attempt_squat_res_ek = second_attempt_squat_res
    #                 cur_comp.second_attempt_squat_off_ek = second_attempt_squat_off
    #                 cur_comp.third_attempt_squat_res_ek = third_attempt_squat_res
    #                 cur_comp.third_attempt_squat_off_ek = third_attempt_squat_off
    #                 cur_comp.first_attempt_bpress_res_ek = first_attempt_bpress_res
    #                 cur_comp.first_attempt_bpress_off_ek = first_attempt_bpress_off
    #                 cur_comp.second_attempt_bpress_res_ek = second_attempt_bpress_res
    #                 cur_comp.second_attempt_bpress_off_ek = second_attempt_bpress_off
    #                 cur_comp.third_attempt_bpress_res_ek = third_attempt_bpress_res
    #                 cur_comp.third_attempt_bpress_off_ek = third_attempt_bpress_off
    #                 cur_comp.first_attempt_dlift_res_ek = first_attempt_dlift_res
    #                 cur_comp.first_attempt_dlift_off_ek = first_attempt_dlift_off
    #                 cur_comp.second_attempt_dlift_res_ek = second_attempt_dlift_res
    #                 cur_comp.second_attempt_dlift_off_ek = second_attempt_dlift_off
    #                 cur_comp.third_attempt_dlift_res_ek = third_attempt_dlift_res
    #                 cur_comp.third_attempt_dlift_off_ek = third_attempt_dlift_off
    #                 cur_comp.competitor_translation = comp_transl
    #                 cur_comp.save()
    #                 return JsonResponse({"pil_h_squat": pil_h_squat, "pil_h_bpress": pil_h_bpress, "first_attempt_squat_res": first_attempt_squat_res,
    #                                      "first_attempt_squat_off": first_attempt_squat_off,
    #                                      "second_attempt_squat_res": second_attempt_squat_res,
    #                                      "second_attempt_squat_off": second_attempt_squat_off,
    #                                      "third_attempt_squat_res": third_attempt_squat_res,
    #                                      "third_attempt_squat_off": third_attempt_squat_off,
    #                                      "first_attempt_bpress_res": first_attempt_bpress_res,
    #                                      "first_attempt_bpress_off": first_attempt_bpress_off,
    #                                      "second_attempt_bpress_res": second_attempt_bpress_res,
    #                                      "second_attempt_bpress_off": second_attempt_bpress_off,
    #                                      "third_attempt_bpress_res": third_attempt_bpress_res,
    #                                      "third_attempt_bpress_off": third_attempt_bpress_off,
    #                                      "first_attempt_dlift_res": first_attempt_dlift_res,
    #                                      "first_attempt_dlift_off": first_attempt_dlift_off,
    #                                      "second_attempt_dlift_res": second_attempt_dlift_res,
    #                                      "second_attempt_dlift_off": second_attempt_dlift_off,
    #                                      "third_attempt_dlift_res": third_attempt_dlift_res,
    #                                      "third_attempt_dlift_off": third_attempt_dlift_off,
    #                                      "comp_transl": comp_transl}, status=200)
    #             if cur_transl_val == "Жим 2 экип" and second_attempt_bpress_off:
    #                 cur_comp.pillar_height_squat = pil_h_squat
    #                 cur_comp.pillar_height_bpress = pil_h_bpress
    #                 cur_comp.first_attempt_squat_res_ek = first_attempt_squat_res
    #                 cur_comp.first_attempt_squat_off_ek = first_attempt_squat_off
    #                 cur_comp.second_attempt_squat_res_ek = second_attempt_squat_res
    #                 cur_comp.second_attempt_squat_off_ek = second_attempt_squat_off
    #                 cur_comp.third_attempt_squat_res_ek = third_attempt_squat_res
    #                 cur_comp.third_attempt_squat_off_ek = third_attempt_squat_off
    #                 cur_comp.first_attempt_bpress_res_ek = first_attempt_bpress_res
    #                 cur_comp.first_attempt_bpress_off_ek = first_attempt_bpress_off
    #                 cur_comp.second_attempt_bpress_res_ek = second_attempt_bpress_res
    #                 cur_comp.second_attempt_bpress_off_ek = second_attempt_bpress_off
    #                 cur_comp.third_attempt_bpress_res_ek = third_attempt_bpress_res
    #                 cur_comp.third_attempt_bpress_off_ek = third_attempt_bpress_off
    #                 cur_comp.first_attempt_dlift_res_ek = first_attempt_dlift_res
    #                 cur_comp.first_attempt_dlift_off_ek = first_attempt_dlift_off
    #                 cur_comp.second_attempt_dlift_res_ek = second_attempt_dlift_res
    #                 cur_comp.second_attempt_dlift_off_ek = second_attempt_dlift_off
    #                 cur_comp.third_attempt_dlift_res_ek = third_attempt_dlift_res
    #                 cur_comp.third_attempt_dlift_off_ek = third_attempt_dlift_off
    #                 cur_comp.competitor_translation = comp_transl
    #                 cur_comp.save()
    #                 return JsonResponse({"pil_h_squat": pil_h_squat, "pil_h_bpress": pil_h_bpress, "first_attempt_squat_res": first_attempt_squat_res,
    #                                      "first_attempt_squat_off": first_attempt_squat_off,
    #                                      "second_attempt_squat_res": second_attempt_squat_res,
    #                                      "second_attempt_squat_off": second_attempt_squat_off,
    #                                      "third_attempt_squat_res": third_attempt_squat_res,
    #                                      "third_attempt_squat_off": third_attempt_squat_off,
    #                                      "first_attempt_bpress_res": first_attempt_bpress_res,
    #                                      "first_attempt_bpress_off": first_attempt_bpress_off,
    #                                      "second_attempt_bpress_res": second_attempt_bpress_res,
    #                                      "second_attempt_bpress_off": second_attempt_bpress_off,
    #                                      "third_attempt_bpress_res": third_attempt_bpress_res,
    #                                      "third_attempt_bpress_off": third_attempt_bpress_off,
    #                                      "first_attempt_dlift_res": first_attempt_dlift_res,
    #                                      "first_attempt_dlift_off": first_attempt_dlift_off,
    #                                      "second_attempt_dlift_res": second_attempt_dlift_res,
    #                                      "second_attempt_dlift_off": second_attempt_dlift_off,
    #                                      "third_attempt_dlift_res": third_attempt_dlift_res,
    #                                      "third_attempt_dlift_off": third_attempt_dlift_off,
    #                                      "comp_transl": comp_transl}, status=200)
    #             if cur_transl_val == "Жим 3 экип" and third_attempt_bpress_off:
    #                 cur_comp.pillar_height_squat = pil_h_squat
    #                 cur_comp.pillar_height_bpress = pil_h_bpress
    #                 cur_comp.first_attempt_squat_res_ek = first_attempt_squat_res
    #                 cur_comp.first_attempt_squat_off_ek = first_attempt_squat_off
    #                 cur_comp.second_attempt_squat_res_ek = second_attempt_squat_res
    #                 cur_comp.second_attempt_squat_off_ek = second_attempt_squat_off
    #                 cur_comp.third_attempt_squat_res_ek = third_attempt_squat_res
    #                 cur_comp.third_attempt_squat_off_ek = third_attempt_squat_off
    #                 cur_comp.first_attempt_bpress_res_ek = first_attempt_bpress_res
    #                 cur_comp.first_attempt_bpress_off_ek = first_attempt_bpress_off
    #                 cur_comp.second_attempt_bpress_res_ek = second_attempt_bpress_res
    #                 cur_comp.second_attempt_bpress_off_ek = second_attempt_bpress_off
    #                 cur_comp.third_attempt_bpress_res_ek = third_attempt_bpress_res
    #                 cur_comp.third_attempt_bpress_off_ek = third_attempt_bpress_off
    #                 cur_comp.first_attempt_dlift_res_ek = first_attempt_dlift_res
    #                 cur_comp.first_attempt_dlift_off_ek = first_attempt_dlift_off
    #                 cur_comp.second_attempt_dlift_res_ek = second_attempt_dlift_res
    #                 cur_comp.second_attempt_dlift_off_ek = second_attempt_dlift_off
    #                 cur_comp.third_attempt_dlift_res_ek = third_attempt_dlift_res
    #                 cur_comp.third_attempt_dlift_off_ek = third_attempt_dlift_off
    #                 cur_comp.competitor_translation = comp_transl
    #                 cur_comp.save()
    #                 return JsonResponse({"pil_h_squat": pil_h_squat, "pil_h_bpress": pil_h_bpress, "first_attempt_squat_res": first_attempt_squat_res,
    #                                      "first_attempt_squat_off": first_attempt_squat_off,
    #                                      "second_attempt_squat_res": second_attempt_squat_res,
    #                                      "second_attempt_squat_off": second_attempt_squat_off,
    #                                      "third_attempt_squat_res": third_attempt_squat_res,
    #                                      "third_attempt_squat_off": third_attempt_squat_off,
    #                                      "first_attempt_bpress_res": first_attempt_bpress_res,
    #                                      "first_attempt_bpress_off": first_attempt_bpress_off,
    #                                      "second_attempt_bpress_res": second_attempt_bpress_res,
    #                                      "second_attempt_bpress_off": second_attempt_bpress_off,
    #                                      "third_attempt_bpress_res": third_attempt_bpress_res,
    #                                      "third_attempt_bpress_off": third_attempt_bpress_off,
    #                                      "first_attempt_dlift_res": first_attempt_dlift_res,
    #                                      "first_attempt_dlift_off": first_attempt_dlift_off,
    #                                      "second_attempt_dlift_res": second_attempt_dlift_res,
    #                                      "second_attempt_dlift_off": second_attempt_dlift_off,
    #                                      "third_attempt_dlift_res": third_attempt_dlift_res,
    #                                      "third_attempt_dlift_off": third_attempt_dlift_off,
    #                                      "comp_transl": comp_transl}, status=200)
    #             if cur_transl_val == "Тяга 1 экип" and first_attempt_dlift_off:
    #                 cur_comp.pillar_height_squat = pil_h_squat
    #                 cur_comp.pillar_height_bpress = pil_h_bpress
    #                 cur_comp.first_attempt_squat_res_ek = first_attempt_squat_res
    #                 cur_comp.first_attempt_squat_off_ek = first_attempt_squat_off
    #                 cur_comp.second_attempt_squat_res_ek = second_attempt_squat_res
    #                 cur_comp.second_attempt_squat_off_ek = second_attempt_squat_off
    #                 cur_comp.third_attempt_squat_res_ek = third_attempt_squat_res
    #                 cur_comp.third_attempt_squat_off_ek = third_attempt_squat_off
    #                 cur_comp.first_attempt_bpress_res_ek = first_attempt_bpress_res
    #                 cur_comp.first_attempt_bpress_off_ek = first_attempt_bpress_off
    #                 cur_comp.second_attempt_bpress_res_ek = second_attempt_bpress_res
    #                 cur_comp.second_attempt_bpress_off_ek = second_attempt_bpress_off
    #                 cur_comp.third_attempt_bpress_res_ek = third_attempt_bpress_res
    #                 cur_comp.third_attempt_bpress_off_ek = third_attempt_bpress_off
    #                 cur_comp.first_attempt_dlift_res_ek = first_attempt_dlift_res
    #                 cur_comp.first_attempt_dlift_off_ek = first_attempt_dlift_off
    #                 cur_comp.second_attempt_dlift_res_ek = second_attempt_dlift_res
    #                 cur_comp.second_attempt_dlift_off_ek = second_attempt_dlift_off
    #                 cur_comp.third_attempt_dlift_res_ek = third_attempt_dlift_res
    #                 cur_comp.third_attempt_dlift_off_ek = third_attempt_dlift_off
    #                 cur_comp.competitor_translation = comp_transl
    #                 cur_comp.save()
    #                 return JsonResponse({"pil_h_squat": pil_h_squat, "pil_h_bpress": pil_h_bpress, "first_attempt_squat_res": first_attempt_squat_res,
    #                                      "first_attempt_squat_off": first_attempt_squat_off,
    #                                      "second_attempt_squat_res": second_attempt_squat_res,
    #                                      "second_attempt_squat_off": second_attempt_squat_off,
    #                                      "third_attempt_squat_res": third_attempt_squat_res,
    #                                      "third_attempt_squat_off": third_attempt_squat_off,
    #                                      "first_attempt_bpress_res": first_attempt_bpress_res,
    #                                      "first_attempt_bpress_off": first_attempt_bpress_off,
    #                                      "second_attempt_bpress_res": second_attempt_bpress_res,
    #                                      "second_attempt_bpress_off": second_attempt_bpress_off,
    #                                      "third_attempt_bpress_res": third_attempt_bpress_res,
    #                                      "third_attempt_bpress_off": third_attempt_bpress_off,
    #                                      "first_attempt_dlift_res": first_attempt_dlift_res,
    #                                      "first_attempt_dlift_off": first_attempt_dlift_off,
    #                                      "second_attempt_dlift_res": second_attempt_dlift_res,
    #                                      "second_attempt_dlift_off": second_attempt_dlift_off,
    #                                      "third_attempt_dlift_res": third_attempt_dlift_res,
    #                                      "third_attempt_dlift_off": third_attempt_dlift_off,
    #                                      "comp_transl": comp_transl}, status=200)
    #             if cur_transl_val == "Тяга 2 экип" and second_attempt_dlift_off:
    #                 cur_comp.pillar_height_squat = pil_h_squat
    #                 cur_comp.pillar_height_bpress = pil_h_bpress
    #                 cur_comp.first_attempt_squat_res_ek = first_attempt_squat_res
    #                 cur_comp.first_attempt_squat_off_ek = first_attempt_squat_off
    #                 cur_comp.second_attempt_squat_res_ek = second_attempt_squat_res
    #                 cur_comp.second_attempt_squat_off_ek = second_attempt_squat_off
    #                 cur_comp.third_attempt_squat_res_ek = third_attempt_squat_res
    #                 cur_comp.third_attempt_squat_off_ek = third_attempt_squat_off
    #                 cur_comp.first_attempt_bpress_res_ek = first_attempt_bpress_res
    #                 cur_comp.first_attempt_bpress_off_ek = first_attempt_bpress_off
    #                 cur_comp.second_attempt_bpress_res_ek = second_attempt_bpress_res
    #                 cur_comp.second_attempt_bpress_off_ek = second_attempt_bpress_off
    #                 cur_comp.third_attempt_bpress_res_ek = third_attempt_bpress_res
    #                 cur_comp.third_attempt_bpress_off_ek = third_attempt_bpress_off
    #                 cur_comp.first_attempt_dlift_res_ek = first_attempt_dlift_res
    #                 cur_comp.first_attempt_dlift_off_ek = first_attempt_dlift_off
    #                 cur_comp.second_attempt_dlift_res_ek = second_attempt_dlift_res
    #                 cur_comp.second_attempt_dlift_off_ek = second_attempt_dlift_off
    #                 cur_comp.third_attempt_dlift_res_ek = third_attempt_dlift_res
    #                 cur_comp.third_attempt_dlift_off_ek = third_attempt_dlift_off
    #                 cur_comp.competitor_translation = comp_transl
    #                 cur_comp.save()
    #                 return JsonResponse({"pil_h_squat": pil_h_squat, "pil_h_bpress": pil_h_bpress, "first_attempt_squat_res": first_attempt_squat_res,
    #                                      "first_attempt_squat_off": first_attempt_squat_off,
    #                                      "second_attempt_squat_res": second_attempt_squat_res,
    #                                      "second_attempt_squat_off": second_attempt_squat_off,
    #                                      "third_attempt_squat_res": third_attempt_squat_res,
    #                                      "third_attempt_squat_off": third_attempt_squat_off,
    #                                      "first_attempt_bpress_res": first_attempt_bpress_res,
    #                                      "first_attempt_bpress_off": first_attempt_bpress_off,
    #                                      "second_attempt_bpress_res": second_attempt_bpress_res,
    #                                      "second_attempt_bpress_off": second_attempt_bpress_off,
    #                                      "third_attempt_bpress_res": third_attempt_bpress_res,
    #                                      "third_attempt_bpress_off": third_attempt_bpress_off,
    #                                      "first_attempt_dlift_res": first_attempt_dlift_res,
    #                                      "first_attempt_dlift_off": first_attempt_dlift_off,
    #                                      "second_attempt_dlift_res": second_attempt_dlift_res,
    #                                      "second_attempt_dlift_off": second_attempt_dlift_off,
    #                                      "third_attempt_dlift_res": third_attempt_dlift_res,
    #                                      "third_attempt_dlift_off": third_attempt_dlift_off,
    #                                      "comp_transl": comp_transl}, status=200)
    #             if cur_transl_val == "Тяга 3 экип" and third_attempt_dlift_off:
    #                 cur_comp.pillar_height_squat = pil_h_squat
    #                 cur_comp.pillar_height_bpress = pil_h_bpress
    #                 cur_comp.first_attempt_squat_res_ek = first_attempt_squat_res
    #                 cur_comp.first_attempt_squat_off_ek = first_attempt_squat_off
    #                 cur_comp.second_attempt_squat_res_ek = second_attempt_squat_res
    #                 cur_comp.second_attempt_squat_off_ek = second_attempt_squat_off
    #                 cur_comp.third_attempt_squat_res_ek = third_attempt_squat_res
    #                 cur_comp.third_attempt_squat_off_ek = third_attempt_squat_off
    #                 cur_comp.first_attempt_bpress_res_ek = first_attempt_bpress_res
    #                 cur_comp.first_attempt_bpress_off_ek = first_attempt_bpress_off
    #                 cur_comp.second_attempt_bpress_res_ek = second_attempt_bpress_res
    #                 cur_comp.second_attempt_bpress_off_ek = second_attempt_bpress_off
    #                 cur_comp.third_attempt_bpress_res_ek = third_attempt_bpress_res
    #                 cur_comp.third_attempt_bpress_off_ek = third_attempt_bpress_off
    #                 cur_comp.first_attempt_dlift_res_ek = first_attempt_dlift_res
    #                 cur_comp.first_attempt_dlift_off_ek = first_attempt_dlift_off
    #                 cur_comp.second_attempt_dlift_res_ek = second_attempt_dlift_res
    #                 cur_comp.second_attempt_dlift_off_ek = second_attempt_dlift_off
    #                 cur_comp.third_attempt_dlift_res_ek = third_attempt_dlift_res
    #                 cur_comp.third_attempt_dlift_off_ek = third_attempt_dlift_off
    #                 cur_comp.competitor_translation = comp_transl
    #                 cur_comp.save()
    #                 return JsonResponse({"pil_h_squat": pil_h_squat, "pil_h_bpress": pil_h_bpress, "first_attempt_squat_res": first_attempt_squat_res,
    #                                      "first_attempt_squat_off": first_attempt_squat_off,
    #                                      "second_attempt_squat_res": second_attempt_squat_res,
    #                                      "second_attempt_squat_off": second_attempt_squat_off,
    #                                      "third_attempt_squat_res": third_attempt_squat_res,
    #                                      "third_attempt_squat_off": third_attempt_squat_off,
    #                                      "first_attempt_bpress_res": first_attempt_bpress_res,
    #                                      "first_attempt_bpress_off": first_attempt_bpress_off,
    #                                      "second_attempt_bpress_res": second_attempt_bpress_res,
    #                                      "second_attempt_bpress_off": second_attempt_bpress_off,
    #                                      "third_attempt_bpress_res": third_attempt_bpress_res,
    #                                      "third_attempt_bpress_off": third_attempt_bpress_off,
    #                                      "first_attempt_dlift_res": first_attempt_dlift_res,
    #                                      "first_attempt_dlift_off": first_attempt_dlift_off,
    #                                      "second_attempt_dlift_res": second_attempt_dlift_res,
    #                                      "second_attempt_dlift_off": second_attempt_dlift_off,
    #                                      "third_attempt_dlift_res": third_attempt_dlift_res,
    #                                      "third_attempt_dlift_off": third_attempt_dlift_off,
    #                                      "comp_transl": comp_transl}, status=200)
    #     if comp_transl:
    #         cur_comp.pillar_height_squat = pil_h_squat
    #         cur_comp.pillar_height_bpress = pil_h_bpress
    #         cur_comp.first_attempt_squat_res_ek = first_attempt_squat_res
    #         cur_comp.first_attempt_squat_off_ek = first_attempt_squat_off
    #         cur_comp.second_attempt_squat_res_ek = second_attempt_squat_res
    #         cur_comp.second_attempt_squat_off_ek = second_attempt_squat_off
    #         cur_comp.third_attempt_squat_res_ek = third_attempt_squat_res
    #         cur_comp.third_attempt_squat_off_ek = third_attempt_squat_off
    #         cur_comp.first_attempt_bpress_res_ek = first_attempt_bpress_res
    #         cur_comp.first_attempt_bpress_off_ek = first_attempt_bpress_off
    #         cur_comp.second_attempt_bpress_res_ek = second_attempt_bpress_res
    #         cur_comp.second_attempt_bpress_off_ek = second_attempt_bpress_off
    #         cur_comp.third_attempt_bpress_res_ek = third_attempt_bpress_res
    #         cur_comp.third_attempt_bpress_off_ek = third_attempt_bpress_off
    #         cur_comp.first_attempt_dlift_res_ek = first_attempt_dlift_res
    #         cur_comp.first_attempt_dlift_off_ek = first_attempt_dlift_off
    #         cur_comp.second_attempt_dlift_res_ek = second_attempt_dlift_res
    #         cur_comp.second_attempt_dlift_off_ek = second_attempt_dlift_off
    #         cur_comp.third_attempt_dlift_res_ek = third_attempt_dlift_res
    #         cur_comp.third_attempt_dlift_off_ek = third_attempt_dlift_off
    #         cur_comp.competitor_translation = comp_transl
    #         cur_comp.save()
    #         return JsonResponse(
    #             {"pil_h_squat": pil_h_squat, "pil_h_bpress": pil_h_bpress, "first_attempt_squat_res": first_attempt_squat_res,
    #              "first_attempt_squat_off": first_attempt_squat_off,
    #              "second_attempt_squat_res": second_attempt_squat_res,
    #              "second_attempt_squat_off": second_attempt_squat_off,
    #              "third_attempt_squat_res": third_attempt_squat_res,
    #              "third_attempt_squat_off": third_attempt_squat_off,
    #              "first_attempt_bpress_res": first_attempt_bpress_res,
    #              "first_attempt_bpress_off": first_attempt_bpress_off,
    #              "second_attempt_bpress_res": second_attempt_bpress_res,
    #              "second_attempt_bpress_off": second_attempt_bpress_off,
    #              "third_attempt_bpress_res": third_attempt_bpress_res,
    #              "third_attempt_bpress_off": third_attempt_bpress_off,
    #              "first_attempt_dlift_res": first_attempt_dlift_res,
    #              "first_attempt_dlift_off": first_attempt_dlift_off,
    #              "second_attempt_dlift_res": second_attempt_dlift_res,
    #              "second_attempt_dlift_off": second_attempt_dlift_off,
    #              "third_attempt_dlift_res": third_attempt_dlift_res,
    #              "third_attempt_dlift_off": third_attempt_dlift_off,
    #              "comp_transl": comp_transl}, status=200)
    #     else:
    #         cur_comp.pillar_height_squat = pil_h_squat
    #         cur_comp.pillar_height_bpress = pil_h_bpress
    #         cur_comp.first_attempt_squat_res_ek = first_attempt_squat_res
    #         cur_comp.first_attempt_squat_off_ek = first_attempt_squat_off
    #         cur_comp.second_attempt_squat_res_ek = second_attempt_squat_res
    #         cur_comp.second_attempt_squat_off_ek = second_attempt_squat_off
    #         cur_comp.third_attempt_squat_res_ek = third_attempt_squat_res
    #         cur_comp.third_attempt_squat_off_ek = third_attempt_squat_off
    #         cur_comp.first_attempt_bpress_res_ek = first_attempt_bpress_res
    #         cur_comp.first_attempt_bpress_off_ek = first_attempt_bpress_off
    #         cur_comp.second_attempt_bpress_res_ek = second_attempt_bpress_res
    #         cur_comp.second_attempt_bpress_off_ek = second_attempt_bpress_off
    #         cur_comp.third_attempt_bpress_res_ek = third_attempt_bpress_res
    #         cur_comp.third_attempt_bpress_off_ek = third_attempt_bpress_off
    #         cur_comp.first_attempt_dlift_res_ek = first_attempt_dlift_res
    #         cur_comp.first_attempt_dlift_off_ek = first_attempt_dlift_off
    #         cur_comp.second_attempt_dlift_res_ek = second_attempt_dlift_res
    #         cur_comp.second_attempt_dlift_off_ek = second_attempt_dlift_off
    #         cur_comp.third_attempt_dlift_res_ek = third_attempt_dlift_res
    #         cur_comp.third_attempt_dlift_off_ek = third_attempt_dlift_off
    #         cur_comp.competitor_translation = None
    #         cur_comp.save()
    #         return JsonResponse(
    #             {"pil_h_squat": pil_h_squat, "pil_h_bpress": pil_h_bpress, "first_attempt_squat_res": first_attempt_squat_res,
    #              "first_attempt_squat_off": first_attempt_squat_off,
    #              "second_attempt_squat_res": second_attempt_squat_res,
    #              "second_attempt_squat_off": second_attempt_squat_off,
    #              "third_attempt_squat_res": third_attempt_squat_res,
    #              "third_attempt_squat_off": third_attempt_squat_off,
    #              "first_attempt_bpress_res": first_attempt_bpress_res,
    #              "first_attempt_bpress_off": first_attempt_bpress_off,
    #              "second_attempt_bpress_res": second_attempt_bpress_res,
    #              "second_attempt_bpress_off": second_attempt_bpress_off,
    #              "third_attempt_bpress_res": third_attempt_bpress_res,
    #              "third_attempt_bpress_off": third_attempt_bpress_off,
    #              "first_attempt_dlift_res": first_attempt_dlift_res,
    #              "first_attempt_dlift_off": first_attempt_dlift_off,
    #              "second_attempt_dlift_res": second_attempt_dlift_res,
    #              "second_attempt_dlift_off": second_attempt_dlift_off,
    #              "third_attempt_dlift_res": third_attempt_dlift_res,
    #              "third_attempt_dlift_off": third_attempt_dlift_off,
    #              "comp_transl": comp_transl}, status=200)
    # if 'secr-page_form-btn-bpek' in request.POST:
    #     comp_pk = request.POST.get('secr-page_form-btn-bpek')
    #     pil_h_bpress = request.POST.get('secr-page_form-pil_h_bpress')
    #     if pil_h_bpress:
    #         pil_h_bpress = pil_h_bpress
    #     else:
    #         pil_h_bpress = None
    #     first_attempt_bpress_res = request.POST.get('secr-page_form-fat-bpress-res')
    #     if first_attempt_bpress_res:
    #         first_attempt_bpress_res = Decimal(first_attempt_bpress_res)
    #     else:
    #         first_attempt_bpress_res = None
    #     first_attempt_bpress_off = request.POST.get('secr-page_form-fat-bpress-off')
    #     second_attempt_bpress_res = request.POST.get('secr-page_form-sat-bpress-res')
    #     if second_attempt_bpress_res:
    #         second_attempt_bpress_res = Decimal(second_attempt_bpress_res)
    #     else:
    #         second_attempt_bpress_res = None
    #     second_attempt_bpress_off = request.POST.get('secr-page_form-sat-bpress-off')
    #     third_attempt_bpress_res = request.POST.get('secr-page_form-tat-bpress-res')
    #     if third_attempt_bpress_res:
    #         third_attempt_bpress_res = Decimal(third_attempt_bpress_res)
    #     else:
    #         third_attempt_bpress_res = None
    #     third_attempt_bpress_off = request.POST.get('secr-page_form-tat-bpress-off')
    #     cur_comp = CompetitionProtocols.objects.get(pk=comp_pk)
    #     comp_transl = request.POST.get('secr-page_chckbx')
    #     cur_transl = CompetitionProtocols.objects.exclude(competitor_translation=None)
    #     if comp_transl and cur_transl:
    #         if comp_transl != cur_transl[0].competitor_translation:
    #             return JsonResponse(
    #                 {"error": "Нельзя запускать две трансляции одновременно. Сначала закончите существующую."},
    #                 status=302)
    #     if cur_transl:
    #         cur_transl_val = cur_transl.values_list('competitor_translation', flat=True)[0]
    #         cur_transl_pk = cur_transl.values_list('id', flat=True)[0]
    #         if comp_pk == str(cur_transl_pk):
    #             if cur_transl_val == "Жим 1 экип" and first_attempt_bpress_off:
    #                 cur_comp.pillar_height_bpress = pil_h_bpress
    #                 cur_comp.first_attempt_bpress_res_ek = first_attempt_bpress_res
    #                 cur_comp.first_attempt_bpress_off_ek = first_attempt_bpress_off
    #                 cur_comp.second_attempt_bpress_res_ek = second_attempt_bpress_res
    #                 cur_comp.second_attempt_bpress_off_ek = second_attempt_bpress_off
    #                 cur_comp.third_attempt_bpress_res_ek = third_attempt_bpress_res
    #                 cur_comp.third_attempt_bpress_off_ek = third_attempt_bpress_off
    #                 cur_comp.competitor_translation = comp_transl
    #                 cur_comp.save()
    #                 return JsonResponse({"pil_h_bpress": pil_h_bpress, "first_attempt_bpress_res": first_attempt_bpress_res,
    #                                      "first_attempt_bpress_off": first_attempt_bpress_off,
    #                                      "second_attempt_bpress_res": second_attempt_bpress_res,
    #                                      "second_attempt_bpress_off": second_attempt_bpress_off,
    #                                      "third_attempt_bpress_res": third_attempt_bpress_res,
    #                                      "third_attempt_bpress_off": third_attempt_bpress_off,
    #                                      "comp_transl": comp_transl}, status=200)
    #             if cur_transl_val == "Жим 2 экип" and second_attempt_bpress_off:
    #                 cur_comp.pillar_height_bpress = pil_h_bpress
    #                 cur_comp.first_attempt_bpress_res_ek = first_attempt_bpress_res
    #                 cur_comp.first_attempt_bpress_off_ek = first_attempt_bpress_off
    #                 cur_comp.second_attempt_bpress_res_ek = second_attempt_bpress_res
    #                 cur_comp.second_attempt_bpress_off_ek = second_attempt_bpress_off
    #                 cur_comp.third_attempt_bpress_res_ek = third_attempt_bpress_res
    #                 cur_comp.third_attempt_bpress_off_ek = third_attempt_bpress_off
    #                 cur_comp.competitor_translation = comp_transl
    #                 cur_comp.save()
    #                 return JsonResponse({"pil_h_bpress": pil_h_bpress, "first_attempt_bpress_res": first_attempt_bpress_res,
    #                                      "first_attempt_bpress_off": first_attempt_bpress_off,
    #                                      "second_attempt_bpress_res": second_attempt_bpress_res,
    #                                      "second_attempt_bpress_off": second_attempt_bpress_off,
    #                                      "third_attempt_bpress_res": third_attempt_bpress_res,
    #                                      "third_attempt_bpress_off": third_attempt_bpress_off,
    #                                      "comp_transl": comp_transl}, status=200)
    #             if cur_transl_val == "Жим 3 экип" and third_attempt_bpress_off:
    #                 cur_comp.pillar_height_bpress = pil_h_bpress
    #                 cur_comp.first_attempt_bpress_res_ek = first_attempt_bpress_res
    #                 cur_comp.first_attempt_bpress_off_ek = first_attempt_bpress_off
    #                 cur_comp.second_attempt_bpress_res_ek = second_attempt_bpress_res
    #                 cur_comp.second_attempt_bpress_off_ek = second_attempt_bpress_off
    #                 cur_comp.third_attempt_bpress_res_ek = third_attempt_bpress_res
    #                 cur_comp.third_attempt_bpress_off_ek = third_attempt_bpress_off
    #                 cur_comp.competitor_translation = comp_transl
    #                 cur_comp.save()
    #                 return JsonResponse({"pil_h_bpress": pil_h_bpress, "first_attempt_bpress_res": first_attempt_bpress_res,
    #                                      "first_attempt_bpress_off": first_attempt_bpress_off,
    #                                      "second_attempt_bpress_res": second_attempt_bpress_res,
    #                                      "second_attempt_bpress_off": second_attempt_bpress_off,
    #                                      "third_attempt_bpress_res": third_attempt_bpress_res,
    #                                      "third_attempt_bpress_off": third_attempt_bpress_off,
    #                                      "comp_transl": comp_transl}, status=200)
    #     if comp_transl:
    #         cur_comp.pillar_height_bpress = pil_h_bpress
    #         cur_comp.first_attempt_bpress_res_ek = first_attempt_bpress_res
    #         cur_comp.first_attempt_bpress_off_ek = first_attempt_bpress_off
    #         cur_comp.second_attempt_bpress_res_ek = second_attempt_bpress_res
    #         cur_comp.second_attempt_bpress_off_ek = second_attempt_bpress_off
    #         cur_comp.third_attempt_bpress_res_ek = third_attempt_bpress_res
    #         cur_comp.third_attempt_bpress_off_ek = third_attempt_bpress_off
    #         cur_comp.competitor_translation = comp_transl
    #         cur_comp.save()
    #         return JsonResponse(
    #             {"pil_h_bpress": pil_h_bpress, "first_attempt_bpress_res": first_attempt_bpress_res,
    #              "first_attempt_bpress_off": first_attempt_bpress_off,
    #              "second_attempt_bpress_res": second_attempt_bpress_res,
    #              "second_attempt_bpress_off": second_attempt_bpress_off,
    #              "third_attempt_bpress_res": third_attempt_bpress_res,
    #              "third_attempt_bpress_off": third_attempt_bpress_off,
    #              "comp_transl": comp_transl}, status=200)
    #     else:
    #         cur_comp.pillar_height_bpress = pil_h_bpress
    #         cur_comp.first_attempt_bpress_res_ek = first_attempt_bpress_res
    #         cur_comp.first_attempt_bpress_off_ek = first_attempt_bpress_off
    #         cur_comp.second_attempt_bpress_res_ek = second_attempt_bpress_res
    #         cur_comp.second_attempt_bpress_off_ek = second_attempt_bpress_off
    #         cur_comp.third_attempt_bpress_res_ek = third_attempt_bpress_res
    #         cur_comp.third_attempt_bpress_off_ek = third_attempt_bpress_off
    #         cur_comp.competitor_translation = None
    #         cur_comp.save()
    #         return JsonResponse(
    #             {"pil_h_bpress": pil_h_bpress, "first_attempt_bpress_res": first_attempt_bpress_res,
    #              "first_attempt_bpress_off": first_attempt_bpress_off,
    #              "second_attempt_bpress_res": second_attempt_bpress_res,
    #              "second_attempt_bpress_off": second_attempt_bpress_off,
    #              "third_attempt_bpress_res": third_attempt_bpress_res,
    #              "third_attempt_bpress_off": third_attempt_bpress_off,
    #              "comp_transl": comp_transl}, status=200)
    #
    #     # print(first_attempt_squat_res)
    #     # print(first_attempt_squat_off)
    #     # print(second_attempt_squat_res)
    #     # print(second_attempt_squat_off)
    #     # print(third_attempt_squat_res)
    #     # print(third_attempt_squat_off)
    #     # print(first_attempt_bpress_res)
    #     # print(first_attempt_bpress_off)
    #     # print(second_attempt_bpress_res)
    #     # print(second_attempt_bpress_off)
    #     # print(third_attempt_bpress_res)
    #     # print(third_attempt_bpress_off)
    #     # print(first_attempt_dlift_res)
    #     # print(first_attempt_dlift_off)
    #     # print(second_attempt_dlift_res)
    #     # print(second_attempt_dlift_off)
    #     # print(third_attempt_dlift_res)
    #     # print(third_attempt_dlift_off)
    #     # print(CompetitionProtocols.objects.get(pk=comp_pk))
    #
    #     # print(CompetitionProtocols.objects.exclude(competitor_translation=None))
    return render(request, 'pwrlftmain/secretary_page.html', response_data)


@login_required
def scoreboard_page(request, competition_slug):
    cur_competition = Competitions.objects.get(competition_slug=competition_slug)
    competitor_exercise = ''
    competitor_exercise_attempt = ''
    competitor_fi = ''
    competitor_weight_cat = ''
    competitor_ordered_weight = ''
    competitor_ordered_weight_offset = ''
    cur_transl_pk = ''
    comp_first_weight = ''
    comp_first_off = ''
    comp_sec_weight = ''
    comp_sec_off = ''
    response_data = {
        'cur_competition': cur_competition,
    }
    if 'score-form_btn' in request.POST:
        cur_transl = CompetitionProtocols.objects.exclude(competitor_translation=None)
        if cur_transl:
            cur_transl_val = cur_transl.values_list('competitor_translation', flat=True)[0]
            cur_transl_pk = cur_transl.values_list('id', flat=True)[0]
            competitor_transl = CompetitionProtocols.objects.get(pk=cur_transl_pk)
            competitor_exercise = cur_transl_val.split(' ')[0].lower()
            competitor_exercise_attempt = f"{cur_transl_val.split(' ')[1]} подход"
            competitor_fi = f"{competitor_transl.competitor.surname_comp.capitalize()} {competitor_transl.competitor.name_comp.capitalize()}"
            competitor_weight_cat = competitor_transl.competitor.weight_cat
            if cur_transl_val == "Присед 1 класс":
                competitor_ordered_weight = competitor_transl.first_attempt_squat_res
                competitor_ordered_weight_offset = competitor_transl.first_attempt_squat_off
            if cur_transl_val == "Присед 2 класс":
                competitor_ordered_weight = competitor_transl.second_attempt_squat_res
                competitor_ordered_weight_offset = competitor_transl.second_attempt_squat_off
                comp_first_weight = competitor_transl.first_attempt_squat_res
                comp_first_off = competitor_transl.first_attempt_squat_off
            if cur_transl_val == "Присед 3 класс":
                competitor_ordered_weight = competitor_transl.third_attempt_squat_res
                competitor_ordered_weight_offset = competitor_transl.third_attempt_squat_off
                comp_first_weight = competitor_transl.first_attempt_squat_res
                comp_first_off = competitor_transl.first_attempt_squat_off
                comp_sec_weight = competitor_transl.second_attempt_squat_res
                comp_sec_off = competitor_transl.second_attempt_squat_off
            if cur_transl_val == "Жим 1 класс":
                competitor_ordered_weight = competitor_transl.first_attempt_bpress_res
                competitor_ordered_weight_offset = competitor_transl.first_attempt_bpress_off
            if cur_transl_val == "Жим 2 класс":
                competitor_ordered_weight = competitor_transl.second_attempt_bpress_res
                competitor_ordered_weight_offset = competitor_transl.second_attempt_bpress_off
                comp_first_weight = competitor_transl.first_attempt_bpress_res
                comp_first_off = competitor_transl.first_attempt_bpress_off
            if cur_transl_val == "Жим 3 класс":
                competitor_ordered_weight = competitor_transl.third_attempt_bpress_res
                competitor_ordered_weight_offset = competitor_transl.third_attempt_bpress_off
                comp_first_weight = competitor_transl.first_attempt_bpress_res
                comp_first_off = competitor_transl.first_attempt_bpress_off
                comp_sec_weight = competitor_transl.second_attempt_bpress_res
                comp_sec_off = competitor_transl.second_attempt_bpress_off
            if cur_transl_val == "Тяга 1 класс":
                competitor_ordered_weight = competitor_transl.first_attempt_dlift_res
                competitor_ordered_weight_offset = competitor_transl.first_attempt_dlift_off
            if cur_transl_val == "Тяга 2 класс":
                competitor_ordered_weight = competitor_transl.second_attempt_dlift_res
                competitor_ordered_weight_offset = competitor_transl.second_attempt_dlift_off
                comp_first_weight = competitor_transl.first_attempt_dlift_res
                comp_first_off = competitor_transl.first_attempt_dlift_off
            if cur_transl_val == "Тяга 3 класс":
                competitor_ordered_weight = competitor_transl.third_attempt_dlift_res
                competitor_ordered_weight_offset = competitor_transl.third_attempt_dlift_off
                comp_first_weight = competitor_transl.first_attempt_dlift_res
                comp_first_off = competitor_transl.first_attempt_dlift_off
                comp_sec_weight = competitor_transl.second_attempt_dlift_res
                comp_sec_off = competitor_transl.second_attempt_dlift_off
            if cur_transl_val == "Присед 1 экип":
                competitor_ordered_weight = competitor_transl.first_attempt_squat_res_ek
                competitor_ordered_weight_offset = competitor_transl.first_attempt_squat_off_ek
            if cur_transl_val == "Присед 2 экип":
                competitor_ordered_weight = competitor_transl.second_attempt_squat_res_ek
                competitor_ordered_weight_offset = competitor_transl.second_attempt_squat_off_ek
                comp_first_weight = competitor_transl.first_attempt_squat_res_ek
                comp_first_off = competitor_transl.first_attempt_squat_off_ek
            if cur_transl_val == "Присед 3 экип":
                competitor_ordered_weight = competitor_transl.third_attempt_squat_res_ek
                competitor_ordered_weight_offset = competitor_transl.third_attempt_squat_off_ek
                comp_first_weight = competitor_transl.first_attempt_squat_res_ek
                comp_first_off = competitor_transl.first_attempt_squat_off_ek
                comp_sec_weight = competitor_transl.second_attempt_squat_res_ek
                comp_sec_off = competitor_transl.second_attempt_squat_off_ek
            if cur_transl_val == "Жим 1 экип":
                competitor_ordered_weight = competitor_transl.first_attempt_bpress_res_ek
                competitor_ordered_weight_offset = competitor_transl.first_attempt_bpress_off_ek
            if cur_transl_val == "Жим 2 экип":
                competitor_ordered_weight = competitor_transl.second_attempt_bpress_res_ek
                competitor_ordered_weight_offset = competitor_transl.second_attempt_bpress_off_ek
                comp_first_weight = competitor_transl.first_attempt_bpress_res_ek
                comp_first_off = competitor_transl.first_attempt_bpress_off_ek
            if cur_transl_val == "Жим 3 экип":
                competitor_ordered_weight = competitor_transl.third_attempt_bpress_res_ek
                competitor_ordered_weight_offset = competitor_transl.third_attempt_bpress_off_ek
                comp_first_weight = competitor_transl.first_attempt_bpress_res_ek
                comp_first_off = competitor_transl.first_attempt_bpress_off_ek
                comp_sec_weight = competitor_transl.second_attempt_bpress_res_ek
                comp_sec_off = competitor_transl.second_attempt_bpress_off_ek
            if cur_transl_val == "Тяга 1 экип":
                competitor_ordered_weight = competitor_transl.first_attempt_dlift_res_ek
                competitor_ordered_weight_offset = competitor_transl.first_attempt_dlift_off_ek
            if cur_transl_val == "Тяга 2 экип":
                competitor_ordered_weight = competitor_transl.second_attempt_dlift_res_ek
                competitor_ordered_weight_offset = competitor_transl.second_attempt_dlift_off_ek
                comp_first_weight = competitor_transl.first_attempt_dlift_res_ek
                comp_first_off = competitor_transl.first_attempt_dlift_off_ek
            if cur_transl_val == "Тяга 3 экип":
                competitor_ordered_weight = competitor_transl.third_attempt_dlift_res_ek
                competitor_ordered_weight_offset = competitor_transl.third_attempt_dlift_off_ek
                comp_first_weight = competitor_transl.first_attempt_dlift_res_ek
                comp_first_off = competitor_transl.first_attempt_dlift_off_ek
                comp_sec_weight = competitor_transl.second_attempt_dlift_res_ek
                comp_sec_off = competitor_transl.second_attempt_dlift_off_ek
            if not competitor_ordered_weight_offset:
                competitor_ordered_weight_offset = ""
            return JsonResponse(
                {"competitor_exercise": competitor_exercise, "competitor_exercise_attempt": competitor_exercise_attempt,
                 "competitor_fi": competitor_fi,
                 "competitor_weight_cat": competitor_weight_cat, "competitor_ordered_weight": competitor_ordered_weight,
                 "competitor_ordered_weight_offset": competitor_ordered_weight_offset, "cur_transl_pk": cur_transl_pk, "comp_first_weight": comp_first_weight,
                 "comp_first_off": comp_first_off, "comp_sec_weight": comp_sec_weight, "comp_sec_off": comp_sec_off},
                status=200)
        else:
            cur_flow = CompetitorsFlow.objects.first().comp_flow
            cur_flow_competitors_st = CompetitionProtocols.objects.filter(competitor_stream=cur_flow, competitor__comp_title__competition_slug=competition_slug)
            competitors_stypes_st = cur_flow_competitors_st.values_list('competitor__sports_type__title', flat=True)
            competitors_stypes = []
            for csts in competitors_stypes_st:
                if csts not in competitors_stypes:
                    competitors_stypes.append(csts)
            if "Троеборье классическое" in competitors_stypes:
                cur_flow_competitors = cur_flow_competitors_st.order_by(F("first_attempt_squat_res").asc(nulls_last=True))
                # if None not in cur_flow_competitors.filter(Q(competitor__sports_type__title="Троеборье классическое") & Q(competitor__sports_type__title="Троеборье (экип.)")).values_list('third_attempt_squat_off', flat=True):
                if None not in cur_flow_competitors.filter(competitor__sports_type__title="Троеборье классическое").values_list('third_attempt_squat_off', flat=True):
                    cur_flow_competitors = cur_flow_competitors_st.order_by(F("first_attempt_bpress_res").asc(nulls_last=True))
                    # if None not in cur_flow_competitors.filter(Q(competitor__sports_type__title="Троеборье классическое") & Q(competitor__sports_type__title="Троеборье (экип.)")).values_list('third_attempt_bpress_off', flat=True):
                    if None not in cur_flow_competitors.filter(competitor__sports_type__title="Троеборье классическое").values_list('third_attempt_bpress_off', flat=True):
                        cur_flow_competitors = cur_flow_competitors_st.order_by(F("first_attempt_dlift_res").asc(nulls_last=True))
            elif "Троеборье (экип.)" in competitors_stypes:
                cur_flow_competitors = cur_flow_competitors_st.order_by(F("first_attempt_squat_res_ek").asc(nulls_last=True))
                if None not in cur_flow_competitors.filter(competitor__sports_type__title="Троеборье (экип.)").values_list('third_attempt_squat_off_ek', flat=True):
                    cur_flow_competitors = cur_flow_competitors_st.order_by(F("first_attempt_bpress_res_ek").asc(nulls_last=True))
                    if None not in cur_flow_competitors.filter(competitor__sports_type__title="Троеборье (экип.)").values_list('third_attempt_bpress_off_ek', flat=True):
                        cur_flow_competitors = cur_flow_competitors_st.order_by(F("first_attempt_dlift_res_ek").asc(nulls_last=True))
            elif "Жим лёжа (без экип.)" in competitors_stypes:
                cur_flow_competitors = cur_flow_competitors_st.order_by(F("first_attempt_bpress_res").asc(nulls_last=True))
            else:
                cur_flow_competitors = cur_flow_competitors_st.order_by(F("first_attempt_bpress_res_ek").asc(nulls_last=True))
            cur_flow_competitors_json = serializers.serialize("json", cur_flow_competitors, use_natural_foreign_keys=True)
            return JsonResponse({"cur_flow_competitors_json": cur_flow_competitors_json, }, status=200)
            # competitor_sports_types = [i.title for i in competitor_transl.competitor.sports_type.all()]
            # if "Троеборье классическое" in competitor_sports_types:
            #     if cur_transl_val == "Присед 1":
            #         competitor_ordered_weight = competitor_transl.first_attempt_squat_res
            #     if cur_transl_val == "Присед 2":
            #         competitor_ordered_weight = competitor_transl.second_attempt_squat_res
            #     if cur_transl_val == "Присед 3":
            #         competitor_ordered_weight = competitor_transl.third_attempt_squat_res
            #     if cur_transl_val == "Жим 1":
            #         competitor_ordered_weight = competitor_transl.first_attempt_bpress_res
            #     if cur_transl_val == "Жим 2":
            #         competitor_ordered_weight = competitor_transl.second_attempt_bpress_res
            #     if cur_transl_val == "Жим 3":
            #         competitor_ordered_weight = competitor_transl.third_attempt_bpress_res
            #     if cur_transl_val == "Тяга 1":
            #         competitor_ordered_weight = competitor_transl.first_attempt_dlift_res
            #     if cur_transl_val == "Тяга 2":
            #         competitor_ordered_weight = competitor_transl.second_attempt_dlift_res
            #     if cur_transl_val == "Тяга 3":
            #         competitor_ordered_weight = competitor_transl.third_attempt_dlift_res
            # if "Троеборье (экип.)" in competitor_sports_types:
            #     if cur_transl_val == "Присед 1":
            #         competitor_ordered_weight = competitor_transl.first_attempt_squat_res
            #     if cur_transl_val == "Присед 2":
            #         competitor_ordered_weight = competitor_transl.second_attempt_squat_res
            #     if cur_transl_val == "Присед 3":
            #         competitor_ordered_weight = competitor_transl.third_attempt_squat_res
            #     if cur_transl_val == "Жим 1":
            #         competitor_ordered_weight = competitor_transl.first_attempt_bpress_res
            #     if cur_transl_val == "Жим 2":
            #         competitor_ordered_weight = competitor_transl.second_attempt_bpress_res
            #     if cur_transl_val == "Жим 3":
            #         competitor_ordered_weight = competitor_transl.third_attempt_bpress_res
            #     if cur_transl_val == "Тяга 1":
            #         competitor_ordered_weight = competitor_transl.first_attempt_dlift_res
            #     if cur_transl_val == "Тяга 2":
            #         competitor_ordered_weight = competitor_transl.second_attempt_dlift_res
            #     if cur_transl_val == "Тяга 3":
            #         competitor_ordered_weight = competitor_transl.third_attempt_dlift_res
            # print(cur_transl_val)
            # print(cur_transl_val.split(' ')[0].lower())
            # print(cur_transl_val.split(' ')[1])
            # print(cur_transl_pk)
            # competitor_transl = CompetitionProtocols.objects.get(pk=cur_transl_pk)
            # print(f"{competitor_transl.competitor.surname_comp.capitalize()} {competitor_transl.competitor.name_comp.capitalize()}")
            # print(competitor_transl.competitor.weight_cat)
            # print(competitor_transl.competitor.sports_type.all())
            # print([i.title for i in competitor_transl.competitor.sports_type.all()])
        # print(cur_transl_val)
        # print(cur_transl_pk)
        # print(request.POST)
        # return JsonResponse({"cur_transl_val": cur_transl_val, "cur_transl_pk": cur_transl_pk}, status=200)
        # return JsonResponse({"competitor_exercise": competitor_exercise, "competitor_exercise_attempt": competitor_exercise_attempt, "competitor_fi": competitor_fi,
        #                      "competitor_weight_cat": competitor_weight_cat, "competitor_ordered_weight": competitor_ordered_weight, "competitor_ordered_weight_offset": competitor_ordered_weight_offset, "cur_transl_pk": cur_transl_pk}, status=200)
    return render(request, 'pwrlftmain/scoreboard_page.html', response_data)


@login_required
def scoreboard_comp_page(request, competition_slug):
    cur_competition = Competitions.objects.get(competition_slug=competition_slug)
    response_data = {
        'cur_competition': cur_competition,
    }
    if 'score-form_btn' in request.POST:
        cur_flow = CompetitorsFlow.objects.first().comp_flow
        cur_flow_competitors_st = CompetitionProtocols.objects.filter(competitor__comp_title__competition_slug=competition_slug, competitor_stream=cur_flow)
        competitors_stypes_st = cur_flow_competitors_st.values_list('competitor__sports_type__title', flat=True)
        competitors_stypes = []
        for csts in competitors_stypes_st:
            if csts not in competitors_stypes:
                competitors_stypes.append(csts)
        if "Троеборье классическое" in competitors_stypes:
            cur_flow_competitors = cur_flow_competitors_st.order_by(F("first_attempt_squat_res").asc(nulls_last=True))
            # if None not in cur_flow_competitors.filter(Q(competitor__sports_type__title="Троеборье классическое") & Q(competitor__sports_type__title="Троеборье (экип.)")).values_list('third_attempt_squat_off', flat=True):
            if None not in cur_flow_competitors.filter(competitor__sports_type__title="Троеборье классическое").values_list('third_attempt_squat_off', flat=True):
                cur_flow_competitors = cur_flow_competitors_st.order_by(F("first_attempt_bpress_res").asc(nulls_last=True))
                # if None not in cur_flow_competitors.filter(Q(competitor__sports_type__title="Троеборье классическое") & Q(competitor__sports_type__title="Троеборье (экип.)")).values_list('third_attempt_bpress_off', flat=True):
                if None not in cur_flow_competitors.filter(competitor__sports_type__title="Троеборье классическое").values_list('third_attempt_bpress_off', flat=True):
                    cur_flow_competitors = cur_flow_competitors_st.order_by(F("first_attempt_dlift_res").asc(nulls_last=True))
        elif "Троеборье (экип.)" in competitors_stypes:
            cur_flow_competitors = cur_flow_competitors_st.order_by(F("first_attempt_squat_res_ek").asc(nulls_last=True))
            if None not in cur_flow_competitors.filter(competitor__sports_type__title="Троеборье (экип.)").values_list('third_attempt_squat_off_ek', flat=True):
                cur_flow_competitors = cur_flow_competitors_st.order_by(F("first_attempt_bpress_res_ek").asc(nulls_last=True))
                if None not in cur_flow_competitors.filter(competitor__sports_type__title="Троеборье (экип.)").values_list('third_attempt_bpress_off_ek', flat=True):
                    cur_flow_competitors = cur_flow_competitors_st.order_by(F("first_attempt_dlift_res_ek").asc(nulls_last=True))
        elif "Жим лёжа (без экип.)" in competitors_stypes:
            cur_flow_competitors = cur_flow_competitors_st.order_by(F("first_attempt_bpress_res").asc(nulls_last=True))
        else:
            cur_flow_competitors = cur_flow_competitors_st.order_by(F("first_attempt_bpress_res_ek").asc(nulls_last=True))
        cur_flow_competitors_json = serializers.serialize("json", cur_flow_competitors, use_natural_foreign_keys=True)
        return JsonResponse({"cur_flow_competitors_json": cur_flow_competitors_json, }, status=200)
    return render(request, 'pwrlftmain/scoreboard_comp_page.html', response_data)


def create_excel_protocol(request, competition_slug):
    cur_competition = Competitions.objects.get(competition_slug=competition_slug)
    competitors = CompetitionProtocols.objects.filter(competitor__comp_title_id=cur_competition.id)
    competitiors_pwlclass_m = ""
    competitiors_pwlclass_f = ""
    competitiors_pwlekip_m = ""
    competitiors_pwlekip_f = ""
    competitiors_bpekip_m = ""
    competitiors_bpekip_f = ""
    competitiors_bp_m = ""
    competitiors_bp_f = ""
    competitiors_pwlclass_m_wcat = ""
    competitiors_pwlclass_f_wcat = ""
    competitiors_pwlekip_m_wcat = ""
    competitiors_pwlekip_f_wcat = ""
    competitiors_bpekip_m_wcat = ""
    competitiors_bpekip_f_wcat = ""
    competitiors_bp_m_wcat = ""
    competitiors_bp_f_wcat = ""
    for sp_type in cur_competition.sport_types.all():
        if sp_type.title == "Троеборье классическое":
            competitiors_pwlclass_m = competitors.filter(competitor__sports_type__title="Троеборье классическое",
                                                         competitor__gender="Мужской").order_by("-best_sum_res",
                                                                                                "competitor_weight")
            competitiors_pwlclass_m_wcat_st = competitiors_pwlclass_m.values_list("competitor__weight_cat", flat=True)
            competitiors_pwlclass_m_wcat = []
            for comppclmwc in competitiors_pwlclass_m_wcat_st:
                if comppclmwc not in competitiors_pwlclass_m_wcat:
                    competitiors_pwlclass_m_wcat.append(comppclmwc)
            competitiors_pwlclass_f = competitors.filter(competitor__sports_type__title="Троеборье классическое",
                                                         competitor__gender="Женский").order_by("-best_sum_res",
                                                                                                "competitor_weight")
            competitiors_pwlclass_f_wcat_st = competitiors_pwlclass_f.values_list("competitor__weight_cat", flat=True)
            competitiors_pwlclass_f_wcat = []
            for comppclfwc in competitiors_pwlclass_f_wcat_st:
                if comppclfwc not in competitiors_pwlclass_f_wcat:
                    competitiors_pwlclass_f_wcat.append(comppclfwc)
        if sp_type.title == "Троеборье (экип.)":
            competitiors_pwlekip_m = competitors.filter(competitor__sports_type__title="Троеборье (экип.)",
                                                        competitor__gender="Мужской").order_by("-best_sum_res_ek",
                                                                                               "competitor_weight")
            competitiors_pwlekip_m_wcat_st = competitiors_pwlekip_m.values_list("competitor__weight_cat", flat=True)
            competitiors_pwlekip_m_wcat = []
            for comppekmwc in competitiors_pwlekip_m_wcat_st:
                if comppekmwc not in competitiors_pwlekip_m_wcat:
                    competitiors_pwlekip_m_wcat.append(comppekmwc)
            competitiors_pwlekip_f = competitors.filter(competitor__sports_type__title="Троеборье (экип.)",
                                                        competitor__gender="Женский").order_by("-best_sum_res_ek",
                                                                                               "competitor_weight")
            competitiors_pwlekip_f_wcat_st = competitiors_pwlekip_f.values_list("competitor__weight_cat", flat=True)
            competitiors_pwlekip_f_wcat = []
            for comppekfwc in competitiors_pwlekip_f_wcat_st:
                if comppekfwc not in competitiors_pwlekip_f_wcat:
                    competitiors_pwlekip_f_wcat.append(comppekfwc)
        if sp_type.title == "Жим лёжа (экип.)":
            competitiors_bpekip_m = competitors.filter(competitor__sports_type__title="Жим лёжа (экип.)",
                                                       competitor__gender="Мужской").order_by("-best_bpress_res_ek",
                                                                                              "competitor_weight")
            competitiors_bpekip_m_wcat_st = competitiors_bpekip_m.values_list("competitor__weight_cat", flat=True)
            competitiors_bpekip_m_wcat = []
            for combpekmwc in competitiors_bpekip_m_wcat_st:
                if combpekmwc not in competitiors_bpekip_m_wcat:
                    competitiors_bpekip_m_wcat.append(combpekmwc)
            competitiors_bpekip_f = competitors.filter(competitor__sports_type__title="Жим лёжа (экип.)",
                                                       competitor__gender="Женский").order_by("-best_bpress_res_ek",
                                                                                              "competitor_weight")
            competitiors_bpekip_f_wcat_st = competitiors_bpekip_f.values_list("competitor__weight_cat", flat=True)
            competitiors_bpekip_f_wcat = []
            for combpekfwc in competitiors_bpekip_f_wcat_st:
                if combpekfwc not in competitiors_bpekip_f_wcat:
                    competitiors_bpekip_f_wcat.append(combpekfwc)
        if sp_type.title == "Жим лёжа (без экип.)":
            competitiors_bp_m = competitors.filter(competitor__sports_type__title="Жим лёжа (без экип.)",
                                                   competitor__gender="Мужской").order_by("-best_bpress_res",
                                                                                          "competitor_weight")
            competitiors_bp_m_wcat_st = competitiors_bp_m.values_list("competitor__weight_cat", flat=True)
            competitiors_bp_m_wcat = []
            for combpclmwc in competitiors_bp_m_wcat_st:
                if combpclmwc not in competitiors_bp_m_wcat:
                    competitiors_bp_m_wcat.append(combpclmwc)
            competitiors_bp_f = competitors.filter(competitor__sports_type__title="Жим лёжа (без экип.)",
                                                   competitor__gender="Женский").order_by("-best_bpress_res",
                                                                                          "competitor_weight")
            competitiors_bp_f_wcat_st = competitiors_bp_f.values_list("competitor__weight_cat", flat=True)
            competitiors_bp_f_wcat = []
            for combpclfwc in competitiors_bp_f_wcat_st:
                if combpclfwc not in competitiors_bp_f_wcat:
                    competitiors_bp_f_wcat.append(combpclfwc)
    wb = Workbook()
    font_title = Font(bold=True, color='000000', name='Arial', size=14)
    font_sub_title = Font(bold=True, color='000000', name='Arial', size=12)
    font_title_table = Font(bold=True, color='000000', name='Arial', size=10)
    font_table_text = Font(bold=False, color='000000', name='Arial', size=10)
    alignment_center = Alignment(horizontal='center', vertical='center')
    thins = Side(border_style='thin', color='000000')
    border = Border(top=thins, bottom=thins, left=thins, right=thins)
    def create_sheet_pl_cl(wbook, name_disc, name_disc_full, wcats, competitors_wcats, font_title, font_sub_title, font_title_table, font_table_text, alignment_center, border, cur_competition_title, competitors_pl):
        wbook.create_sheet(name_disc)
        ws = wbook[name_disc]
        ws.insert_rows(1)
        ws.merge_cells('A2:I2')
        title_sheet = ws['A2']
        title_sheet.value = cur_competition_title
        title_sheet.font = font_title
        title_sheet.alignment = alignment_center
        ws.merge_cells('A3:I3')
        sub_title_sheet = ws['A3']
        sub_title_sheet.value = name_disc_full
        sub_title_sheet.font = font_sub_title
        sub_title_sheet.alignment = alignment_center
        ws.insert_rows(4)
        start_counter = 5
        for wc in wcats:
            if wc in competitors_wcats:
                ws.merge_cells(f'A{start_counter}:I{start_counter}')
                title_table = ws[f'A{start_counter}']
                title_table.value = f'Весовая категория {wc} кг'
                title_table.font = font_sub_title
                title_table.alignment = alignment_center
                start_counter += 1
                table_title_place = ws[f'A{start_counter}']
                table_title_place.value = 'место №'
                table_title_place.font = font_title_table
                table_title_place.border = border
                table_title_place.alignment = alignment_center
                table_title_fio = ws[f'B{start_counter}']
                table_title_fio.value = 'фио'
                table_title_fio.font = font_title_table
                table_title_fio.border = border
                table_title_fio.alignment = alignment_center
                table_title_birthday = ws[f'C{start_counter}']
                table_title_birthday.value = 'дата рождения'
                table_title_birthday.font = font_title_table
                table_title_birthday.border = border
                table_title_birthday.alignment = alignment_center
                table_title_sport_club = ws[f'D{start_counter}']
                table_title_sport_club.value = 'спортивный клуб'
                table_title_sport_club.font = font_title_table
                table_title_sport_club.border = border
                table_title_sport_club.alignment = alignment_center
                table_title_weight = ws[f'E{start_counter}']
                table_title_weight.value = 'вес'
                table_title_weight.font = font_title_table
                table_title_weight.border = border
                table_title_weight.alignment = alignment_center
                table_title_squat = ws[f'F{start_counter}']
                table_title_squat.value = 'присед'
                table_title_squat.font = font_title_table
                table_title_squat.border = border
                table_title_squat.alignment = alignment_center
                table_title_bpress = ws[f'G{start_counter}']
                table_title_bpress.value = 'жим'
                table_title_bpress.font = font_title_table
                table_title_bpress.border = border
                table_title_bpress.alignment = alignment_center
                table_title_dlift = ws[f'H{start_counter}']
                table_title_dlift.value = 'тяга'
                table_title_dlift.font = font_title_table
                table_title_dlift.border = border
                table_title_dlift.alignment = alignment_center
                table_title_total = ws[f'I{start_counter}']
                table_title_total.value = 'сумма'
                table_title_total.font = font_title_table
                table_title_total.border = border
                table_title_total.alignment = alignment_center
                start_counter += 1
                start_place_counter = 0
                for comp in competitors_pl:
                    if comp.competitor.weight_cat == wc:
                        start_place_counter += 1
                        table_cell_place = ws[f'A{start_counter}']
                        table_cell_place.value = start_place_counter
                        table_cell_place.font = font_table_text
                        table_cell_place.border = border
                        table_cell_fio = ws[f'B{start_counter}']
                        table_cell_fio.value = f"{comp.competitor.surname_comp} {comp.competitor.name_comp} {comp.competitor.patronymic_comp}"
                        table_cell_fio.font = font_table_text
                        table_cell_fio.border = border
                        table_cell_birthday = ws[f'C{start_counter}']
                        table_cell_birthday.number_format = 'DD.MM.YYYY'
                        table_cell_birthday.value = comp.competitor.birthday
                        table_cell_birthday.font = font_table_text
                        table_cell_birthday.border = border
                        table_cell_sport_club = ws[f'D{start_counter}']
                        table_cell_sport_club.value = "ProSport"
                        comp_sport_club = comp.competitor.club
                        if comp_sport_club:
                            table_cell_sport_club.value = comp_sport_club
                        table_cell_sport_club.font = font_table_text
                        table_cell_sport_club.border = border
                        table_cell_weight = ws[f'E{start_counter}']
                        table_cell_weight.value = comp.competitor_weight
                        table_cell_weight.font = font_table_text
                        table_cell_weight.border = border
                        table_cell_squat = ws[f'F{start_counter}']
                        table_cell_squat.value = comp.best_squat_res
                        table_cell_squat.font = font_table_text
                        table_cell_squat.border = border
                        table_cell_bpress = ws[f'G{start_counter}']
                        table_cell_bpress.value = comp.best_bpress_res
                        table_cell_bpress.font = font_table_text
                        table_cell_bpress.border = border
                        table_cell_dlift = ws[f'H{start_counter}']
                        table_cell_dlift.value = comp.best_dlift_res
                        table_cell_dlift.font = font_table_text
                        table_cell_dlift.border = border
                        table_cell_total = ws[f'I{start_counter}']
                        table_cell_total.value = comp.best_sum_res
                        table_cell_total.font = font_table_text
                        table_cell_total.border = border
                        start_counter += 1
                ws.merge_cells(f'A{start_counter}:I{start_counter}')
                ws[f'A{start_counter}'] = ''
                start_counter += 1
        dims = {}
        dims1 = {}
        dims2 = {}
        for row in ws.rows:
            for cell in row:
                if cell.coordinate[0] == 'B' and cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
                if cell.coordinate[0] == 'C' and cell.value:
                    dims1[cell.column] = max((dims1.get(cell.column, 0), len(str(cell.value))))
                if cell.coordinate[0] == 'D' and cell.value:
                    dims2[cell.column] = max((dims2.get(cell.column, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions['B'].width = value + 1
        for col, value in dims1.items():
            ws.column_dimensions['C'].width = value + 2
        for col, value in dims2.items():
            ws.column_dimensions['D'].width = value + 3
        return "Done"
    # def create_sheet_pl_cl(wbook, name_disc, name_disc_full, wcats, competitors_wcats, font_title, font_sub_title, font_title_table, font_table_text, alignment_center, border, cur_competition_title, competitors_pl):
    #     wbook.create_sheet(name_disc)
    #     ws = wbook[name_disc]
    #     ws.insert_rows(1)
    #     ws.merge_cells('A2:H2')
    #     title_sheet = ws['A2']
    #     title_sheet.value = cur_competition_title
    #     title_sheet.font = font_title
    #     title_sheet.alignment = alignment_center
    #     ws.merge_cells('A3:H3')
    #     sub_title_sheet = ws['A3']
    #     sub_title_sheet.value = name_disc_full
    #     sub_title_sheet.font = font_sub_title
    #     sub_title_sheet.alignment = alignment_center
    #     ws.insert_rows(4)
    #     start_counter = 5
    #     for wc in wcats:
    #         if wc in competitors_wcats:
    #             ws.merge_cells(f'A{start_counter}:H{start_counter}')
    #             title_table = ws[f'A{start_counter}']
    #             title_table.value = f'Весовая категория {wc} кг'
    #             title_table.font = font_sub_title
    #             title_table.alignment = alignment_center
    #             start_counter += 1
    #             table_title_place = ws[f'A{start_counter}']
    #             table_title_place.value = 'место №'
    #             table_title_place.font = font_title_table
    #             table_title_place.border = border
    #             table_title_place.alignment = alignment_center
    #             table_title_fio = ws[f'B{start_counter}']
    #             table_title_fio.value = 'фио'
    #             table_title_fio.font = font_title_table
    #             table_title_fio.border = border
    #             table_title_fio.alignment = alignment_center
    #             table_title_birthday = ws[f'C{start_counter}']
    #             table_title_birthday.value = 'дата рождения'
    #             table_title_birthday.font = font_title_table
    #             table_title_birthday.border = border
    #             table_title_birthday.alignment = alignment_center
    #             table_title_weight = ws[f'D{start_counter}']
    #             table_title_weight.value = 'вес'
    #             table_title_weight.font = font_title_table
    #             table_title_weight.border = border
    #             table_title_weight.alignment = alignment_center
    #             table_title_squat = ws[f'E{start_counter}']
    #             table_title_squat.value = 'присед'
    #             table_title_squat.font = font_title_table
    #             table_title_squat.border = border
    #             table_title_squat.alignment = alignment_center
    #             table_title_bpress = ws[f'F{start_counter}']
    #             table_title_bpress.value = 'жим'
    #             table_title_bpress.font = font_title_table
    #             table_title_bpress.border = border
    #             table_title_bpress.alignment = alignment_center
    #             table_title_dlift = ws[f'G{start_counter}']
    #             table_title_dlift.value = 'тяга'
    #             table_title_dlift.font = font_title_table
    #             table_title_dlift.border = border
    #             table_title_dlift.alignment = alignment_center
    #             table_title_total = ws[f'H{start_counter}']
    #             table_title_total.value = 'сумма'
    #             table_title_total.font = font_title_table
    #             table_title_total.border = border
    #             table_title_total.alignment = alignment_center
    #             start_counter += 1
    #             start_place_counter = 0
    #             for comp in competitors_pl:
    #                 if comp.competitor.weight_cat == wc:
    #                     start_place_counter += 1
    #                     table_cell_place = ws[f'A{start_counter}']
    #                     table_cell_place.value = start_place_counter
    #                     table_cell_place.font = font_table_text
    #                     table_cell_place.border = border
    #                     table_cell_fio = ws[f'B{start_counter}']
    #                     table_cell_fio.value = f"{comp.competitor.surname_comp} {comp.competitor.name_comp} {comp.competitor.patronymic_comp}"
    #                     table_cell_fio.font = font_table_text
    #                     table_cell_fio.border = border
    #                     table_cell_birthday = ws[f'C{start_counter}']
    #                     table_cell_birthday.number_format = 'DD.MM.YYYY'
    #                     table_cell_birthday.value = comp.competitor.birthday
    #                     table_cell_birthday.font = font_table_text
    #                     table_cell_birthday.border = border
    #                     table_cell_weight = ws[f'D{start_counter}']
    #                     table_cell_weight.value = comp.competitor_weight
    #                     table_cell_weight.font = font_table_text
    #                     table_cell_weight.border = border
    #                     table_cell_squat = ws[f'E{start_counter}']
    #                     table_cell_squat.value = comp.best_squat_res
    #                     table_cell_squat.font = font_table_text
    #                     table_cell_squat.border = border
    #                     table_cell_bpress = ws[f'F{start_counter}']
    #                     table_cell_bpress.value = comp.best_bpress_res
    #                     table_cell_bpress.font = font_table_text
    #                     table_cell_bpress.border = border
    #                     table_cell_dlift = ws[f'G{start_counter}']
    #                     table_cell_dlift.value = comp.best_dlift_res
    #                     table_cell_dlift.font = font_table_text
    #                     table_cell_dlift.border = border
    #                     table_cell_total = ws[f'H{start_counter}']
    #                     table_cell_total.value = comp.best_sum_res
    #                     table_cell_total.font = font_table_text
    #                     table_cell_total.border = border
    #                     start_counter += 1
    #             ws.merge_cells(f'A{start_counter}:H{start_counter}')
    #             ws[f'A{start_counter}'] = ''
    #             start_counter += 1
    #     dims = {}
    #     dims1 = {}
    #     for row in ws.rows:
    #         for cell in row:
    #             if cell.coordinate[0] == 'B' and cell.value:
    #                 dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
    #             if cell.coordinate[0] == 'C' and cell.value:
    #                 dims1[cell.column] = max((dims1.get(cell.column, 0), len(str(cell.value))))
    #     for col, value in dims.items():
    #         ws.column_dimensions['B'].width = value + 1
    #     for col, value in dims1.items():
    #         ws.column_dimensions['C'].width = value + 2
    #     return "Done"
    def create_sheet_pl(wbook, name_disc, name_disc_full, wcats, competitors_wcats, font_title, font_sub_title, font_title_table, font_table_text, alignment_center, border, cur_competition_title, competitors_pl):
        wbook.create_sheet(name_disc)
        ws = wbook[name_disc]
        ws.insert_rows(1)
        ws.merge_cells('A2:I2')
        title_sheet = ws['A2']
        title_sheet.value = cur_competition_title
        title_sheet.font = font_title
        title_sheet.alignment = alignment_center
        ws.merge_cells('A3:I3')
        sub_title_sheet = ws['A3']
        sub_title_sheet.value = name_disc_full
        sub_title_sheet.font = font_sub_title
        sub_title_sheet.alignment = alignment_center
        ws.insert_rows(4)
        start_counter = 5
        for wc in wcats:
            if wc in competitors_wcats:
                ws.merge_cells(f'A{start_counter}:I{start_counter}')
                title_table = ws[f'A{start_counter}']
                title_table.value = f'Весовая категория {wc} кг'
                title_table.font = font_sub_title
                title_table.alignment = alignment_center
                start_counter += 1
                table_title_place = ws[f'A{start_counter}']
                table_title_place.value = 'место №'
                table_title_place.font = font_title_table
                table_title_place.border = border
                table_title_place.alignment = alignment_center
                table_title_fio = ws[f'B{start_counter}']
                table_title_fio.value = 'фио'
                table_title_fio.font = font_title_table
                table_title_fio.border = border
                table_title_fio.alignment = alignment_center
                table_title_birthday = ws[f'C{start_counter}']
                table_title_birthday.value = 'дата рождения'
                table_title_birthday.font = font_title_table
                table_title_birthday.border = border
                table_title_birthday.alignment = alignment_center
                table_title_sport_club = ws[f'D{start_counter}']
                table_title_sport_club.value = 'спортивный клуб'
                table_title_sport_club.font = font_title_table
                table_title_sport_club.border = border
                table_title_sport_club.alignment = alignment_center
                table_title_weight = ws[f'E{start_counter}']
                table_title_weight.value = 'вес'
                table_title_weight.font = font_title_table
                table_title_weight.border = border
                table_title_weight.alignment = alignment_center
                table_title_squat = ws[f'F{start_counter}']
                table_title_squat.value = 'присед'
                table_title_squat.font = font_title_table
                table_title_squat.border = border
                table_title_squat.alignment = alignment_center
                table_title_bpress = ws[f'G{start_counter}']
                table_title_bpress.value = 'жим'
                table_title_bpress.font = font_title_table
                table_title_bpress.border = border
                table_title_bpress.alignment = alignment_center
                table_title_dlift = ws[f'H{start_counter}']
                table_title_dlift.value = 'тяга'
                table_title_dlift.font = font_title_table
                table_title_dlift.border = border
                table_title_dlift.alignment = alignment_center
                table_title_total = ws[f'I{start_counter}']
                table_title_total.value = 'сумма'
                table_title_total.font = font_title_table
                table_title_total.border = border
                table_title_total.alignment = alignment_center
                start_counter += 1
                start_place_counter = 0
                for comp in competitors_pl:
                    if comp.competitor.weight_cat == wc:
                        start_place_counter += 1
                        table_cell_place = ws[f'A{start_counter}']
                        table_cell_place.value = start_place_counter
                        table_cell_place.font = font_table_text
                        table_cell_place.border = border
                        table_cell_fio = ws[f'B{start_counter}']
                        table_cell_fio.value = f"{comp.competitor.surname_comp} {comp.competitor.name_comp} {comp.competitor.patronymic_comp}"
                        table_cell_fio.font = font_table_text
                        table_cell_fio.border = border
                        table_cell_birthday = ws[f'C{start_counter}']
                        table_cell_birthday.number_format = 'DD.MM.YYYY'
                        table_cell_birthday.value = comp.competitor.birthday
                        table_cell_birthday.font = font_table_text
                        table_cell_birthday.border = border
                        table_cell_sport_club = ws[f'D{start_counter}']
                        table_cell_sport_club.value = "ProSport"
                        comp_sport_club = comp.competitor.club
                        if comp_sport_club:
                            table_cell_sport_club.value = comp_sport_club
                        table_cell_sport_club.font = font_table_text
                        table_cell_sport_club.border = border
                        table_cell_weight = ws[f'E{start_counter}']
                        table_cell_weight.value = comp.competitor_weight
                        table_cell_weight.font = font_table_text
                        table_cell_weight.border = border
                        table_cell_squat = ws[f'F{start_counter}']
                        table_cell_squat.value = comp.best_squat_res_ek
                        table_cell_squat.font = font_table_text
                        table_cell_squat.border = border
                        table_cell_bpress = ws[f'G{start_counter}']
                        table_cell_bpress.value = comp.best_bpress_res_ek
                        table_cell_bpress.font = font_table_text
                        table_cell_bpress.border = border
                        table_cell_dlift = ws[f'H{start_counter}']
                        table_cell_dlift.value = comp.best_dlift_res_ek
                        table_cell_dlift.font = font_table_text
                        table_cell_dlift.border = border
                        table_cell_total = ws[f'I{start_counter}']
                        table_cell_total.value = comp.best_sum_res_ek
                        table_cell_total.font = font_table_text
                        table_cell_total.border = border
                        start_counter += 1
                ws.merge_cells(f'A{start_counter}:I{start_counter}')
                ws[f'A{start_counter}'] = ''
                start_counter += 1
        dims = {}
        dims1 = {}
        dims2 = {}
        for row in ws.rows:
            for cell in row:
                if cell.coordinate[0] == 'B' and cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
                if cell.coordinate[0] == 'C' and cell.value:
                    dims1[cell.column] = max((dims1.get(cell.column, 0), len(str(cell.value))))
                if cell.coordinate[0] == 'D' and cell.value:
                    dims2[cell.column] = max((dims2.get(cell.column, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions['B'].width = value + 1
        for col, value in dims1.items():
            ws.column_dimensions['C'].width = value + 2
        for col, value in dims2.items():
            ws.column_dimensions['D'].width = value + 2
        return "Done"
    # def create_sheet_pl(wbook, name_disc, name_disc_full, wcats, competitors_wcats, font_title, font_sub_title, font_title_table, font_table_text, alignment_center, border, cur_competition_title, competitors_pl):
    #     wbook.create_sheet(name_disc)
    #     ws = wbook[name_disc]
    #     ws.insert_rows(1)
    #     ws.merge_cells('A2:H2')
    #     title_sheet = ws['A2']
    #     title_sheet.value = cur_competition_title
    #     title_sheet.font = font_title
    #     title_sheet.alignment = alignment_center
    #     ws.merge_cells('A3:H3')
    #     sub_title_sheet = ws['A3']
    #     sub_title_sheet.value = name_disc_full
    #     sub_title_sheet.font = font_sub_title
    #     sub_title_sheet.alignment = alignment_center
    #     ws.insert_rows(4)
    #     start_counter = 5
    #     for wc in wcats:
    #         if wc in competitors_wcats:
    #             ws.merge_cells(f'A{start_counter}:H{start_counter}')
    #             title_table = ws[f'A{start_counter}']
    #             title_table.value = f'Весовая категория {wc} кг'
    #             title_table.font = font_sub_title
    #             title_table.alignment = alignment_center
    #             start_counter += 1
    #             table_title_place = ws[f'A{start_counter}']
    #             table_title_place.value = 'место №'
    #             table_title_place.font = font_title_table
    #             table_title_place.border = border
    #             table_title_place.alignment = alignment_center
    #             table_title_fio = ws[f'B{start_counter}']
    #             table_title_fio.value = 'фио'
    #             table_title_fio.font = font_title_table
    #             table_title_fio.border = border
    #             table_title_fio.alignment = alignment_center
    #             table_title_birthday = ws[f'C{start_counter}']
    #             table_title_birthday.value = 'дата рождения'
    #             table_title_birthday.font = font_title_table
    #             table_title_birthday.border = border
    #             table_title_birthday.alignment = alignment_center
    #             table_title_weight = ws[f'D{start_counter}']
    #             table_title_weight.value = 'вес'
    #             table_title_weight.font = font_title_table
    #             table_title_weight.border = border
    #             table_title_weight.alignment = alignment_center
    #             table_title_squat = ws[f'E{start_counter}']
    #             table_title_squat.value = 'присед'
    #             table_title_squat.font = font_title_table
    #             table_title_squat.border = border
    #             table_title_squat.alignment = alignment_center
    #             table_title_bpress = ws[f'F{start_counter}']
    #             table_title_bpress.value = 'жим'
    #             table_title_bpress.font = font_title_table
    #             table_title_bpress.border = border
    #             table_title_bpress.alignment = alignment_center
    #             table_title_dlift = ws[f'G{start_counter}']
    #             table_title_dlift.value = 'тяга'
    #             table_title_dlift.font = font_title_table
    #             table_title_dlift.border = border
    #             table_title_dlift.alignment = alignment_center
    #             table_title_total = ws[f'H{start_counter}']
    #             table_title_total.value = 'сумма'
    #             table_title_total.font = font_title_table
    #             table_title_total.border = border
    #             table_title_total.alignment = alignment_center
    #             start_counter += 1
    #             start_place_counter = 0
    #             for comp in competitors_pl:
    #                 if comp.competitor.weight_cat == wc:
    #                     start_place_counter += 1
    #                     table_cell_place = ws[f'A{start_counter}']
    #                     table_cell_place.value = start_place_counter
    #                     table_cell_place.font = font_table_text
    #                     table_cell_place.border = border
    #                     table_cell_fio = ws[f'B{start_counter}']
    #                     table_cell_fio.value = f"{comp.competitor.surname_comp} {comp.competitor.name_comp} {comp.competitor.patronymic_comp}"
    #                     table_cell_fio.font = font_table_text
    #                     table_cell_fio.border = border
    #                     table_cell_birthday = ws[f'C{start_counter}']
    #                     table_cell_birthday.number_format = 'DD.MM.YYYY'
    #                     table_cell_birthday.value = comp.competitor.birthday
    #                     table_cell_birthday.font = font_table_text
    #                     table_cell_birthday.border = border
    #                     table_cell_weight = ws[f'D{start_counter}']
    #                     table_cell_weight.value = comp.competitor_weight
    #                     table_cell_weight.font = font_table_text
    #                     table_cell_weight.border = border
    #                     table_cell_squat = ws[f'E{start_counter}']
    #                     table_cell_squat.value = comp.best_squat_res_ek
    #                     table_cell_squat.font = font_table_text
    #                     table_cell_squat.border = border
    #                     table_cell_bpress = ws[f'F{start_counter}']
    #                     table_cell_bpress.value = comp.best_bpress_res_ek
    #                     table_cell_bpress.font = font_table_text
    #                     table_cell_bpress.border = border
    #                     table_cell_dlift = ws[f'G{start_counter}']
    #                     table_cell_dlift.value = comp.best_dlift_res_ek
    #                     table_cell_dlift.font = font_table_text
    #                     table_cell_dlift.border = border
    #                     table_cell_total = ws[f'H{start_counter}']
    #                     table_cell_total.value = comp.best_sum_res_ek
    #                     table_cell_total.font = font_table_text
    #                     table_cell_total.border = border
    #                     start_counter += 1
    #             ws.merge_cells(f'A{start_counter}:H{start_counter}')
    #             ws[f'A{start_counter}'] = ''
    #             start_counter += 1
    #     dims = {}
    #     dims1 = {}
    #     for row in ws.rows:
    #         for cell in row:
    #             if cell.coordinate[0] == 'B' and cell.value:
    #                 dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
    #             if cell.coordinate[0] == 'C' and cell.value:
    #                 dims1[cell.column] = max((dims1.get(cell.column, 0), len(str(cell.value))))
    #     for col, value in dims.items():
    #         ws.column_dimensions['B'].width = value + 1
    #     for col, value in dims1.items():
    #         ws.column_dimensions['C'].width = value + 2
    #     return "Done"
    def create_sheet_bp(wbook, name_disc, name_disc_full, wcats, competitors_wcats, font_title, font_sub_title, font_title_table, font_table_text, alignment_center, border, cur_competition_title, competitors_pl):
        wbook.create_sheet(name_disc)
        ws = wbook[name_disc]
        ws.insert_rows(1)
        ws.merge_cells('A2:F2')
        title_sheet = ws['A2']
        title_sheet.value = cur_competition_title
        title_sheet.font = font_title
        title_sheet.alignment = alignment_center
        ws.merge_cells('A3:F3')
        sub_title_sheet = ws['A3']
        sub_title_sheet.value = name_disc_full
        sub_title_sheet.font = font_sub_title
        sub_title_sheet.alignment = alignment_center
        ws.insert_rows(4)
        start_counter = 5
        for wc in wcats:
            if wc in competitors_wcats:
                ws.merge_cells(f'A{start_counter}:F{start_counter}')
                title_table = ws[f'A{start_counter}']
                title_table.value = f'Весовая категория {wc} кг'
                title_table.font = font_sub_title
                title_table.alignment = alignment_center
                start_counter += 1
                table_title_place = ws[f'A{start_counter}']
                table_title_place.value = 'место №'
                table_title_place.font = font_title_table
                table_title_place.border = border
                table_title_place.alignment = alignment_center
                table_title_fio = ws[f'B{start_counter}']
                table_title_fio.value = 'фио'
                table_title_fio.font = font_title_table
                table_title_fio.border = border
                table_title_fio.alignment = alignment_center
                table_title_birthday = ws[f'C{start_counter}']
                table_title_birthday.value = 'дата рождения'
                table_title_birthday.font = font_title_table
                table_title_birthday.border = border
                table_title_birthday.alignment = alignment_center
                table_title_sport_club = ws[f'D{start_counter}']
                table_title_sport_club.value = 'спортивный клуб'
                table_title_sport_club.font = font_title_table
                table_title_sport_club.border = border
                table_title_sport_club.alignment = alignment_center
                table_title_weight = ws[f'E{start_counter}']
                table_title_weight.value = 'вес'
                table_title_weight.font = font_title_table
                table_title_weight.border = border
                table_title_weight.alignment = alignment_center
                table_title_bpress = ws[f'F{start_counter}']
                table_title_bpress.value = 'жим'
                table_title_bpress.font = font_title_table
                table_title_bpress.border = border
                table_title_bpress.alignment = alignment_center
                start_counter += 1
                start_place_counter = 0
                for comp in competitors_pl:
                    if comp.competitor.weight_cat == wc:
                        start_place_counter += 1
                        table_cell_place = ws[f'A{start_counter}']
                        table_cell_place.value = start_place_counter
                        table_cell_place.font = font_table_text
                        table_cell_place.border = border
                        table_cell_fio = ws[f'B{start_counter}']
                        table_cell_fio.value = f"{comp.competitor.surname_comp} {comp.competitor.name_comp} {comp.competitor.patronymic_comp}"
                        table_cell_fio.font = font_table_text
                        table_cell_fio.border = border
                        table_cell_birthday = ws[f'C{start_counter}']
                        table_cell_birthday.number_format = 'DD.MM.YYYY'
                        table_cell_birthday.value = comp.competitor.birthday
                        table_cell_birthday.font = font_table_text
                        table_cell_birthday.border = border
                        table_cell_sport_club = ws[f'D{start_counter}']
                        table_cell_sport_club.value = "ProSport"
                        comp_sport_club = comp.competitor.club
                        if comp_sport_club:
                            table_cell_sport_club.value = comp_sport_club
                        table_cell_sport_club.font = font_table_text
                        table_cell_sport_club.border = border
                        table_cell_weight = ws[f'E{start_counter}']
                        table_cell_weight.value = comp.competitor_weight
                        table_cell_weight.font = font_table_text
                        table_cell_weight.border = border
                        table_cell_bpress = ws[f'F{start_counter}']
                        table_cell_bpress.value = comp.best_bpress_res_ek
                        table_cell_bpress.font = font_table_text
                        table_cell_bpress.border = border
                        start_counter += 1
                ws.merge_cells(f'A{start_counter}:F{start_counter}')
                ws[f'A{start_counter}'] = ''
                start_counter += 1
        dims = {}
        dims1 = {}
        dims2 = {}
        for row in ws.rows:
            for cell in row:
                if cell.coordinate[0] == 'B' and cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
                if cell.coordinate[0] == 'C' and cell.value:
                    dims1[cell.column] = max((dims1.get(cell.column, 0), len(str(cell.value))))
                if cell.coordinate[0] == 'D' and cell.value:
                    dims2[cell.column] = max((dims2.get(cell.column, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions['B'].width = value + 1
        for col, value in dims1.items():
            ws.column_dimensions['C'].width = value + 2
        for col, value in dims2.items():
            ws.column_dimensions['D'].width = value + 2
        return "Done"
    # def create_sheet_bp(wbook, name_disc, name_disc_full, wcats, competitors_wcats, font_title, font_sub_title, font_title_table, font_table_text, alignment_center, border, cur_competition_title, competitors_pl):
    #     wbook.create_sheet(name_disc)
    #     ws = wbook[name_disc]
    #     ws.insert_rows(1)
    #     ws.merge_cells('A2:E2')
    #     title_sheet = ws['A2']
    #     title_sheet.value = cur_competition_title
    #     title_sheet.font = font_title
    #     title_sheet.alignment = alignment_center
    #     ws.merge_cells('A3:E3')
    #     sub_title_sheet = ws['A3']
    #     sub_title_sheet.value = name_disc_full
    #     sub_title_sheet.font = font_sub_title
    #     sub_title_sheet.alignment = alignment_center
    #     ws.insert_rows(4)
    #     start_counter = 5
    #     for wc in wcats:
    #         if wc in competitors_wcats:
    #             ws.merge_cells(f'A{start_counter}:E{start_counter}')
    #             title_table = ws[f'A{start_counter}']
    #             title_table.value = f'Весовая категория {wc} кг'
    #             title_table.font = font_sub_title
    #             title_table.alignment = alignment_center
    #             start_counter += 1
    #             table_title_place = ws[f'A{start_counter}']
    #             table_title_place.value = 'место №'
    #             table_title_place.font = font_title_table
    #             table_title_place.border = border
    #             table_title_place.alignment = alignment_center
    #             table_title_fio = ws[f'B{start_counter}']
    #             table_title_fio.value = 'фио'
    #             table_title_fio.font = font_title_table
    #             table_title_fio.border = border
    #             table_title_fio.alignment = alignment_center
    #             table_title_birthday = ws[f'C{start_counter}']
    #             table_title_birthday.value = 'дата рождения'
    #             table_title_birthday.font = font_title_table
    #             table_title_birthday.border = border
    #             table_title_birthday.alignment = alignment_center
    #             table_title_weight = ws[f'D{start_counter}']
    #             table_title_weight.value = 'вес'
    #             table_title_weight.font = font_title_table
    #             table_title_weight.border = border
    #             table_title_weight.alignment = alignment_center
    #             table_title_bpress = ws[f'E{start_counter}']
    #             table_title_bpress.value = 'жим'
    #             table_title_bpress.font = font_title_table
    #             table_title_bpress.border = border
    #             table_title_bpress.alignment = alignment_center
    #             start_counter += 1
    #             start_place_counter = 0
    #             for comp in competitors_pl:
    #                 if comp.competitor.weight_cat == wc:
    #                     start_place_counter += 1
    #                     table_cell_place = ws[f'A{start_counter}']
    #                     table_cell_place.value = start_place_counter
    #                     table_cell_place.font = font_table_text
    #                     table_cell_place.border = border
    #                     table_cell_fio = ws[f'B{start_counter}']
    #                     table_cell_fio.value = f"{comp.competitor.surname_comp} {comp.competitor.name_comp} {comp.competitor.patronymic_comp}"
    #                     table_cell_fio.font = font_table_text
    #                     table_cell_fio.border = border
    #                     table_cell_birthday = ws[f'C{start_counter}']
    #                     table_cell_birthday.number_format = 'DD.MM.YYYY'
    #                     table_cell_birthday.value = comp.competitor.birthday
    #                     table_cell_birthday.font = font_table_text
    #                     table_cell_birthday.border = border
    #                     table_cell_weight = ws[f'D{start_counter}']
    #                     table_cell_weight.value = comp.competitor_weight
    #                     table_cell_weight.font = font_table_text
    #                     table_cell_weight.border = border
    #                     table_cell_bpress = ws[f'E{start_counter}']
    #                     table_cell_bpress.value = comp.best_bpress_res_ek
    #                     table_cell_bpress.font = font_table_text
    #                     table_cell_bpress.border = border
    #                     start_counter += 1
    #             ws.merge_cells(f'A{start_counter}:E{start_counter}')
    #             ws[f'A{start_counter}'] = ''
    #             start_counter += 1
    #     dims = {}
    #     dims1 = {}
    #     for row in ws.rows:
    #         for cell in row:
    #             if cell.coordinate[0] == 'B' and cell.value:
    #                 dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
    #             if cell.coordinate[0] == 'C' and cell.value:
    #                 dims1[cell.column] = max((dims1.get(cell.column, 0), len(str(cell.value))))
    #     for col, value in dims.items():
    #         ws.column_dimensions['B'].width = value + 1
    #     for col, value in dims1.items():
    #         ws.column_dimensions['C'].width = value + 2
    #     return "Done"
    def create_sheet_bp_cl(wbook, name_disc, name_disc_full, wcats, competitors_wcats, font_title, font_sub_title, font_title_table, font_table_text, alignment_center, border, cur_competition_title, competitors_pl):
        wbook.create_sheet(name_disc)
        ws = wbook[name_disc]
        ws.insert_rows(1)
        ws.merge_cells('A2:F2')
        title_sheet = ws['A2']
        title_sheet.value = cur_competition_title
        title_sheet.font = font_title
        title_sheet.alignment = alignment_center
        ws.merge_cells('A3:F3')
        sub_title_sheet = ws['A3']
        sub_title_sheet.value = name_disc_full
        sub_title_sheet.font = font_sub_title
        sub_title_sheet.alignment = alignment_center
        ws.insert_rows(4)
        start_counter = 5
        for wc in wcats:
            if wc in competitors_wcats:
                ws.merge_cells(f'A{start_counter}:F{start_counter}')
                title_table = ws[f'A{start_counter}']
                title_table.value = f'Весовая категория {wc} кг'
                title_table.font = font_sub_title
                title_table.alignment = alignment_center
                start_counter += 1
                table_title_place = ws[f'A{start_counter}']
                table_title_place.value = 'место №'
                table_title_place.font = font_title_table
                table_title_place.border = border
                table_title_place.alignment = alignment_center
                table_title_fio = ws[f'B{start_counter}']
                table_title_fio.value = 'фио'
                table_title_fio.font = font_title_table
                table_title_fio.border = border
                table_title_fio.alignment = alignment_center
                table_title_birthday = ws[f'C{start_counter}']
                table_title_birthday.value = 'дата рождения'
                table_title_birthday.font = font_title_table
                table_title_birthday.border = border
                table_title_birthday.alignment = alignment_center
                table_title_sport_club = ws[f'D{start_counter}']
                table_title_sport_club.value = 'спортивный клуб'
                table_title_sport_club.font = font_title_table
                table_title_sport_club.border = border
                table_title_sport_club.alignment = alignment_center
                table_title_weight = ws[f'E{start_counter}']
                table_title_weight.value = 'вес'
                table_title_weight.font = font_title_table
                table_title_weight.border = border
                table_title_weight.alignment = alignment_center
                table_title_bpress = ws[f'F{start_counter}']
                table_title_bpress.value = 'жим'
                table_title_bpress.font = font_title_table
                table_title_bpress.border = border
                table_title_bpress.alignment = alignment_center
                start_counter += 1
                start_place_counter = 0
                for comp in competitors_pl:
                    if comp.competitor.weight_cat == wc:
                        start_place_counter += 1
                        table_cell_place = ws[f'A{start_counter}']
                        table_cell_place.value = start_place_counter
                        table_cell_place.font = font_table_text
                        table_cell_place.border = border
                        table_cell_fio = ws[f'B{start_counter}']
                        table_cell_fio.value = f"{comp.competitor.surname_comp} {comp.competitor.name_comp} {comp.competitor.patronymic_comp}"
                        table_cell_fio.font = font_table_text
                        table_cell_fio.border = border
                        table_cell_birthday = ws[f'C{start_counter}']
                        table_cell_birthday.number_format = 'DD.MM.YYYY'
                        table_cell_birthday.value = comp.competitor.birthday
                        table_cell_birthday.font = font_table_text
                        table_cell_birthday.border = border
                        table_cell_sport_club = ws[f'D{start_counter}']
                        table_cell_sport_club.value = "ProSport"
                        comp_sport_club = comp.competitor.club
                        if comp_sport_club:
                            table_cell_sport_club.value = comp_sport_club
                        table_cell_sport_club.font = font_table_text
                        table_cell_sport_club.border = border
                        table_cell_weight = ws[f'E{start_counter}']
                        table_cell_weight.value = comp.competitor_weight
                        table_cell_weight.font = font_table_text
                        table_cell_weight.border = border
                        table_cell_bpress = ws[f'F{start_counter}']
                        table_cell_bpress.value = comp.best_bpress_res
                        table_cell_bpress.font = font_table_text
                        table_cell_bpress.border = border
                        start_counter += 1
                ws.merge_cells(f'A{start_counter}:F{start_counter}')
                ws[f'A{start_counter}'] = ''
                start_counter += 1
        dims = {}
        dims1 = {}
        dims2 = {}
        for row in ws.rows:
            for cell in row:
                if cell.coordinate[0] == 'B' and cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
                if cell.coordinate[0] == 'C' and cell.value:
                    dims1[cell.column] = max((dims1.get(cell.column, 0), len(str(cell.value))))
                if cell.coordinate[0] == 'D' and cell.value:
                    dims2[cell.column] = max((dims2.get(cell.column, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions['B'].width = value + 1
        for col, value in dims1.items():
            ws.column_dimensions['C'].width = value + 2
        for col, value in dims2.items():
            ws.column_dimensions['D'].width = value + 2
        return "Done"
    # def create_sheet_bp_cl(wbook, name_disc, name_disc_full, wcats, competitors_wcats, font_title, font_sub_title, font_title_table, font_table_text, alignment_center, border, cur_competition_title, competitors_pl):
    #     wbook.create_sheet(name_disc)
    #     ws = wbook[name_disc]
    #     ws.insert_rows(1)
    #     ws.merge_cells('A2:E2')
    #     title_sheet = ws['A2']
    #     title_sheet.value = cur_competition_title
    #     title_sheet.font = font_title
    #     title_sheet.alignment = alignment_center
    #     ws.merge_cells('A3:E3')
    #     sub_title_sheet = ws['A3']
    #     sub_title_sheet.value = name_disc_full
    #     sub_title_sheet.font = font_sub_title
    #     sub_title_sheet.alignment = alignment_center
    #     ws.insert_rows(4)
    #     start_counter = 5
    #     for wc in wcats:
    #         if wc in competitors_wcats:
    #             ws.merge_cells(f'A{start_counter}:E{start_counter}')
    #             title_table = ws[f'A{start_counter}']
    #             title_table.value = f'Весовая категория {wc} кг'
    #             title_table.font = font_sub_title
    #             title_table.alignment = alignment_center
    #             start_counter += 1
    #             table_title_place = ws[f'A{start_counter}']
    #             table_title_place.value = 'место №'
    #             table_title_place.font = font_title_table
    #             table_title_place.border = border
    #             table_title_place.alignment = alignment_center
    #             table_title_fio = ws[f'B{start_counter}']
    #             table_title_fio.value = 'фио'
    #             table_title_fio.font = font_title_table
    #             table_title_fio.border = border
    #             table_title_fio.alignment = alignment_center
    #             table_title_birthday = ws[f'C{start_counter}']
    #             table_title_birthday.value = 'дата рождения'
    #             table_title_birthday.font = font_title_table
    #             table_title_birthday.border = border
    #             table_title_birthday.alignment = alignment_center
    #             table_title_weight = ws[f'D{start_counter}']
    #             table_title_weight.value = 'вес'
    #             table_title_weight.font = font_title_table
    #             table_title_weight.border = border
    #             table_title_weight.alignment = alignment_center
    #             table_title_bpress = ws[f'E{start_counter}']
    #             table_title_bpress.value = 'жим'
    #             table_title_bpress.font = font_title_table
    #             table_title_bpress.border = border
    #             table_title_bpress.alignment = alignment_center
    #             start_counter += 1
    #             start_place_counter = 0
    #             for comp in competitors_pl:
    #                 if comp.competitor.weight_cat == wc:
    #                     start_place_counter += 1
    #                     table_cell_place = ws[f'A{start_counter}']
    #                     table_cell_place.value = start_place_counter
    #                     table_cell_place.font = font_table_text
    #                     table_cell_place.border = border
    #                     table_cell_fio = ws[f'B{start_counter}']
    #                     table_cell_fio.value = f"{comp.competitor.surname_comp} {comp.competitor.name_comp} {comp.competitor.patronymic_comp}"
    #                     table_cell_fio.font = font_table_text
    #                     table_cell_fio.border = border
    #                     table_cell_birthday = ws[f'C{start_counter}']
    #                     table_cell_birthday.number_format = 'DD.MM.YYYY'
    #                     table_cell_birthday.value = comp.competitor.birthday
    #                     table_cell_birthday.font = font_table_text
    #                     table_cell_birthday.border = border
    #                     table_cell_weight = ws[f'D{start_counter}']
    #                     table_cell_weight.value = comp.competitor_weight
    #                     table_cell_weight.font = font_table_text
    #                     table_cell_weight.border = border
    #                     table_cell_bpress = ws[f'E{start_counter}']
    #                     table_cell_bpress.value = comp.best_bpress_res
    #                     table_cell_bpress.font = font_table_text
    #                     table_cell_bpress.border = border
    #                     start_counter += 1
    #             ws.merge_cells(f'A{start_counter}:E{start_counter}')
    #             ws[f'A{start_counter}'] = ''
    #             start_counter += 1
    #     dims = {}
    #     dims1 = {}
    #     for row in ws.rows:
    #         for cell in row:
    #             if cell.coordinate[0] == 'B' and cell.value:
    #                 dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
    #             if cell.coordinate[0] == 'C' and cell.value:
    #                 dims1[cell.column] = max((dims1.get(cell.column, 0), len(str(cell.value))))
    #     for col, value in dims.items():
    #         ws.column_dimensions['B'].width = value + 1
    #     for col, value in dims1.items():
    #         ws.column_dimensions['C'].width = value + 2
    #     return "Done"
    if competitiors_pwlclass_m:
        create_sheet_pl_cl(wbook=wb, name_disc='пауэрлифтинг класс муж', name_disc_full='Пауэрлифтинг классический (муж)', wcats=['53', '59', '66', '74', '83', '93', '105', '120', '120+'],
                     competitors_wcats=competitiors_pwlclass_m_wcat, font_title=font_title, font_sub_title=font_sub_title, font_title_table=font_title_table, font_table_text=font_table_text,
                     alignment_center=alignment_center, border=border, cur_competition_title=cur_competition.title, competitors_pl=competitiors_pwlclass_m)
    if competitiors_pwlclass_f:
        create_sheet_pl_cl(wbook=wb, name_disc='пауэрлифтинг класс жен', name_disc_full='Пауэрлифтинг классический (жен)',
                     wcats=['43', '47', '52', '57', '63', '69', '76', '84', '84+'],
                     competitors_wcats=competitiors_pwlclass_f_wcat, font_title=font_title,
                     font_sub_title=font_sub_title, font_title_table=font_title_table, font_table_text=font_table_text,
                     alignment_center=alignment_center, border=border, cur_competition_title=cur_competition.title,
                     competitors_pl=competitiors_pwlclass_f)
    if competitiors_pwlekip_m:
        create_sheet_pl(wbook=wb, name_disc='пауэрлифтинг муж', name_disc_full='Пауэрлифтинг (муж)',
                        wcats=['53', '59', '66', '74', '83', '93', '105', '120', '120+'],
                        competitors_wcats=competitiors_pwlekip_m_wcat, font_title=font_title,
                        font_sub_title=font_sub_title, font_title_table=font_title_table,
                        font_table_text=font_table_text,
                        alignment_center=alignment_center, border=border, cur_competition_title=cur_competition.title,
                        competitors_pl=competitiors_pwlekip_m)
    if competitiors_pwlekip_f:
        create_sheet_pl(wbook=wb, name_disc='пауэрлифтинг жен', name_disc_full='Пауэрлифтинг (жен)',
                        wcats=['43', '47', '52', '57', '63', '69', '76', '84', '84+'],
                        competitors_wcats=competitiors_pwlekip_f_wcat, font_title=font_title,
                        font_sub_title=font_sub_title, font_title_table=font_title_table,
                        font_table_text=font_table_text,
                        alignment_center=alignment_center, border=border, cur_competition_title=cur_competition.title,
                        competitors_pl=competitiors_pwlekip_f)
    if competitiors_bpekip_m:
        create_sheet_bp(wbook=wb, name_disc='жим муж', name_disc_full='Жим (муж)',
                        wcats=['53', '59', '66', '74', '83', '93', '105', '120', '120+'],
                        competitors_wcats=competitiors_bpekip_m_wcat, font_title=font_title,
                        font_sub_title=font_sub_title, font_title_table=font_title_table,
                        font_table_text=font_table_text,
                        alignment_center=alignment_center, border=border, cur_competition_title=cur_competition.title,
                        competitors_pl=competitiors_bpekip_m)
    if competitiors_bpekip_f:
        create_sheet_bp(wbook=wb, name_disc='жим жен', name_disc_full='Жим (жен)',
                        wcats=['43', '47', '52', '57', '63', '69', '76', '84', '84+'],
                        competitors_wcats=competitiors_bpekip_f_wcat, font_title=font_title,
                        font_sub_title=font_sub_title, font_title_table=font_title_table,
                        font_table_text=font_table_text,
                        alignment_center=alignment_center, border=border, cur_competition_title=cur_competition.title,
                        competitors_pl=competitiors_bpekip_f)
    if competitiors_bp_m:
        create_sheet_bp_cl(wbook=wb, name_disc='жим класс муж', name_disc_full='Жим классический (муж)',
                        wcats=['53', '59', '66', '74', '83', '93', '105', '120', '120+'],
                        competitors_wcats=competitiors_bp_m_wcat, font_title=font_title,
                        font_sub_title=font_sub_title, font_title_table=font_title_table,
                        font_table_text=font_table_text,
                        alignment_center=alignment_center, border=border, cur_competition_title=cur_competition.title,
                        competitors_pl=competitiors_bp_m)
    if competitiors_bp_f:
        create_sheet_bp_cl(wbook=wb, name_disc='жим класс жен', name_disc_full='Жим классический (жен)',
                           wcats=['43', '47', '52', '57', '63', '69', '76', '84', '84+'],
                           competitors_wcats=competitiors_bp_f_wcat, font_title=font_title,
                           font_sub_title=font_sub_title, font_title_table=font_title_table,
                           font_table_text=font_table_text,
                           alignment_center=alignment_center, border=border,
                           cur_competition_title=cur_competition.title,
                           competitors_pl=competitiors_bp_f)
    wb.remove_sheet(wb['Sheet'])
    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{cur_competition.competition_slug}_protocol.xlsx"'
    return response


@login_required
def register_participants_page(request, competition_slug):
    cur_competition = Competitions.objects.get(competition_slug=competition_slug)
    cur_participants = Competitors.objects.filter(comp_title__competition_slug=competition_slug).values('id', 'surname_comp', 'name_comp', 'patronymic_comp', 'birthday')
    cur_participants_cnt = cur_participants.count()
    cur_compprotocols_cnt = CompetitionProtocols.objects.filter(competitor__comp_title__competition_slug=competition_slug).count()
    cur_flow = CompetitorsFlow.objects.first()
    response_data = {
        'cur_competition': cur_competition,
        'competition_slug': competition_slug,
        'cur_participants': cur_participants,
        'cur_participants_cnt': cur_participants_cnt,
        'cur_compprotocols_cnt': cur_compprotocols_cnt,
        'cur_flow': cur_flow,
    }
    if 'search' in request.POST:
        search_val = request.POST.get('searchval')
        search_res = ""
        if search_val:
            search_res = Competitors.objects.filter(surname_comp__icontains=search_val.strip(), comp_title__competition_slug=competition_slug).values('id', 'surname_comp', 'name_comp', 'patronymic_comp', 'birthday')
            cur_participants_cnt = search_res.count()
            cur_compprotocols_cnt = CompetitionProtocols.objects.filter(competitor__comp_title__competition_slug=competition_slug, competitor__surname_comp__icontains=search_val.strip()).count()
        return JsonResponse({"answer": list(search_res), 'cur_participants_cnt': cur_participants_cnt, 'cur_compprotocols_cnt': cur_compprotocols_cnt}, status=200)
    if 'searchstop' in request.POST:
        return JsonResponse({"answer": list(cur_participants), 'cur_participants_cnt': cur_participants_cnt, 'cur_compprotocols_cnt': cur_compprotocols_cnt}, status=200)
    if 'filtershow' in request.POST:
        targ_gen = request.POST.get('gen')
        targ_wcat = request.POST.get('wcat')
        if targ_gen == 'male':
            if targ_wcat == 'change_cat':
                cur_participants = Competitors.objects.filter(gender='Мужской', comp_title__competition_slug=competition_slug).values('id', 'surname_comp', 'name_comp', 'patronymic_comp', 'birthday')
                cur_participants_cnt = cur_participants.count()
                cur_compprotocols_cnt = CompetitionProtocols.objects.filter(competitor__comp_title__competition_slug=competition_slug, competitor__gender='Мужской').count()
                return JsonResponse({"answer": list(cur_participants), 'cur_participants_cnt': cur_participants_cnt, 'cur_compprotocols_cnt': cur_compprotocols_cnt}, status=200)
            else:
                cur_participants = Competitors.objects.filter(gender='Мужской', comp_title__competition_slug=competition_slug, weight_cat=targ_wcat).values('id', 'surname_comp', 'name_comp', 'patronymic_comp', 'birthday')
                cur_participants_cnt = cur_participants.count()
                cur_compprotocols_cnt = CompetitionProtocols.objects.filter(competitor__comp_title__competition_slug=competition_slug, competitor__gender='Мужской', competitor__weight_cat=targ_wcat).count()
                return JsonResponse({"answer": list(cur_participants), 'cur_participants_cnt': cur_participants_cnt, 'cur_compprotocols_cnt': cur_compprotocols_cnt}, status=200)
        if targ_gen == 'female':
            if targ_wcat == 'change_cat':
                cur_participants = Competitors.objects.filter(gender='Женский', comp_title__competition_slug=competition_slug).values('id', 'surname_comp', 'name_comp', 'patronymic_comp', 'birthday')
                cur_participants_cnt = cur_participants.count()
                cur_compprotocols_cnt = CompetitionProtocols.objects.filter(competitor__comp_title__competition_slug=competition_slug, competitor__gender='Женский').count()
                return JsonResponse({"answer": list(cur_participants), 'cur_participants_cnt': cur_participants_cnt, 'cur_compprotocols_cnt': cur_compprotocols_cnt}, status=200)
            else:
                cur_participants = Competitors.objects.filter(gender='Женский', comp_title__competition_slug=competition_slug, weight_cat=targ_wcat).values('id', 'surname_comp', 'name_comp', 'patronymic_comp', 'birthday')
                cur_participants_cnt = cur_participants.count()
                cur_compprotocols_cnt = CompetitionProtocols.objects.filter(competitor__comp_title__competition_slug=competition_slug, competitor__gender='Женский', competitor__weight_cat=targ_wcat).count()
                return JsonResponse({"answer": list(cur_participants), 'cur_participants_cnt': cur_participants_cnt, 'cur_compprotocols_cnt': cur_compprotocols_cnt}, status=200)
    if 'setflow' in request.POST:
        new_flow = request.POST.get('setflow')
        cur_flow.comp_flow = new_flow
        cur_flow.save()
    return render(request, 'pwrlftmain/register_participants.html', response_data)


@login_required
def register_participant_page(request, competition_slug, pk):
    cur_competition = Competitions.objects.get(competition_slug=competition_slug)
    competitor = Competitors.objects.get(pk=pk)
    if CompetitionProtocols.objects.filter(competitor=competitor).exists():
        cur_comp_protocol = CompetitionProtocols.objects.get(competitor=competitor)
    else:
        cur_comp_protocol = None
    response_data = {
        'cur_competition': cur_competition,
        'competitor': competitor,
        'cur_comp_protocol': cur_comp_protocol,
        'competition_slug': competition_slug,
    }
    if 'competition_registry-form_btn' in request.POST:
        gender_type = request.POST.get('competition_entry-form_gender')
        if gender_type == 'competition_registry-form_gender-male':
            competitor.gender = 'Мужской'
            competitor.weight_cat = request.POST.get('competition_registry-form_wcat-m')
        else:
            competitor.gender = 'Женский'
            competitor.weight_cat = request.POST.get('competition_registry-form_wcat-f')
        comp_surname = request.POST.get('competition_registry-form_sur-input')
        comp_name = request.POST.get('competition_registry-form_name-input')
        comp_patr = request.POST.get('competition_registry-form_patr-input')
        comp_birth = request.POST.get('competition_registry-form_birthday-input')
        comp_scat = request.POST.get('competition_registry-form_sp-cat')
        comp_stypes = request.POST.getlist('competition_registry-form_sp-t_chckbx')
        comp_bpek = 0
        comp_bpcl = 0
        comp_bbpek = 0
        comp_bbp = 0
        comp_bpek_st = request.POST.get('competition_registry-form_pwlekip-input')
        if comp_bpek_st != '0.00':
            comp_bpek = comp_bpek_st
        comp_bpcl_st = request.POST.get('competition_registry-form_pwlclass-input')
        if comp_bpcl_st != '0.00':
            comp_bpcl = comp_bpcl_st
        comp_bbpek_st = request.POST.get('competition_registry-form_bpekip-input')
        if comp_bbpek_st != '0.00':
            comp_bbpek = comp_bbpek_st
        comp_bbp_st = request.POST.get('competition_registry-form_bp-input')
        if comp_bbp_st != '0.00':
            comp_bbp = comp_bbp_st
        comp_surt = request.POST.get('competition_registry-form_sur-trn-input')
        comp_namet = request.POST.get('competition_registry-form_name-trn-input')
        comp_patrt = request.POST.get('competition_registry-form_patr-trn-input')
        comp_sclub = request.POST.get('competition_registry-form_club-input')
        comp_phone = request.POST.get('competition_registry-form_phone-input')
        comp_mail = request.POST.get('competition_registry-form_mail-input')
        comp_weight = request.POST.get('competition_registry-form_w-weight')
        if not comp_weight:
            comp_weight = 0
        comp_fbp = request.POST.get('competition_registry-form_bpress')
        if not comp_fbp:
            comp_fbp = 0
        if 'Троеборье классическое' in comp_stypes:
            comp_fsq = request.POST.get('competition_registry-form_squat')
            comp_fdl = request.POST.get('competition_registry-form_dlift')
            if not comp_fsq:
                comp_fsq = 0
            if not comp_fdl:
                comp_fdl = 0
        if 'Троеборье (экип.)' in comp_stypes:
            comp_fsq = request.POST.get('competition_registry-form_squat')
            comp_fdl = request.POST.get('competition_registry-form_dlift')
            if not comp_fsq:
                comp_fsq = 0
            if not comp_fdl:
                comp_fdl = 0
        comp_flow = None
        comp_flow_st = request.POST.get('competition_registry-form_flow')
        if comp_flow_st != 'Поток':
            comp_flow = comp_flow_st
        competitor.surname_comp = comp_surname
        competitor.name_comp = comp_name
        competitor.patronymic_comp = comp_patr
        competitor.birthday = comp_birth
        competitor.sports_cat = comp_scat
        competitor.best_res_pwlft_ek = comp_bpek
        competitor.best_res_pwlft_cl = comp_bpcl
        competitor.best_res_bpress_ek = comp_bbpek
        competitor.best_res_bpress_cl = comp_bbp
        competitor.surname_trainer = comp_surt
        competitor.name_trainer = comp_namet
        competitor.patronymic_trainer = comp_patrt
        competitor.club = comp_sclub
        competitor.phone = comp_phone
        competitor.mail = comp_mail
        competitor.save()
        competitor.sports_type.clear()
        for c_s_t in comp_stypes:
            sport_type_id = SportType.objects.get(title=c_s_t).id
            competitor.sports_type.add(sport_type_id)
        if CompetitionProtocols.objects.filter(competitor=competitor).exists():
            cur_protocol = CompetitionProtocols.objects.get(competitor=competitor)
        else:
            cur_protocol = CompetitionProtocols()
        cur_protocol.competitor_weight = comp_weight
        cur_protocol.competitor_stream = comp_flow
        if 'Троеборье классическое' in comp_stypes:
            cur_protocol.first_attempt_squat_res = comp_fsq
            cur_protocol.first_attempt_bpress_res = comp_fbp
            cur_protocol.first_attempt_dlift_res = comp_fdl
        elif 'Жим лёжа (без экип.)' in comp_stypes:
            cur_protocol.first_attempt_bpress_res = comp_fbp
        if 'Троеборье (экип.)' in comp_stypes:
            cur_protocol.first_attempt_squat_res_ek = comp_fsq
            cur_protocol.first_attempt_bpress_res_ek = comp_fbp
            cur_protocol.first_attempt_dlift_off_ek = comp_fdl
        elif 'Жим лёжа (экип.)' in comp_stypes:
            cur_protocol.first_attempt_bpress_res_ek = comp_fbp
        cur_protocol.competitor = competitor
        cur_protocol.save()
        return redirect('register_participants', competition_slug=competition_slug)
    if 'delpart' in request.POST:
        if CompetitionProtocols.objects.filter(competitor=competitor).exists():
            cur_protocol = CompetitionProtocols.objects.get(competitor=competitor)
            cur_protocol.delete()
        competitor.delete()
        return JsonResponse({"answer": "ok"}, status=200)
    return render(request, 'pwrlftmain/register_participant.html', response_data)


@login_required
def register_new_participant_page(request, competition_slug):
    cur_competition = Competitions.objects.get(competition_slug=competition_slug)
    response_data = {
        'cur_competition': cur_competition,
    }
    if 'competition_registry-form_btn' in request.POST:
        competitor = Competitors()
        gender_type = request.POST.get('competition_entry-form_gender')
        if gender_type == 'competition_registry-form_gender-male':
            competitor.gender = 'Мужской'
            competitor.weight_cat = request.POST.get('competition_registry-form_wcat-m')
        else:
            competitor.gender = 'Женский'
            competitor.weight_cat = request.POST.get('competition_registry-form_wcat-f')
        comp_surname = request.POST.get('competition_registry-form_sur-input')
        comp_name = request.POST.get('competition_registry-form_name-input')
        comp_patr = request.POST.get('competition_registry-form_patr-input')
        comp_birth = request.POST.get('competition_registry-form_birthday-input')
        comp_scat = request.POST.get('competition_registry-form_sp-cat')
        comp_stypes = request.POST.getlist('competition_registry-form_sp-t_chckbx')
        comp_bpek = 0
        comp_bpcl = 0
        comp_bbpek = 0
        comp_bbp = 0
        comp_bpek_st = request.POST.get('competition_registry-form_pwlekip-input')
        if comp_bpek_st != '0.00':
            comp_bpek = comp_bpek_st
        comp_bpcl_st = request.POST.get('competition_registry-form_pwlclass-input')
        if comp_bpcl_st != '0.00':
            comp_bpcl = comp_bpcl_st
        comp_bbpek_st = request.POST.get('competition_registry-form_bpekip-input')
        if comp_bbpek_st != '0.00':
            comp_bbpek = comp_bbpek_st
        comp_bbp_st = request.POST.get('competition_registry-form_bp-input')
        if comp_bbp_st != '0.00':
            comp_bbp = comp_bbp_st
        comp_surt = request.POST.get('competition_registry-form_sur-trn-input')
        comp_namet = request.POST.get('competition_registry-form_name-trn-input')
        comp_patrt = request.POST.get('competition_registry-form_patr-trn-input')
        comp_sclub = request.POST.get('competition_registry-form_club-input')
        comp_phone = request.POST.get('competition_registry-form_phone-input')
        comp_mail = request.POST.get('competition_registry-form_mail-input')
        comp_weight = request.POST.get('competition_registry-form_w-weight')
        if not comp_weight:
            comp_weight = 0
        comp_fbp = request.POST.get('competition_registry-form_bpress')
        if not comp_fbp:
            comp_fbp = 0
        if 'Троеборье классическое' in comp_stypes:
            comp_fsq = request.POST.get('competition_registry-form_squat')
            comp_fdl = request.POST.get('competition_registry-form_dlift')
            if not comp_fsq:
                comp_fsq = 0
            if not comp_fdl:
                comp_fdl = 0
        if 'Троеборье (экип.)' in comp_stypes:
            comp_fsq = request.POST.get('competition_registry-form_squat')
            comp_fdl = request.POST.get('competition_registry-form_dlift')
            if not comp_fsq:
                comp_fsq = 0
            if not comp_fdl:
                comp_fdl = 0
        comp_flow = None
        comp_flow_st = request.POST.get('competition_registry-form_flow')
        if comp_flow_st != 'Поток':
            comp_flow = comp_flow_st
        competitor.surname_comp = comp_surname
        competitor.name_comp = comp_name
        competitor.patronymic_comp = comp_patr
        competitor.birthday = comp_birth
        competitor.sports_cat = comp_scat
        competitor.best_res_pwlft_ek = comp_bpek
        competitor.best_res_pwlft_cl = comp_bpcl
        competitor.best_res_bpress_ek = comp_bbpek
        competitor.best_res_bpress_cl = comp_bbp
        competitor.surname_trainer = comp_surt
        competitor.name_trainer = comp_namet
        competitor.patronymic_trainer = comp_patrt
        competitor.club = comp_sclub
        competitor.phone = comp_phone
        competitor.mail = comp_mail
        competitor.comp_title = cur_competition
        competitor.save()
        for c_s_t in comp_stypes:
            sport_type_id = SportType.objects.get(title=c_s_t).id
            competitor.sports_type.add(sport_type_id)
        cur_protocol = CompetitionProtocols()
        cur_protocol.competitor_weight = comp_weight
        cur_protocol.competitor_stream = comp_flow
        if 'Троеборье классическое' in comp_stypes:
            cur_protocol.first_attempt_squat_res = comp_fsq
            cur_protocol.first_attempt_bpress_res = comp_fbp
            cur_protocol.first_attempt_dlift_res = comp_fdl
        elif 'Жим лёжа (без экип.)' in comp_stypes:
            cur_protocol.first_attempt_bpress_res = comp_fbp
        if 'Троеборье (экип.)' in comp_stypes:
            cur_protocol.first_attempt_squat_res_ek = comp_fsq
            cur_protocol.first_attempt_bpress_res_ek = comp_fbp
            cur_protocol.first_attempt_dlift_off_ek = comp_fdl
        elif 'Жим лёжа (экип.)' in comp_stypes:
            cur_protocol.first_attempt_bpress_res_ek = comp_fbp
        cur_protocol.competitor = competitor
        cur_protocol.save()
        return redirect('register_participants', competition_slug=competition_slug)
    return render(request, 'pwrlftmain/register_participant_new.html', response_data)
